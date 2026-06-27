#!/usr/bin/env python3
"""
SOV -> Hotels & Motels Supplemental Application (signable PDF)
=============================================================

Reads a HUB hospitality SOV (Excel) and produces a pre-filled, plain-format
"Hotels & Motels" supplemental application as a flat PDF with a signature line
for the insured to sign. Mirrors the layout of the carrier sample form.

- Single location   -> one application, premises/pool/etc. filled from the row.
- Multiple locations -> business info + signature appear once; the per-location
  schedule (Premises, Pools, Recreation, Cooking, Liquor, Bar, Entertainment,
  Security, Gross Receipts) repeats once per location.

Field placement is driven by HEADER text in row 2 of the SOV, so column order
does not matter. Missing columns simply leave the answer blank.

Usage:
    python3 sov_to_application.py "<SOV.xlsx>" "<output.pdf>"
"""

import sys
import datetime
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether,
    HRFlowable,
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os

# ----------------------------------------------- checkbox font (portable)
# Use a TTF with ballot-box glyphs if one is available; otherwise fall back
# to ASCII brackets so the script renders correctly on any machine.
_BOX_FONT = None
for _p in (
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/local/lib/python3.10/dist-packages/matplotlib/mpl-data/fonts/ttf/DejaVuSans.ttf",
    "/Library/Fonts/Arial Unicode.ttf",
    "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
    "/Library/Fonts/DejaVuSans.ttf",
):
    if os.path.exists(_p):
        try:
            pdfmetrics.registerFont(TTFont("BoxFont", _p))
            _BOX_FONT = "BoxFont"
            break
        except Exception:
            pass


def _box(checked):
    """Return paragraph markup for a checkbox, checked or empty."""
    if _BOX_FONT:
        glyph = "☒" if checked else "☐"  # ballot box (with X) / empty
        return f'<font name="{_BOX_FONT}">{glyph}</font>'
    return "[X]" if checked else "[&nbsp;&nbsp;]"

# ---------------------------------------------------------------- HUB brand colors
HUB_CLASSIC_BLUE = colors.HexColor("#263746")
HUB_ELECTRIC_BLUE = colors.HexColor("#0678d5")
HUB_ARCTIC_GRAY = colors.HexColor("#e3e3e3")
HUB_CLASSIC_LIGHT = colors.HexColor("#385263")
HUB_GOLD = colors.HexColor("#f3b921")

APP_TITLE = "Hotel Supplemental Application"
APP_FOOTER = "Hotel Supplemental Application v2026Q2"

# Canonical HUB hotel-program quote subjectivities. Each entry is
# (display text, lowercase substring used to detect it in the SOV string).
# Marked True when the substring is present in the SOV "Quote Subjectivities"
# field; left blank (unchecked) when missing.
SUBJECTIVITIES = [
    ("Over 3 years of operational experience", "3 year"),
    ("No prior losses for Abuse & Molestation, Assault & Battery, and Human Trafficking", "no prior losses"),
    ("Pest control service contract including bed bug prevention / detection", "pest control"),
    ("Background checks on all employees", "background check"),
    ("No homeless shelters", "homeless shelter"),
    ("No hourly rentals", "hourly rental"),
    ("No long term rentals (greater than 30 days)", "long term rental"),
    ("No liability losses over $100,000", "liability losses over"),
    ("Human Trafficking Awareness Program (annual training all employees)", "human trafficking awareness"),
    ("Manager performs walk-throughs at least quarterly to identify and fix hazards", "walk through"),
    ("Written contracts with all third-party contractors/providers with hold harmless and indemnification favorable to the insured", "hold harmless"),
]

# ---------------------------------------------------------------- styles
BODY = ParagraphStyle("body", fontName="Helvetica", fontSize=8.5, leading=11)
BOLD = ParagraphStyle("bold", fontName="Helvetica-Bold", fontSize=8.5, leading=11)
SECTION = ParagraphStyle("section", fontName="Helvetica-Bold", fontSize=10,
                         leading=13, textColor=colors.white)
TITLE = ParagraphStyle("title", fontName="Helvetica-Bold", fontSize=15,
                       leading=18, alignment=1, textColor=colors.white)
LOCHEAD = ParagraphStyle("lochead", fontName="Helvetica-Bold", fontSize=11,
                         leading=14, textColor=HUB_CLASSIC_BLUE)
SMALL = ParagraphStyle("small", fontName="Helvetica", fontSize=7.5, leading=10)
FINE = ParagraphStyle("fine", fontName="Helvetica", fontSize=7.2, leading=9.5)

CONTENT_W = 7.0 * inch  # printable width inside 0.75" margins


# ---------------------------------------------------------------- helpers
def P(text, style=BODY):
    return Paragraph("" if text is None else str(text), style)


def yn(val):
    """Render a Yes/No answer with the applicable box checked.
    Returns a Paragraph. Blank value -> both boxes empty."""
    s = ("" if val is None else str(val)).strip().lower()
    yes = s in ("yes", "y", "true", "1", "1.0")
    no = s in ("no", "n", "false", "0", "0.0", "n/a", "na", "none")
    return Paragraph(f"{_box(yes)} Yes&nbsp;&nbsp;&nbsp;{_box(no)} No", BODY)


def checkbox(checked, label=""):
    return Paragraph(f"{_box(checked)} {label}", BODY)


def money(v):
    try:
        return "$" + format(float(v), ",.0f")
    except (TypeError, ValueError):
        return "" if v in (None, "") else str(v)


def pct(v):
    if v in (None, ""):
        return ""
    try:
        f = float(v)
        if f <= 1:
            f *= 100
        return f"{f:.0f}%"
    except (TypeError, ValueError):
        return str(v)


def year_age(v):
    """Render an update year with its computed age, e.g. '2025 - 1 yr'."""
    if v in (None, ""):
        return ""
    try:
        yr = int(float(str(v).replace(",", "").strip()))
    except (TypeError, ValueError):
        return str(v)
    age = datetime.date.today().year - yr
    if age < 0:
        return str(yr)
    unit = "yr" if age == 1 else "yrs"
    return f"{yr} - {age} {unit}"


def roof_yr_age(v):
    """Compact roof replacement year + age, e.g. '1997 (29 yrs)'."""
    if v in (None, ""):
        return ""
    try:
        yr = int(float(str(v).replace(",", "").strip()))
    except (TypeError, ValueError):
        return str(v)
    age = datetime.date.today().year - yr
    if age < 0:
        return str(yr)
    unit = "yr" if age == 1 else "yrs"
    return f"{yr} ({age} {unit})"


def txt(v):
    if v in (None, ""):
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, datetime.datetime):
        return v.strftime("%m/%d/%Y")
    return str(v)


# ---------------------------------------------------------------- SOV read
def read_sov(path):
    """Return (applicant_dict, [location_dicts]). Each location aggregates the
    SOV rows that share a Loc # (multi-building locations are summed/first)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # auto-detect the header row (some SOVs have a title banner in row 1,
    # others put the field names directly in row 1). Pick the row in the
    # first few that best matches known SOV column names.
    known = {"Corporate Name (LLC)", "DBA", "Full Address", "TIV", "Loc #",
             "# of Rooms", "Construction", "Effective Date", "Building #"}
    header_row, best = 1, -1
    for r in range(1, min(ws.max_row, 6) + 1):
        score = 0
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() in known:
                score += 1
        if score > best:
            best, header_row = score, r

    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=header_row, column=c).value
        if h is not None:
            headers[str(h).strip()] = c

    def cell(row, name):
        col = headers.get(name)
        return ws.cell(row=row, column=col).value if col else None

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        # skip fully empty rows
        if all(ws.cell(row=r, column=c).value in (None, "")
               for c in range(1, ws.max_column + 1)):
            continue
        rows.append(r)

    # group rows by Loc # (fall back to one group if no Loc #)
    groups = {}
    order = []
    for r in rows:
        loc = cell(r, "Loc #")
        key = txt(loc) or f"_{r}"
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(r)

    def g(rowlist, name):
        for r in rowlist:
            v = cell(r, name)
            if v not in (None, ""):
                return v
        return None

    def _num(v):
        """Parse a numeric value that may be a currency/percent string."""
        if v is None or v == "":
            return None
        try:
            return float(str(v).replace("$", "").replace(",", "").replace("%", "").strip())
        except ValueError:
            return None

    def gsum(rowlist, name):
        tot, found = 0, False
        for r in rowlist:
            n = _num(cell(r, name))
            if n is not None:
                tot += n; found = True
        return tot if found else None

    locations = []
    for key in order:
        rl = groups[key]
        loc = {
            "loc_no": g(rl, "Loc #"),
            "corp": g(rl, "Corporate Name (LLC)"),
            "dba": g(rl, "DBA"),
            "address": g(rl, "Full Address"),
            "rooms": gsum(rl, "# of Rooms"),
            "occupancy": g(rl, "Occupancy Ratio"),
            "adr": g(rl, "ADR"),
            "room_rentals": g(rl, "Room Rentals"),
            "num_buildings": len(rl),
            "yr_elec": g(rl, "Year Electrical Updated"),
            "yr_plumb": g(rl, "Year Plumbing Updated"),
            "yr_roof": g(rl, "Year Roof Fully Replaced"),
            "yr_hvac": g(rl, "Year HVAC Updated"),
            "sprinkler": g(rl, "% Sprinklered"),
            "smoke": g(rl, "Smoke Alarm"),
            "fire": g(rl, "Fire Alarm"),
            "wiring": g(rl, "Wiring Type"),
            "floors": g(rl, "# of Floors"),
            "corridor": g(rl, "Corridor"),
            "construction": g(rl, "Construction"),
            "eifs": g(rl, "Exterior Insulating Finishing System (EIFS)"),
            "yr_built": g(rl, "Yr Built"),
            "sqft": gsum(rl, "SqFt"),
            "bldg_limit": gsum(rl, "Building Limit"),
            "contents_limit": gsum(rl, "Contents Limit"),
            "bi_limit": gsum(rl, "Business Income Limit"),
            "tiv": gsum(rl, "TIV"),
            # pools
            "num_pools": gsum(rl, "# of Pools"),
            "pool_amen": g(rl, "Pool Amenities"),
            # recreation / fitness
            "fitness": g(rl, "Fitness Room"),
            # cooking / restaurant
            "restaurant": g(rl, "Restaurant"),
            "nfpa": g(rl, "NFPA 90 or UL 300 Equivalent Standards (Cooking)"),
            "hours_open": g(rl, "Hours - Open"),
            "hours_closed": g(rl, "Hours - Closed"),
            # liquor / bar / entertainment
            "liquor": g(rl, "Liquor Liability"),
            "happy_hours": g(rl, "Happy Hours?"),
            "live_ent": g(rl, "Live Entertainment"),
            "dance_floor": g(rl, "Dance Floor"),
            "bouncer": g(rl, "Bouncer"),
            # receipts
            "hotel_sales": gsum(rl, "Hotel Sales"),
            "restaurant_sales": gsum(rl, "Restaurant Sales"),
            "liquor_sales": gsum(rl, "Liquor Sales"),
            "sundry_sales": gsum(rl, "Sundry Sales (Marketplace)"),
            "other_sales": gsum(rl, "Other Sales"),
            "total_sales": gsum(rl, "Total Sales"),
            # employees (per location)
            "ft": gsum(rl, "# Full Time Employees"),
            "pt": gsum(rl, "# Part Time Employees"),
        }
        locations.append(loc)

    # sort locations by numeric Loc # (fall back to string order)
    def _locsort(l):
        try:
            return (0, int(float(str(l["loc_no"]).strip())))
        except (TypeError, ValueError):
            return (1, str(l["loc_no"]))
    locations.sort(key=_locsort)

    first = rows[0] if rows else header_row + 1
    applicant = {
        "corp": cell(first, "Corporate Name (LLC)"),
        "dba": cell(first, "DBA"),
        "address": cell(first, "Full Address"),
        "eff_date": cell(first, "Effective Date"),
        "fein": cell(first, "FEIN"),
        "contact": cell(first, "Application Contact Name"),
        "phone": cell(first, "Application - Phone"),
        "subjectivities": cell(first, "Quote Subjectivities"),
        "ft": cell(first, "# Full Time Employees"),
        "pt": cell(first, "# Part Time Employees"),
        # HNOA
        "hnoa": cell(first, "HNOA Controls"),
        "owned_autos": cell(first, "# of Owned Autos"),
        "guest_transport": cell(first, "Guest Transportation"),
        "valet": cell(first, "Valet Parking"),
    }
    return applicant, locations


# ---------------------------------------------------------------- layout bits
def section_bar(title):
    t = Table([[P(title, SECTION)]], colWidths=[CONTENT_W])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), HUB_ELECTRIC_BLUE),
        ("LINEBEFORE", (0, 0), (0, -1), 3, HUB_CLASSIC_BLUE),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    return t


def kv_table(rows, col_widths):
    """rows: list of cell-lists already wrapped as Paragraphs/strings."""
    t = Table(rows, colWidths=col_widths)
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONT", (0, 0), (-1, -1), "Helvetica", 8.5),
        ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#bdbdbd")),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    return t


def lbl(text):
    return Paragraph(text, BOLD)


def ans(text):
    """Filled answer shown with a light underline feel via box border."""
    return Paragraph(f"<u>{'' if text in (None,'') else text}</u>" if text not in (None, "")
                     else "____________________", BODY)


# ---------------------------------------------------------------- builders
def title_bar(text):
    t = Table([[P(text, TITLE)]], colWidths=[CONTENT_W])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), HUB_CLASSIC_BLUE),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    return t


def build_business_info(story, app):
    story.append(title_bar(APP_TITLE))
    story.append(Spacer(1, 6))
    half = CONTENT_W / 2
    rows = [
        [lbl("Applicant Name:"), ans(txt(app["corp"]) + (f"  (DBA {txt(app['dba'])})" if app["dba"] else "")),
         lbl("Date:"), ans(datetime.date.today().strftime("%m/%d/%Y"))],
    ]
    t = Table(rows, colWidths=[1.2 * inch, half + 0.3 * inch, 0.6 * inch, 1.3 * inch])
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBELOW", (1, 0), (1, 0), 0.4, colors.grey),
        ("LINEBELOW", (3, 0), (3, 0), 0.4, colors.grey),
        ("LEFTPADDING", (0, 0), (-1, -1), 2), ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(t)

    story.append(kv_table([
        [lbl("Location / Mailing Address:"), ans(txt(app["address"]))],
        [lbl("FEIN:"), ans(txt(app["fein"]))],
        [lbl("Effective Date:"), ans(txt(app["eff_date"]))],
        [lbl("Contact:"), ans(txt(app["contact"]) + ("  " + txt(app["phone"]) if app["phone"] else ""))],
    ], [2.0 * inch, CONTENT_W - 2.0 * inch]))

    story.append(Spacer(1, 6))
    story.append(P("Business Information:", BOLD))
    exp = ""
    if app["subjectivities"]:
        exp = txt(app["subjectivities"])
    story.append(kv_table([
        [lbl("Years of experience in this industry:"),
         ans("Over 3 years" if exp and "3 year" in exp.lower() else "")],
        [lbl("Number of full-time / part-time employees:"),
         ans(f"{txt(app['ft'])} FT  /  {txt(app['pt'])} PT")],
        [lbl("Are background checks performed on all employees?"),
         yn("Yes" if exp and "background check" in exp.lower() else "")],
        [lbl("Is multi-factor authentication (MFA) required for all employees "
             "accessing email, networks, or company systems remotely (from "
             "outside the hotel)?"), yn("")],
    ], [2.6 * inch, CONTENT_W - 2.6 * inch]))
    story.append(Spacer(1, 6))

    # Quote subjectivities broken out, each marked True (checked) when present
    # in the SOV, blank when missing.
    story.append(P("Quote Subjectivities "
                   "<font size=7>(checked = confirmed on SOV)</font>:", BOLD))
    story.append(Spacer(1, 2))
    src = (exp or "").lower()
    sub_rows = []
    for disp, key in SUBJECTIVITIES:
        present = key in src
        sub_rows.append([checkbox(present, ""), P(disp, SMALL)])
    st = Table(sub_rows, colWidths=[0.3 * inch, CONTENT_W - 0.3 * inch])
    st.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#dddddd")),
        ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5), ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
    ]))
    story.append(st)
    story.append(Spacer(1, 8))


def build_property_schedule(story, locations):
    """Summary schedule table across all locations."""
    story.append(section_bar(f"Property Schedule &mdash; {len(locations)} Locations"))
    story.append(Spacer(1, 3))
    th = ParagraphStyle("psth", fontName="Helvetica-Bold", fontSize=6.8,
                        leading=8.2, textColor=colors.white)
    td = ParagraphStyle("pstd", fontName="Helvetica", fontSize=6.6, leading=8)
    tdb = ParagraphStyle("pstdb", fontName="Helvetica-Bold", fontSize=6.6, leading=8)
    tot = ParagraphStyle("pstot", fontName="Helvetica-Bold", fontSize=6.6,
                         leading=8, textColor=HUB_CLASSIC_BLUE)
    cols = ["Loc", "Property / DBA", "Construction", "Yr<br/>Blt", "Spr<br/>%",
            "Building", "Contents", "Business<br/>Income", "TIV", "Roof Yr<br/>(Age)"]
    data = [[Paragraph(c, th) for c in cols]]
    tb = tc = tbi = tt = 0
    for i, loc in enumerate(locations, start=1):
        tb += loc["bldg_limit"] or 0; tc += loc["contents_limit"] or 0
        tbi += loc["bi_limit"] or 0; tt += loc["tiv"] or 0
        data.append([
            Paragraph(txt(loc["loc_no"]) or str(i), td),
            Paragraph(txt(loc["dba"]), tdb),
            Paragraph(txt(loc["construction"]), td),
            Paragraph(txt(loc["yr_built"]), td),
            Paragraph(pct(loc["sprinkler"]), td),
            Paragraph(money(loc["bldg_limit"]), td),
            Paragraph(money(loc["contents_limit"]), td),
            Paragraph(money(loc["bi_limit"]), td),
            Paragraph(money(loc["tiv"]), tdb),
            Paragraph(roof_yr_age(loc["yr_roof"]), td),
        ])
    data.append([
        Paragraph("", td), Paragraph("PORTFOLIO TOTAL", tot), Paragraph("", td),
        Paragraph("", td), Paragraph("", td), Paragraph(money(tb), tot),
        Paragraph(money(tc), tot), Paragraph(money(tbi), tot),
        Paragraph(money(tt), tot), Paragraph("", td),
    ])
    cw = [0.3, 1.45, 0.92, 0.4, 0.4, 0.76, 0.7, 0.72, 0.8, 0.82]
    scale = CONTENT_W / (sum(cw) * inch)
    cw = [c * inch * scale for c in cw]
    t = Table(data, colWidths=cw, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HUB_CLASSIC_BLUE),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#eef3f7")]),
        ("BACKGROUND", (0, -1), (-1, -1), HUB_ARCTIC_GRAY),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#bfbfbf")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (3, 1), (4, -1), "CENTER"), ("ALIGN", (5, 1), (8, -1), "RIGHT"),
        ("ALIGN", (9, 1), (9, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5), ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(t)
    story.append(Spacer(1, 12))


def build_liability_schedule(story, locations):
    """Exposure/liability schedule: rooms, sales mix, and employee counts."""
    bar = section_bar(f"Liability Schedule &mdash; {len(locations)} Locations")
    th = ParagraphStyle("lsth", fontName="Helvetica-Bold", fontSize=6.8,
                        leading=8.2, textColor=colors.white)
    td = ParagraphStyle("lstd", fontName="Helvetica", fontSize=6.6, leading=8)
    tdb = ParagraphStyle("lstdb", fontName="Helvetica-Bold", fontSize=6.6, leading=8)
    tot = ParagraphStyle("lstot", fontName="Helvetica-Bold", fontSize=6.6,
                         leading=8, textColor=HUB_CLASSIC_BLUE)
    cols = ["Loc", "Property / DBA", "# of<br/>Rooms", "Hotel", "Restaurant",
            "Liquor", "Sundry", "Total<br/>Sales", "Employees<br/>FT / PT"]
    data = [[Paragraph(c, th) for c in cols]]
    trooms = thot = tres = tliq = tsun = ttot = tft = tpt = 0
    for i, loc in enumerate(locations, start=1):
        trooms += loc["rooms"] or 0; thot += loc["hotel_sales"] or 0
        tres += loc["restaurant_sales"] or 0; tliq += loc["liquor_sales"] or 0
        tsun += loc["sundry_sales"] or 0; ttot += loc["total_sales"] or 0
        tft += loc["ft"] or 0; tpt += loc["pt"] or 0
        data.append([
            Paragraph(txt(loc["loc_no"]) or str(i), td),
            Paragraph(txt(loc["dba"]), tdb),
            Paragraph(txt(int(loc["rooms"])) if loc["rooms"] else "", td),
            Paragraph(money(loc["hotel_sales"]), td),
            Paragraph(money(loc["restaurant_sales"]), td),
            Paragraph(money(loc["liquor_sales"]), td),
            Paragraph(money(loc["sundry_sales"]), td),
            Paragraph(money(loc["total_sales"]), tdb),
            Paragraph(f"{int(loc['ft'] or 0)} / {int(loc['pt'] or 0)}", td),
        ])
    data.append([
        Paragraph("", td), Paragraph("PORTFOLIO TOTAL", tot),
        Paragraph(str(int(trooms)), tot), Paragraph(money(thot), tot),
        Paragraph(money(tres), tot), Paragraph(money(tliq), tot),
        Paragraph(money(tsun), tot), Paragraph(money(ttot), tot),
        Paragraph(f"{int(tft)} / {int(tpt)}", tot),
    ])
    cw = [0.3, 1.55, 0.5, 0.82, 0.82, 0.72, 0.72, 0.86, 0.74]
    scale = CONTENT_W / (sum(cw) * inch)
    cw = [c * inch * scale for c in cw]
    t = Table(data, colWidths=cw, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HUB_CLASSIC_BLUE),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#eef3f7")]),
        ("BACKGROUND", (0, -1), (-1, -1), HUB_ARCTIC_GRAY),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#bfbfbf")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (2, 1), (2, -1), "CENTER"), ("ALIGN", (3, 1), (7, -1), "RIGHT"),
        ("ALIGN", (8, 1), (8, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5), ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(KeepTogether([bar, Spacer(1, 3), t]))
    story.append(Spacer(1, 12))


def location_bar(loc, idx):
    """Prominent full-width divider marking the start of a location.
    Holds the location number + DBA, with the Named Insured inside the band."""
    label = f"LOCATION {txt(loc['loc_no']) or idx}"
    name = txt(loc["dba"])
    corp = txt(loc.get("corp"))
    line1 = ParagraphStyle("locbar1", fontName="Helvetica-Bold", fontSize=12.5,
                           leading=15, textColor=colors.white)
    line2 = ParagraphStyle("locbar2", fontName="Helvetica", fontSize=8.5,
                           leading=11, textColor=HUB_ARCTIC_GRAY)
    head = label + (f'&nbsp;&nbsp;&mdash;&nbsp;&nbsp;<font size=11>{name}</font>' if name else '')
    rows = [[Paragraph(head, line1)]]
    if corp:
        rows.append([Paragraph(f"Named Insured: {corp}", line2)])
    t = Table(rows, colWidths=[CONTENT_W])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), HUB_CLASSIC_BLUE),
        ("LINEBELOW", (0, -1), (-1, -1), 3.5, HUB_GOLD),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (0, 0), 7),
        ("BOTTOMPADDING", (0, -1), (-1, -1), 7),
        ("TOPPADDING", (0, 1), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 1),
    ]))
    return t


def build_location(story, loc, idx, multi):
    block = []
    if multi:
        block.append(Spacer(1, 8))
        block.append(location_bar(loc, idx))
        block.append(Spacer(1, 7))
        block.append(P(txt(loc["address"]), SMALL))
        block.append(Spacer(1, 7))

    # --- Premises Information
    block.append(section_bar("Premises Information"))
    w1, w2 = 2.7 * inch, CONTENT_W - 2.7 * inch
    block.append(kv_table([
        [lbl("Number of rooms:"), ans(txt(loc["rooms"]))],
        [lbl("Number of buildings at this location:"), ans(txt(loc["num_buildings"]))],
        [lbl("Construction / Year built / Floors:"),
         ans(", ".join([x for x in [txt(loc["construction"]), txt(loc["yr_built"]), (txt(loc["floors"]) + " floors" if loc["floors"] else "")] if x]))],
        [lbl("EIFS (Exterior Insulating Finishing System):"), ans(txt(loc["eifs"]))],
        [lbl("Corridor (Interior / Exterior):"), ans(txt(loc["corridor"]))],
        [lbl("Updates &ndash; Electrical:"), ans(year_age(loc["yr_elec"]))],
        [lbl("Updates &ndash; Plumbing:"), ans(year_age(loc["yr_plumb"]))],
        [lbl("Updates &ndash; Roofing:"), ans(year_age(loc["yr_roof"]))],
        [lbl("Updates &ndash; HVAC:"), ans(year_age(loc["yr_hvac"]))],
    ], [w1, w2]))
    block.append(kv_table([
        [lbl("Are buildings sprinklered?"), yn("Yes" if _truthy(loc["sprinkler"]) else ""),
         lbl("Percentage:"), ans(pct(loc["sprinkler"]))],
        [lbl("Smoke detectors?"), yn("Yes" if loc["smoke"] else ""),
         lbl("Type:"), ans(txt(loc["smoke"]))],
        [lbl("Fire alarms?"), yn("Yes" if loc["fire"] else ""),
         lbl("Type:"), ans(txt(loc["fire"]))],
        [lbl("Aluminum wiring on premises?"),
         yn("No" if loc["wiring"] and "alum" not in str(loc["wiring"]).lower() else ""),
         lbl("Wiring type:"), ans(txt(loc["wiring"]))],
    ], [1.8 * inch, 1.3 * inch, 1.1 * inch, CONTENT_W - 4.2 * inch]))
    block.append(Spacer(1, 4))

    # Property values being used (limits applied for this location)
    block.append(P("Property Values Being Used:", BOLD))
    pv = Table([
        [P("Building", BOLD), P("Contents", BOLD), P("Business Income", BOLD), P("TIV", BOLD)],
        [P(money(loc["bldg_limit"])), P(money(loc["contents_limit"])),
         P(money(loc["bi_limit"])), P(money(loc["tiv"]))],
    ], colWidths=[CONTENT_W / 4.0] * 4)
    pv.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), HUB_ARCTIC_GRAY),
        ("TEXTCOLOR", (0, 0), (-1, 0), HUB_CLASSIC_BLUE),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    block.append(pv)
    block.append(Spacer(1, 6))

    # --- Pools
    no_pools = not _truthy(loc["num_pools"])
    block.append(section_bar("Pools"))
    block.append(kv_table([
        [checkbox(no_pools, "Check here if no Pools"), lbl("How many swimming pools?"), ans(txt(loc["num_pools"]))],
    ], [2.6 * inch, 2.0 * inch, CONTENT_W - 4.6 * inch]))
    if not no_pools and loc["pool_amen"]:
        block.append(kv_table([[lbl("Pool details / safety:"), P(txt(loc["pool_amen"]), SMALL)]],
                              [1.6 * inch, CONTENT_W - 1.6 * inch]))
    block.append(Spacer(1, 6))

    # --- Recreational Facilities
    block.append(section_bar("Recreational Facilities"))
    fit = loc["fitness"]
    has_fit = bool(fit) and str(fit).strip().lower() not in ("none", "no", "n/a")
    block.append(kv_table([
        [lbl("Are there any exercise facilities?"), yn("Yes" if has_fit else "No"),
         lbl("Describe:"), ans(txt(fit) if has_fit else "")],
    ], [2.2 * inch, 1.2 * inch, 0.9 * inch, CONTENT_W - 4.3 * inch]))
    block.append(Spacer(1, 6))

    # --- Restaurant / Cooking
    restaurant_val = loc["restaurant"] if loc["restaurant"] not in (None, "") else "No"
    block.append(section_bar("Restaurant / Cooking Exposure"))
    block.append(kv_table([
        [lbl("Restaurant on premises?"), yn(restaurant_val),
         lbl("NFPA 90 or UL 300 equivalent standards (cooking)?"), ans(txt(loc["nfpa"]))],
    ], [1.9 * inch, 1.1 * inch, 2.7 * inch, CONTENT_W - 5.7 * inch]))
    block.append(Spacer(1, 6))

    # --- Liquor Liability (single question + hours of operation)
    liquor_val = loc["liquor"] if loc["liquor"] not in (None, "") else "No"
    hours = ""
    if loc["hours_open"] or loc["hours_closed"]:
        ho, hc = txt(loc["hours_open"]), txt(loc["hours_closed"])
        hours = f"{ho} &ndash; {hc}".strip(" &ndash;")
    block.append(section_bar("Liquor Liability"))
    block.append(kv_table([
        [lbl("Liquor sold / furnished?"), yn(liquor_val),
         lbl("Hours of operation:"), ans(hours)],
    ], [2.0 * inch, 1.1 * inch, 1.6 * inch, CONTENT_W - 4.7 * inch]))
    block.append(Spacer(1, 6))

    # --- Bar / Lounge (Yes/No; blank treated as No)
    bar_val = "No" if (str(loc["dance_floor"]).strip().lower() in ("no", "none", "n/a", "")
                       and str(loc["liquor"]).strip().lower() in ("no", "none", "n/a", "")) else \
              (loc["dance_floor"] if loc["dance_floor"] not in (None, "") else "")
    block.append(section_bar("Bar / Lounge"))
    block.append(kv_table([
        [lbl("Bar / lounge on premises?"), yn(bar_val)],
    ], [2.4 * inch, CONTENT_W - 2.4 * inch]))
    block.append(Spacer(1, 6))

    # --- Live Entertainment (Yes/No; blank always No)
    ent_val = loc["live_ent"] if loc["live_ent"] not in (None, "") else "No"
    block.append(section_bar("LIVE Entertainment"))
    block.append(kv_table([
        [lbl("Live entertainment on premises?"), yn(ent_val)],
    ], [2.4 * inch, CONTENT_W - 2.4 * inch]))
    block.append(Spacer(1, 6))

    # --- Gross Receipts (this location) -- kept together as one block
    gr_start = len(block)
    block.append(section_bar("Gross Receipts &mdash; This Location"))
    block.append(kv_table([
        [lbl("Average room rate (ADR):"),
         ans(money(loc["adr"]) + ("  per " + txt(loc["room_rentals"]) if loc["room_rentals"] else "")),
         lbl("Occupancy rate:"), ans(pct(loc["occupancy"]))],
    ], [1.9 * inch, 1.9 * inch, 1.3 * inch, CONTENT_W - 5.1 * inch]))
    block.append(Spacer(1, 4))
    gr = Table([
        [P("Hotel Operations", BOLD), P("Restaurant", BOLD), P("Liquor", BOLD), P("Other", BOLD), P("Total", BOLD)],
        [P(money(loc["hotel_sales"])), P(money(loc["restaurant_sales"])),
         P(money(loc["liquor_sales"])), P(money(loc["other_sales"])), P(money(loc["total_sales"]))],
    ], colWidths=[CONTENT_W / 5.0] * 5)
    gr.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), HUB_ARCTIC_GRAY),
        ("TEXTCOLOR", (0, 0), (-1, 0), HUB_CLASSIC_BLUE),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    block.append(gr)
    block.append(Spacer(1, 10))

    # keep the location header + premises start together, and keep the entire
    # Gross Receipts section from splitting across a page break.
    story.append(KeepTogether(block[:6]))
    for el in block[6:gr_start]:
        story.append(el)
    story.append(KeepTogether(block[gr_start:]))


def build_hnoa(story, app):
    story.append(section_bar("Hired &amp; Non-Owned Auto"))
    auto = str(app["owned_autos"]).strip()
    block = [
        [lbl("Number of owned autos:"), ans(txt(app["owned_autos"]))],
        [lbl("Do you provide guest shuttle / transportation services?"),
         yn(app["guest_transport"])],
        [lbl("Valet parking service?"), yn(app["valet"])],
    ]
    story.append(kv_table(block, [3.6 * inch, CONTENT_W - 3.6 * inch]))
    if app["hnoa"]:
        story.append(kv_table([[lbl("HNOA controls (from SOV):"), P(txt(app["hnoa"]), SMALL)]],
                              [1.9 * inch, CONTENT_W - 1.9 * inch]))
    story.append(Spacer(1, 12))


def build_signature(story, app):
    story.append(section_bar("Representation &amp; Warranty Statement"))
    story.append(Spacer(1, 4))
    rep = ("I have read this Application and I represent that all of the foregoing statements are true and "
           "accurate and that these statements are offered as the basis upon which Promont is considering "
           "issuance of an insurance policy. Any missing or erroneous information in this Application may "
           "jeopardize coverage in the event of a claim under any policy issued by Promont.")
    warn = ("WARNING: Any person who knowingly and with intent to defraud any insurance company or other "
            "person files an application for insurance or statement of claim containing any materially false "
            "information, or conceals for the purpose of misleading, information concerning any fact material "
            "thereto, commits a fraudulent insurance act, which is a crime, and shall also be subject to a "
            "civil penalty not to exceed five thousand dollars and the stated value of the claim for each "
            "such violation.")
    story.append(P(rep, FINE))
    story.append(Spacer(1, 6))
    story.append(P(warn, FINE))
    story.append(Spacer(1, 22))

    sig = Table([
        [P("X", BODY), "", P("", BODY), "", P("", BODY)],
        [P("Applicant Signature", SMALL), "", P("Title", SMALL), "", P("Date", SMALL)],
    ], colWidths=[3.0 * inch, 0.3 * inch, 1.8 * inch, 0.3 * inch, 1.6 * inch])
    sig.setStyle(TableStyle([
        ("LINEABOVE", (0, 1), (0, 1), 0.6, colors.black),
        ("LINEABOVE", (2, 1), (2, 1), 0.6, colors.black),
        ("LINEABOVE", (4, 1), (4, 1), 0.6, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("TOPPADDING", (0, 0), (0, 0), 14),
    ]))
    story.append(sig)
    story.append(Spacer(1, 8))
    story.append(P(f"Printed name: {txt(app['contact']) or '____________________________'}", SMALL))


def _truthy(v):
    try:
        return float(str(v).replace("$", "").replace(",", "").replace("%", "").strip()) > 0
    except (TypeError, ValueError):
        return str(v).strip().lower() in ("yes", "y", "true")


# ---------------------------------------------------------------- footer
def _footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(HUB_CLASSIC_BLUE)
    canvas.drawString(0.75 * inch, 0.5 * inch, APP_FOOTER)
    canvas.drawRightString(letter[0] - 0.75 * inch, 0.5 * inch, f"Page {doc.page}")
    canvas.restoreState()


# ---------------------------------------------------------------- main
def generate(sov_path, out_path):
    app, locations = read_sov(sov_path)
    multi = len(locations) > 1

    doc = SimpleDocTemplate(
        out_path, pagesize=letter,
        leftMargin=0.75 * inch, rightMargin=0.75 * inch,
        topMargin=0.7 * inch, bottomMargin=0.7 * inch,
        title=APP_FOOTER,
    )
    story = []
    build_business_info(story, app)
    if multi:
        entities = []
        for loc in locations:
            c = txt(loc.get("corp"))
            if c and c not in entities:
                entities.append(c)
        if len(entities) > 1:
            story.append(P("<b>Note:</b> This portfolio includes multiple named "
                           "insureds / entities (see each location). Confirm whether "
                           "a single combined application or separate applications "
                           "per entity are required:", SMALL))
            story.append(P("&nbsp;&nbsp;&bull;&nbsp; " + " &nbsp;&bull;&nbsp; ".join(entities), SMALL))
            story.append(Spacer(1, 6))
        build_property_schedule(story, locations)
        build_liability_schedule(story, locations)
    for i, loc in enumerate(locations, start=1):
        build_location(story, loc, i, multi)
    build_hnoa(story, app)
    build_signature(story, app)

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return len(locations)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 sov_to_application.py <SOV.xlsx> <output.pdf>")
        sys.exit(1)
    n = generate(sys.argv[1], sys.argv[2])
    print(f"Generated {sys.argv[2]} ({n} location(s)).")
