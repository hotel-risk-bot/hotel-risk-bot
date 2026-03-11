#!/usr/bin/env python3
"""
Loss Run Organizer for HUB Hotel Franchise Practice.

Watches a Google Drive "inbox" folder for new loss run files (PDF + Excel),
uses GPT to extract metadata (client, policy type, carrier, valuation date,
policy years), then:
  1. Moves the file into: 0001-1 Client Loss Runs / {Client} / {Year} / {PolicyType}
  2. Renames with convention: {PolicyType} {YY-YY} {ValDate}_{Carrier}.{ext}
  3. Updates a Loss Run Tracker Google Sheet

Designed to run on a schedule via APScheduler inside bot.py on Railway.

Requires:
  - GOOGLE_SERVICE_ACCOUNT_JSON env var (service account with Drive + Sheets scope)
  - OPENAI_API_KEY env var
  - LOSS_RUN_INBOX_FOLDER_ID env var (Google Drive folder ID for the inbox)
  - CLIENT_LOSS_RUNS_FOLDER_ID env var (destination folder for organized files)
  - LOSS_RUN_TRACKER_SHEET_ID env var (Google Sheet ID for the tracker)
  - TELEGRAM_TOKEN + TELEGRAM_CHAT_ID env vars (for notifications)
"""

import os
import io
import json
import logging
import time
import tempfile
import re
from datetime import datetime, date

import jwt
import requests as http_requests

logger = logging.getLogger(__name__)

# ── Configuration ─────────────────────────────────────────────────────────
GOOGLE_SERVICE_ACCOUNT_JSON = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")
LOSS_RUN_INBOX_FOLDER_ID = os.environ.get("LOSS_RUN_INBOX_FOLDER_ID", "")
LOSS_RUN_TRACKER_SHEET_ID = os.environ.get("LOSS_RUN_TRACKER_SHEET_ID", "")
TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "")

# Hardcoded fallback folder IDs (in case env vars have timing issues)
_FALLBACK_CLIENT_LOSS_RUNS_FOLDER_ID = os.environ.get("CLIENT_LOSS_RUNS_FOLDER_ID_FALLBACK", "1v2-Y9pKIY4_Jh3X2_ZJOCB7XdNLptS8x")

# Google API scopes needed
SCOPES = "https://www.googleapis.com/auth/drive https://www.googleapis.com/auth/spreadsheets"

# Supported file extensions
SUPPORTED_EXTENSIONS = (".pdf", ".xlsx", ".xls")

# Policy type mapping (normalized)
POLICY_TYPES = {
    "property": "Property",
    "general liability": "Liability",
    "liability": "Liability",
    "gl": "Liability",
    "commercial general liability": "Liability",
    "workers compensation": "Workers Comp",
    "workers comp": "Workers Comp",
    "wc": "Workers Comp",
    "work comp": "Workers Comp",
    "umbrella": "Umbrella",
    "excess": "Umbrella",
    "excess liability": "Umbrella",
    "auto": "Auto",
    "commercial auto": "Auto",
    "hired non-owned auto": "Auto",
    "epli": "EPLI",
    "employment practices": "EPLI",
    "crime": "Crime",
    "cyber": "Cyber",
    "inland marine": "Inland Marine",
    "equipment breakdown": "Equipment Breakdown",
    "liquor liability": "Liquor Liability",
}

def _get_tracker_sheet_id():
    """Get tracker sheet ID at runtime."""
    return os.environ.get("LOSS_RUN_TRACKER_SHEET_ID", "").strip()


# Tracker sheet headers
TRACKER_HEADERS = [
    "Client", "Policy Type", "Carrier", "Valuation Date",
    "File Name", "Drive Link", "Date Organized", "Year Folder"
]


# ── Client Name Normalization ────────────────────────────────────────────

# Common suffixes to strip for matching
_STRIP_SUFFIXES = [
    "llc", "inc", "corp", "corporation", "company", "co",
    "ltd", "lp", "lllp", "pllc", "pc", "pa",
    "dba", "management", "mgmt", "hospitality", "hotel", "hotels",
    "group", "properties", "enterprises", "enterprise",
    "real estate", "investments", "investors",
]

def _normalize_client_name(name):
    """
    Normalize a client name for comparison purposes.
    Strips punctuation (commas, periods), extra whitespace, and common suffixes
    so that 'Pride Management Inc.' and 'Pride Management, Inc.' match.
    """
    if not name:
        return ""
    # Lowercase
    n = name.strip().lower()
    # Remove common punctuation that varies between sources
    n = n.replace(",", "").replace(".", "").replace("'", "").replace('"', '')
    # Remove anything after "dba" for matching purposes
    if " dba " in n:
        n = n.split(" dba ")[0].strip()
    # Remove semicolons and everything after (e.g., "Kautilya Management LLC; Vinit")
    if ";" in n:
        n = n.split(";")[0].strip()
    # Collapse multiple spaces
    n = re.sub(r'\s+', ' ', n).strip()
    return n


def _extract_core_name(name):
    """
    Extract the core business name by removing common legal suffixes.
    Used for fuzzy matching when exact normalized match fails.
    E.g., 'Dalwadi Hospitality Management LLC' -> 'dalwadi'
    """
    n = _normalize_client_name(name)
    if not n:
        return ""
    # Remove common suffixes iteratively
    words = n.split()
    core_words = []
    for w in words:
        if w not in _STRIP_SUFFIXES:
            core_words.append(w)
    return " ".join(core_words).strip()


def _client_names_match(name1, name2):
    """
    Determine if two client names refer to the same entity.
    Uses multiple strategies:
    1. Exact normalized match
    2. Core name containment
    3. First significant word match (for cases like 'DALWADI HOSPITALITY' vs 'Dalwadi Hospitality Management LLC')
    """
    norm1 = _normalize_client_name(name1)
    norm2 = _normalize_client_name(name2)

    # Exact normalized match
    if norm1 == norm2:
        return True

    # One contains the other
    if norm1 in norm2 or norm2 in norm1:
        return True

    # Core name match
    core1 = _extract_core_name(name1)
    core2 = _extract_core_name(name2)
    if core1 and core2 and (core1 == core2 or core1 in core2 or core2 in core1):
        return True

    # First significant word match (for company groups like Kautilya, Dalwadi, etc.)
    # Only if the first word is distinctive enough (> 4 chars)
    words1 = core1.split() if core1 else []
    words2 = core2.split() if core2 else []
    if words1 and words2 and len(words1[0]) > 4 and words1[0] == words2[0]:
        return True

    return False


# ── Google API Auth ───────────────────────────────────────────────────────

def _parse_service_account_json(raw):
    """Parse service account JSON, handling Railway env var mangling."""
    if not raw:
        return None
    raw = raw.strip()
    # Fix truncated JSON: if it starts with { but doesn't end with }, add it
    if raw.startswith('{') and not raw.endswith('}'):
        raw = raw + '\n}'
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass
    try:
        fixed = re.sub(
            r'"((?:[^"\\]|\\.)*?)"',
            lambda m: '"' + m.group(1).replace('\n', '\\n').replace('\r', '') + '"',
            raw,
            flags=re.DOTALL
        )
        return json.loads(fixed)
    except Exception:
        pass
    try:
        lines = raw.split('\n')
        rebuilt = []
        in_pk = False
        for line in lines:
            s = line.strip()
            if '"private_key"' in s:
                in_pk = True
                rebuilt.append(s)
            elif in_pk:
                if (s.startswith('"') and not s.startswith('"-----')) or s in ('}', '},'):
                    in_pk = False
                    rebuilt.append(s)
                else:
                    if rebuilt:
                        rebuilt[-1] = rebuilt[-1].rstrip() + '\\n' + s
                    else:
                        rebuilt.append(s)
            else:
                rebuilt.append(s)
        rejoined = ' '.join(rebuilt)
        return json.loads(rejoined)
    except Exception as e:
        logger.error(f"All JSON parse attempts failed: {e}")
        return None


def _get_access_token():
    """Get Google API access token with Drive + Sheets scope."""
    sa_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    if not sa_json:
        logger.error("GOOGLE_SERVICE_ACCOUNT_JSON not set")
        return None

    creds = _parse_service_account_json(sa_json)
    if not creds:
        logger.error("Could not parse GOOGLE_SERVICE_ACCOUNT_JSON")
        return None

    # Ensure private_key has proper newline characters
    pk = creds.get("private_key", "")
    if pk and '\n' not in pk and '\\n' in pk:
        creds["private_key"] = pk.replace('\\n', '\n')

    now = int(time.time())
    payload = {
        "iss": creds["client_email"],
        "scope": SCOPES,
        "aud": "https://oauth2.googleapis.com/token",
        "iat": now,
        "exp": now + 3600,
    }

    try:
        signed_jwt = jwt.encode(payload, creds["private_key"], algorithm="RS256")
    except Exception as e:
        logger.error(f"JWT signing failed: {e}")
        return None

    resp = http_requests.post(
        "https://oauth2.googleapis.com/token",
        data={
            "grant_type": "urn:ietf:params:oauth:grant-type:jwt-bearer",
            "assertion": signed_jwt,
        },
        timeout=15,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


def _auth_headers():
    """Get authorization headers."""
    token = _get_access_token()
    if not token:
        return None
    return {"Authorization": f"Bearer {token}"}


# ── Google Drive API Helpers ──────────────────────────────────────────────

def drive_list_files(folder_id, mime_type=None):
    """List files in a Google Drive folder."""
    headers = _auth_headers()
    if not headers:
        logger.error("No auth headers available")
        return []

    q = f"'{folder_id}' in parents and trashed = false"
    logger.info(f"Querying Drive folder {folder_id}")

    files = []
    page_token = None

    while True:
        params = {
            "q": q,
            "fields": "nextPageToken, files(id, name, mimeType, modifiedTime, webViewLink)",
            "pageSize": 100,
        }
        if page_token:
            params["pageToken"] = page_token

        resp = http_requests.get(
            "https://www.googleapis.com/drive/v3/files",
            headers=headers, params=params, timeout=30,
        )

        if resp.status_code != 200:
            logger.error(f"File list request failed. Status: {resp.status_code}")
            break

        data = resp.json()
        batch = data.get("files", [])
        files.extend(batch)
        page_token = data.get("nextPageToken")
        if not page_token:
            break

    logger.info(f"Found {len(files)} items in folder {folder_id}")
    return files


def drive_download_file(file_id):
    """Download a file's content as bytes."""
    headers = _auth_headers()
    if not headers:
        return None

    resp = http_requests.get(
        f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media",
        headers=headers, timeout=120,
    )
    resp.raise_for_status()
    return resp.content


def drive_create_folder(name, parent_id):
    """Create a folder in Google Drive. Returns the new folder ID."""
    headers = _auth_headers()
    if not headers:
        return None

    metadata = {
        "name": name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [parent_id],
    }

    resp = http_requests.post(
        "https://www.googleapis.com/drive/v3/files",
        headers={**headers, "Content-Type": "application/json"},
        json=metadata, timeout=30,
    )
    resp.raise_for_status()
    return resp.json()["id"]


def drive_find_or_create_folder(name, parent_id):
    """Find an existing subfolder by name, or create it."""
    headers = _auth_headers()
    if not headers:
        return None

    q = (
        f"'{parent_id}' in parents and "
        f"name = '{name}' and "
        f"mimeType = 'application/vnd.google-apps.folder' and "
        f"trashed = false"
    )
    resp = http_requests.get(
        "https://www.googleapis.com/drive/v3/files",
        headers=headers,
        params={"q": q, "fields": "files(id, name)", "pageSize": 1},
        timeout=30,
    )
    resp.raise_for_status()
    files = resp.json().get("files", [])

    if files:
        return files[0]["id"]
    return drive_create_folder(name, parent_id)


def drive_find_or_create_folder_normalized(client_name, parent_id):
    """
    Find an existing client subfolder using normalized name matching,
    or create a new one. This prevents duplicates like
    'Pride Management Inc.' vs 'Pride Management, Inc.'
    or 'DALWADI HOSPITALITY' vs 'Dalwadi Hospitality Management, LLC'
    """
    headers = _auth_headers()
    if not headers:
        return None, None

    # List all folders in the parent
    q = (
        f"'{parent_id}' in parents and "
        f"mimeType = 'application/vnd.google-apps.folder' and "
        f"trashed = false"
    )
    resp = http_requests.get(
        "https://www.googleapis.com/drive/v3/files",
        headers=headers,
        params={"q": q, "fields": "files(id, name)", "pageSize": 500},
        timeout=30,
    )
    resp.raise_for_status()
    folders = resp.json().get("files", [])

    # Try to find a match using our smart matching
    best_match = None
    best_score = 0

    for folder in folders:
        if _client_names_match(client_name, folder["name"]):
            # Calculate a score to prefer the best match
            norm_client = _normalize_client_name(client_name)
            norm_folder = _normalize_client_name(folder["name"])

            # Exact normalized match gets highest score
            if norm_client == norm_folder:
                logger.info(f"Exact match: '{client_name}' -> '{folder['name']}'")
                return folder["id"], folder["name"]

            # Score by similarity
            score = min(len(norm_client), len(norm_folder)) / max(len(norm_client), len(norm_folder), 1)
            if score > best_score:
                best_score = score
                best_match = folder

    if best_match:
        logger.info(f"Matched client '{client_name}' to existing folder '{best_match['name']}' (score={best_score:.2f})")
        return best_match["id"], best_match["name"]

    # No match — create new folder with a clean version of the client name
    clean_name = client_name.strip()
    # Title-case if all uppercase
    if clean_name == clean_name.upper() and len(clean_name) > 5:
        clean_name = clean_name.title()
    logger.info(f"No matching folder for '{client_name}' — creating '{clean_name}'")
    new_id = drive_create_folder(clean_name, parent_id)
    return new_id, clean_name


def drive_move_file(file_id, new_parent_id, new_name=None):
    """Move a file to a new folder and optionally rename it."""
    headers = _auth_headers()
    if not headers:
        return False

    # Get current parents
    resp = http_requests.get(
        f"https://www.googleapis.com/drive/v3/files/{file_id}",
        headers=headers, params={"fields": "parents"}, timeout=15,
    )
    resp.raise_for_status()
    current_parents = ",".join(resp.json().get("parents", []))

    # Build update
    params = {
        "addParents": new_parent_id,
        "removeParents": current_parents,
        "fields": "id, parents, name",
    }
    body = {}
    if new_name:
        body["name"] = new_name

    resp = http_requests.patch(
        f"https://www.googleapis.com/drive/v3/files/{file_id}",
        headers={**headers, "Content-Type": "application/json"},
        params=params, json=body, timeout=30,
    )
    resp.raise_for_status()
    return True


def drive_get_web_link(file_id):
    """Get the web view link for a file."""
    headers = _auth_headers()
    if not headers:
        return None

    resp = http_requests.get(
        f"https://www.googleapis.com/drive/v3/files/{file_id}",
        headers=headers, params={"fields": "webViewLink"}, timeout=15,
    )
    resp.raise_for_status()
    return resp.json().get("webViewLink", "")


# ── Text Extraction ──────────────────────────────────────────────────────

def _is_garbled_text(text):
    """
    Detect if extracted text is garbled (e.g., '/j0 /j0 /j0' patterns
    from PDFs with custom font encoding that standard extractors can't decode).
    """
    if not text or len(text.strip()) < 50:
        return True

    # Count the ratio of recognizable words vs gibberish
    # Garbled text typically has lots of /j0, /0, /1 patterns
    garble_pattern = re.findall(r'/[a-z]?\d+', text)
    if len(garble_pattern) > len(text) / 10:
        return True

    # Check if text has very few actual English words
    words = re.findall(r'[a-zA-Z]{3,}', text)
    if len(words) < 10 and len(text) > 200:
        return True

    return False


def _extract_text_from_pdf(file_bytes, filename):
    """Extract text from all pages of a PDF."""
    text = ""
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                text += f"\n--- Page {page.page_number} ---\n"
                text += page_text
    except Exception as e:
        logger.warning(f"pdfplumber extraction failed for {filename}: {e}")
    return text


def _extract_text_from_excel(file_bytes, filename):
    """Extract text from an Excel file for GPT analysis."""
    text = ""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames[:5]:  # Limit to first 5 sheets
            ws = wb[sheet_name]
            text += f"\n--- Sheet: {sheet_name} ---\n"
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                row_str = " | ".join(str(c) if c is not None else "" for c in row)
                if row_str.strip():
                    text += row_str + "\n"
                row_count += 1
                if row_count > 200:  # Limit rows per sheet
                    break
        wb.close()
    except Exception as e:
        logger.warning(f"Excel extraction failed for {filename}: {e}")
    return text


# ── GPT Extraction ───────────────────────────────────────────────────────

def extract_loss_run_metadata(file_bytes, filename):
    """
    Use GPT to extract client name, policy type, carrier, valuation date,
    and policy years from a loss run file (PDF or Excel).
    """
    if not OPENAI_API_KEY:
        logger.error("OPENAI_API_KEY not set")
        return None

    # Extract text based on file type
    lower_name = filename.lower()
    if lower_name.endswith(".pdf"):
        text = _extract_text_from_pdf(file_bytes, filename)
        # If text is garbled or too short, use OCR fallback
        if _is_garbled_text(text):
            logger.info(f"Garbled/insufficient text for {filename} ({len(text)} chars), using OCR")
            return _extract_via_ocr(file_bytes, filename)
    elif lower_name.endswith((".xlsx", ".xls")):
        text = _extract_text_from_excel(file_bytes, filename)
    else:
        logger.warning(f"Unsupported file type: {filename}")
        return None

    if not text or len(text.strip()) < 50:
        logger.warning(f"No text extracted from {filename}")
        return None

    # Ask GPT to extract metadata including policy years
    prompt = f"""You are analyzing an insurance loss run document. This document may contain MULTIPLE policy terms/periods across different pages.

Extract the following:

1. **Client Name** (the insured/policyholder, NOT the insurance company or agent)
2. **Policy Type** (one of: Property, Liability, Workers Comp, Umbrella, Auto, EPLI, Crime, Cyber, Inland Marine, Equipment Breakdown, Liquor Liability)
3. **Carrier** (the insurance company that issued the loss run)
4. **Valuation Date** (also called "report date", "as of date", or "valued as of" — the date through which losses are reported. Format as YYYY-MM-DD)
5. **Policy Years** — IMPORTANT: Look at ALL policy terms across ALL pages. Find the EARLIEST policy start year and the LATEST policy end year. Express as "YY-YY" using 2-digit years.
   - Example: If you see terms "12/20/2021 to 04/01/2022", "04/01/2022 to 04/01/2023", and "04/01/2023 to 04/01/2024", the policy years are "21-24"
   - Example: If there's only one term "01/01/2025 to 01/01/2026", the policy years are "25-26"
   - Use the START year of the earliest term and the END year of the latest term

Original filename: {filename}

Document text (all pages):
{text[:12000]}

Return ONLY valid JSON with these exact keys:
{{"client_name": "...", "policy_type": "...", "carrier": "...", "valuation_date": "YYYY-MM-DD", "policy_years": "YY-YY"}}

If you cannot determine a field, use "Unknown" for strings, "1900-01-01" for date, or "00-00" for policy_years.
"""

    try:
        resp = http_requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {OPENAI_API_KEY}",
                "Content-Type": "application/json",
            },
            json={
                "model": "gpt-4.1-mini",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.0,
                "max_tokens": 500,
            },
            timeout=60,
        )
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]

        # Parse JSON from response (handle nested braces)
        json_match = re.search(r'\{[^{}]*\}', content, re.DOTALL)
        if json_match:
            metadata = json.loads(json_match.group())
            # Normalize policy type
            metadata["policy_type"] = _normalize_policy_type(metadata.get("policy_type", "Unknown"))
            # Ensure policy_years is present
            if "policy_years" not in metadata:
                metadata["policy_years"] = "00-00"
            logger.info(f"Extracted metadata for {filename}: {metadata}")
            return metadata
    except Exception as e:
        logger.error(f"GPT extraction failed for {filename}: {e}")

    return None


def _extract_via_ocr(pdf_bytes, filename):
    """OCR fallback using GPT Vision for scanned/garbled PDFs."""
    try:
        from pdf2image import convert_from_bytes
        import base64

        # Convert up to 3 pages for better coverage
        images = convert_from_bytes(pdf_bytes, first_page=1, last_page=3, dpi=150, fmt="jpeg")
        if not images:
            logger.warning(f"pdf2image returned no images for {filename}")
            return None

        # Build image content for GPT Vision (send up to 2 pages)
        image_content = []
        image_content.append({"type": "text", "text": (
            "These are pages from a scanned insurance loss run document. Extract:\n"
            "1. Client Name (the insured/policyholder, NOT the insurance company or agent)\n"
            "2. Policy Type (Property, Liability, Workers Comp, Umbrella, Auto, EPLI, etc.)\n"
            "3. Carrier (insurance company that issued the loss run)\n"
            "4. Valuation Date (report date / as of date, format as YYYY-MM-DD)\n"
            "5. Policy Years (earliest start year to latest end year as YY-YY, e.g. 21-24)\n\n"
            f"Original filename: {filename}\n\n"
            'Return ONLY valid JSON: {"client_name": "...", "policy_type": "...", "carrier": "...", "valuation_date": "YYYY-MM-DD", "policy_years": "YY-YY"}\n'
            'If you cannot determine a field, use "Unknown" for strings, "1900-01-01" for date, or "00-00" for policy_years.'
        )})

        for i, img in enumerate(images[:2]):
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=70)
            img_b64 = base64.b64encode(buf.getvalue()).decode()
            image_content.append({
                "type": "image_url",
                "image_url": {"url": f"data:image/jpeg;base64,{img_b64}", "detail": "low"}
            })

        resp = http_requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {OPENAI_API_KEY}",
                "Content-Type": "application/json",
            },
            json={
                "model": "gpt-4.1-mini",
                "messages": [{"role": "user", "content": image_content}],
                "temperature": 0.0,
                "max_tokens": 500,
            },
            timeout=90,
        )
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]
        json_match = re.search(r'\{[^{}]*\}', content, re.DOTALL)
        if json_match:
            metadata = json.loads(json_match.group())
            metadata["policy_type"] = _normalize_policy_type(metadata.get("policy_type", "Unknown"))
            if "policy_years" not in metadata:
                metadata["policy_years"] = "00-00"
            logger.info(f"OCR extracted metadata for {filename}: {metadata}")
            return metadata
    except ImportError:
        logger.error("pdf2image not installed - OCR fallback unavailable")
    except Exception as e:
        logger.error(f"OCR extraction failed for {filename}: {e}")

    return None


def _normalize_policy_type(raw_type):
    """Normalize policy type string to standard categories."""
    if not raw_type:
        return "Other"
    key = raw_type.strip().lower()
    return POLICY_TYPES.get(key, raw_type.title())


# ── File Naming ──────────────────────────────────────────────────────────

def _build_filename(policy_type, policy_years, valuation_date, carrier, original_ext):
    """
    Build the standardized filename.
    Format: {PolicyType} {YY-YY} {ValDate}_{CarrierName}.{ext}
    Example: Liability 21-24 2026-02-27_Trisura_Specialty_Ins.pdf
    """
    safe_carrier = re.sub(r'[^\w\s-]', '', carrier).strip().replace(' ', '_')
    # Truncate carrier name if too long (keep it readable)
    if len(safe_carrier) > 40:
        safe_carrier = safe_carrier[:40]
    safe_policy = policy_type.replace(' ', '_')

    # Build the name
    if policy_years and policy_years != "00-00":
        new_name = f"{safe_policy} {policy_years} {valuation_date}_{safe_carrier}{original_ext}"
    else:
        new_name = f"{safe_policy} {valuation_date}_{safe_carrier}{original_ext}"

    return new_name


# ── Tracker Sheet ─────────────────────────────────────────────────────────

def _sheets_headers():
    """Get auth headers for Sheets API (reuses Drive token with combined scope)."""
    headers = _auth_headers()
    if not headers:
        return None
    return {**headers, "Content-Type": "application/json"}


def tracker_initialize():
    """Ensure the tracker sheet has headers."""
    sheet_id = os.environ.get("LOSS_RUN_TRACKER_SHEET_ID", "").strip()
    if not sheet_id:
        logger.warning("LOSS_RUN_TRACKER_SHEET_ID not set - tracker disabled")
        return False

    headers = _sheets_headers()
    if not headers:
        return False

    # Check if headers exist
    sid = _get_tracker_sheet_id()
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{sid}/values/Sheet1!A1:H1"
    try:
        resp = http_requests.get(url, headers=headers, timeout=15)
        resp.raise_for_status()
        values = resp.json().get("values", [])
        if not values or values[0] != TRACKER_HEADERS:
            # Write headers
            body = {"values": [TRACKER_HEADERS]}
            url = (
                f"https://sheets.googleapis.com/v4/spreadsheets/{sid}"
                f"/values/Sheet1!A1:H1?valueInputOption=USER_ENTERED"
            )
            resp = http_requests.put(url, headers=headers, json=body, timeout=15)
            resp.raise_for_status()
            logger.info("Tracker sheet headers initialized")
        return True
    except Exception as e:
        logger.error(f"Tracker initialization failed: {e}")
        return False


def tracker_add_entry(client, policy_type, carrier, valuation_date, filename, drive_link, year_folder):
    """Add or update an entry in the tracker sheet."""
    if not os.environ.get("LOSS_RUN_TRACKER_SHEET_ID", "").strip():
        return False

    headers = _sheets_headers()
    if not headers:
        return False

    today = date.today().isoformat()
    row = [client, policy_type, carrier, valuation_date, filename, drive_link, today, year_folder]

    # First, check if there's an existing row for this client + policy type to update
    try:
        url = f"https://sheets.googleapis.com/v4/spreadsheets/{_get_tracker_sheet_id()}/values/Sheet1!A:H"
        resp = http_requests.get(url, headers=headers, timeout=15)
        resp.raise_for_status()
        all_rows = resp.json().get("values", [])

        # Look for existing entry (same client + policy type) — update if this valuation is newer
        for i, existing_row in enumerate(all_rows[1:], start=2):  # skip header
            if len(existing_row) >= 2:
                if (_client_names_match(existing_row[0], client) and
                        existing_row[1].strip().lower() == policy_type.strip().lower()):
                    # Found match — check if new valuation is more recent
                    existing_val = existing_row[3] if len(existing_row) > 3 else ""
                    if valuation_date >= existing_val:
                        # Update this row
                        range_str = f"Sheet1!A{i}:H{i}"
                        url = (
                            f"https://sheets.googleapis.com/v4/spreadsheets/{_get_tracker_sheet_id()}"
                            f"/values/{range_str}?valueInputOption=USER_ENTERED"
                        )
                        resp = http_requests.put(url, headers=headers, json={"values": [row]}, timeout=15)
                        resp.raise_for_status()
                        logger.info(f"Updated tracker: {client} / {policy_type} -> {valuation_date}")
                        return True
                    else:
                        logger.info(f"Skipping tracker update — existing valuation {existing_val} is newer")
                        return True

        # No existing entry — append new row
        url = (
            f"https://sheets.googleapis.com/v4/spreadsheets/{_get_tracker_sheet_id()}"
            f"/values/Sheet1!A:H:append?valueInputOption=USER_ENTERED&insertDataOption=INSERT_ROWS"
        )
        resp = http_requests.post(url, headers=headers, json={"values": [row]}, timeout=15)
        resp.raise_for_status()
        logger.info(f"Added to tracker: {client} / {policy_type} / {carrier} / {valuation_date}")
        return True

    except Exception as e:
        logger.error(f"Tracker update failed: {e}")
        return False


def tracker_get_all():
    """Get all entries from the tracker sheet."""
    if not os.environ.get("LOSS_RUN_TRACKER_SHEET_ID", "").strip():
        return []

    headers = _sheets_headers()
    if not headers:
        return []

    try:
        url = f"https://sheets.googleapis.com/v4/spreadsheets/{_get_tracker_sheet_id()}/values/Sheet1!A:H"
        resp = http_requests.get(url, headers=headers, timeout=15)
        resp.raise_for_status()
        rows = resp.json().get("values", [])
        if len(rows) <= 1:
            return []
        # Return as list of dicts
        header = rows[0]
        return [dict(zip(header, row + [""] * (len(header) - len(row)))) for row in rows[1:]]
    except Exception as e:
        logger.error(f"Tracker read failed: {e}")
        return []


def tracker_get_client(client_name):
    """Get tracker entries for a specific client."""
    all_entries = tracker_get_all()
    return [e for e in all_entries if _client_names_match(e.get("Client", ""), client_name)]


# ── Main Organizer Logic ─────────────────────────────────────────────────

def organize_loss_runs(accounts_folder_id=None):
    """
    Main entry point. Scans the inbox folder, processes each file (PDF/Excel),
    and organizes into: 0001-1 Client Loss Runs / {Client} / {Year} / {PolicyType}

    Returns a summary dict with counts and details.
    """
    # Read env vars at runtime
    inbox_id = os.environ.get("LOSS_RUN_INBOX_FOLDER_ID", "").strip()
    if not inbox_id:
        logger.error("LOSS_RUN_INBOX_FOLDER_ID not set")
        return {"error": "Inbox folder not configured", "processed": 0}

    # Use CLIENT_LOSS_RUNS_FOLDER_ID as the destination (with hardcoded fallback)
    dest_folder_id = os.environ.get("CLIENT_LOSS_RUNS_FOLDER_ID", "").strip()
    if not dest_folder_id:
        dest_folder_id = _FALLBACK_CLIENT_LOSS_RUNS_FOLDER_ID
        logger.warning(f"CLIENT_LOSS_RUNS_FOLDER_ID not in env, using hardcoded fallback: {dest_folder_id}")

    logger.info(f"Destination folder: {dest_folder_id}")

    # Initialize tracker
    tracker_initialize()

    # Get all files in inbox
    inbox_files = drive_list_files(inbox_id)

    # Filter for supported file types (PDF + Excel)
    supported_files = [
        f for f in inbox_files
        if any(f["name"].lower().endswith(ext) for ext in SUPPORTED_EXTENSIONS)
    ]

    if not supported_files:
        logger.info("No supported files in inbox folder")
        return {"processed": 0, "message": "No loss runs to process"}

    results = {
        "processed": 0,
        "success": [],
        "errors": [],
    }

    for file_info in supported_files:
        file_id = file_info["id"]
        filename = file_info["name"]
        logger.info(f"Processing: {filename}")

        try:
            # Determine file extension
            original_ext = ""
            for ext in SUPPORTED_EXTENSIONS:
                if filename.lower().endswith(ext):
                    original_ext = ext
                    break

            # 1. Download the file
            file_bytes = drive_download_file(file_id)
            if not file_bytes:
                results["errors"].append(f"{filename}: Download failed")
                continue

            # 2. Extract metadata via GPT
            metadata = extract_loss_run_metadata(file_bytes, filename)
            if not metadata:
                results["errors"].append(f"{filename}: Extraction failed")
                continue

            client_name = metadata.get("client_name", "Unknown")
            policy_type = metadata.get("policy_type", "Other")
            carrier = metadata.get("carrier", "Unknown")
            valuation_date = metadata.get("valuation_date", "1900-01-01")
            policy_years = metadata.get("policy_years", "00-00")

            # Determine year from valuation date
            try:
                val_year = str(datetime.strptime(valuation_date, "%Y-%m-%d").year)
            except ValueError:
                val_year = str(date.today().year)

            # 3. Find or create client folder (with normalized matching)
            client_folder_id, client_folder_name = drive_find_or_create_folder_normalized(
                client_name, dest_folder_id
            )
            if not client_folder_id:
                results["errors"].append(f"{filename}: Could not resolve client folder")
                continue

            # 4. Create subfolder path: {Client} / {Year} / {PolicyType}
            year_folder_id = drive_find_or_create_folder(val_year, client_folder_id)
            policy_folder_id = drive_find_or_create_folder(policy_type, year_folder_id)

            # 5. Build new filename
            new_name = _build_filename(policy_type, policy_years, valuation_date, carrier, original_ext)

            # 6. Move file to destination
            drive_move_file(file_id, policy_folder_id, new_name)

            # 7. Get the web link for the tracker
            web_link = drive_get_web_link(file_id)

            # 8. Update tracker sheet
            tracker_add_entry(
                client=client_folder_name or client_name,
                policy_type=policy_type,
                carrier=carrier,
                valuation_date=valuation_date,
                filename=new_name,
                drive_link=web_link or "",
                year_folder=val_year,
            )

            results["processed"] += 1
            results["success"].append({
                "original": filename,
                "new_name": new_name,
                "client": client_folder_name or client_name,
                "policy_type": policy_type,
                "carrier": carrier,
                "valuation_date": valuation_date,
                "policy_years": policy_years,
                "year": val_year,
            })
            logger.info(f"Organized: {filename} -> {client_folder_name}/{val_year}/{policy_type}/{new_name}")

        except Exception as e:
            logger.error(f"Error processing {filename}: {e}")
            results["errors"].append(f"{filename}: {str(e)}")

    return results


# ── Telegram Notification ─────────────────────────────────────────────────

def send_organize_summary(results):
    """Send a summary of the organization run to Telegram."""
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
        return

    lines = ["📂 *Loss Run Organizer Summary*\n"]

    if results.get("error"):
        lines.append(f"⚠️ Error: {results['error']}")
    elif results["processed"] == 0:
        lines.append("No new loss runs to process\\.")
    else:
        lines.append(f"✅ *{results['processed']}* loss run\\(s\\) organized:\n")
        for item in results.get("success", []):
            lines.append(
                f"• *{_escape_md(item['client'])}*\n"
                f"  {_escape_md(item['policy_type'])} \\| {_escape_md(item['carrier'])} \\| {_escape_md(item['valuation_date'])}"
            )

    if results.get("errors"):
        lines.append(f"\n⚠️ *{len(results['errors'])}* error\\(s\\):")
        for err in results["errors"][:5]:
            lines.append(f"  • {_escape_md(err)}")

    text = "\n".join(lines)

    try:
        http_requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
            json={
                "chat_id": TELEGRAM_CHAT_ID,
                "text": text,
                "parse_mode": "MarkdownV2",
            },
            timeout=15,
        )
    except Exception as e:
        logger.error(f"Telegram notification failed: {e}")


def _escape_md(text):
    """Escape Telegram MarkdownV2 special characters."""
    special = r'_*[]()~`>#+-=|{}.!'
    return ''.join(f'\\{c}' if c in special else c for c in str(text))


# ── Scheduled Run Entry Point ─────────────────────────────────────────────

async def scheduled_organize():
    """Entry point for APScheduler. Runs synchronously in a thread."""
    import asyncio
    loop = asyncio.get_event_loop()
    results = await loop.run_in_executor(None, organize_loss_runs)
    send_organize_summary(results)
    return results


def run_organize_sync():
    """Synchronous entry point for testing or manual runs."""
    results = organize_loss_runs()
    send_organize_summary(results)
    return results


# ── CLI for Testing ───────────────────────────────────────────────────────

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    results = run_organize_sync()
    print(json.dumps(results, indent=2))
