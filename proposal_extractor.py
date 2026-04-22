#!/usr/bin/env python3
"""
Hotel Insurance Proposal - Document Extraction & GPT Data Structuring
Extracts text from uploaded PDFs and Excel files, then uses GPT to structure
the data into a standardized format for proposal generation.

Key feature: Smart page-level extraction that identifies quote summary pages
and prioritizes them over forms/endorsements boilerplate.
"""

import os
import json
import logging
import re
import subprocess
import tempfile
from pathlib import Path
from typing import Optional

import openpyxl
from openai import OpenAI
import traceback

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

try:
    from pdf2image import convert_from_path
    HAS_PDF2IMAGE = True
except ImportError:
    HAS_PDF2IMAGE = False

logger = logging.getLogger(__name__)

# OpenAI client (lazy initialization)
_client = None
GPT_MODEL = "gpt-4.1"


def _get_openai_client():
    """Lazy-initialize the OpenAI client."""
    global _client
    if _client is None:
        api_key = os.environ.get("OPENAI_API_KEY")
        if not api_key:
            raise RuntimeError(
                "OPENAI_API_KEY environment variable not set. "
                "Please add it to your Railway environment variables."
            )
        import httpx
        _client = OpenAI(
            api_key=api_key,
            timeout=httpx.Timeout(600.0, connect=30.0)  # 10 min for large extractions
        )
    return _client


# Keywords that indicate a page contains important quote data (high priority)
QUOTE_PAGE_KEYWORDS = [
    "QUOTATION", "QUOTE", "PROPOSAL", "INDICATION",
    "COVERAGE -", "COVERAGE -", "COVERAGE:", "COVERAGE SUMMARY",
    "PREMIUM BREAKDOWN", "PREMIUM SUMMARY", "TOTAL PREMIUM",
    "Total Cost of Policy", "Total General Liability Premium",
    "Total Property Premium", "Total Umbrella Premium",
    "LIMITS OF INSURANCE", "LIMITS OF LIABILITY",
    "GENERAL LIABILITY LIMITS", "PROPERTY LIMITS",
    "DEDUCTIBLE", "DEDUCTIBLES",
    "SCHEDULE OF LOCATIONS", "SCHEDULE OF VALUES",
    "INSURED INFORMATION", "NAMED INSURED",
    "EFFECTIVE DATE", "EXPIRATION DATE",
    "AGENCY INFORMATION", "BROKER",
    "PREMIUM:", "RATE:", "EXPOSURE:",
    "DECLARATIONS", "DECLARATION PAGE",
    "COMMERCIAL GENERAL LIABILITY DECLARATIONS",
    "COMMERCIAL PROPERTY DECLARATIONS",
    "WORKERS COMPENSATION DECLARATIONS",
    "COMMERCIAL AUTO DECLARATIONS",
    "UMBRELLA DECLARATIONS",
    "PREMIUM BREAKDOWN",
    "SCHEDULE OF FORMS",
    "SUBJECTIVITIES", "SUBJECTIVES",
    "CONDITIONS & SUBJECTIVES", "CONDITIONS AND SUBJECTIVES",
    "BINDING CONDITIONS", "BINDING REQUIREMENTS",
    "BINDING SUBJECTIVITIES",
    "CLASS CODE", "CLASSIFICATION",
    "OPTIONAL COVERAGES",
    "ADDITIONAL COVERAGES",
    "SUBLIMITS",
    "SUBLIMITS OF LIABILITY",
    "EXTENSIONS OF COVERAGE",
    "COVERAGE EXTENSIONS",
    "FORMS SCHEDULE",
    "ENDORSEMENT SCHEDULE",
    "FORMS AND ENDORSEMENTS",
    "UNDERLYING INSURANCE",
    "TOWER STRUCTURE",
    "RATING BASIS",
    "PAYROLL",
    "INSURING CLAUSE",
    "INSURING AGREEMENT",
    "EMPLOYEE THEFT",
    "FORGERY OR ALTERATION",
    "SOCIAL ENGINEERING",
    "COMPUTER AND FUNDS TRANSFER",
    "FIDELITY",
    "CRIME COVERAGE",
    "FOREFRONT",
]

# Keywords that indicate boilerplate forms/endorsements (low priority)
BOILERPLATE_KEYWORDS = [
    "THIS ENDORSEMENT CHANGES THE POLICY",
    "PLEASE READ IT CAREFULLY",
    "THIS ENDORSEMENT MODIFIES INSURANCE",
    "COMMON POLICY CONDITIONS",
    "COMMERCIAL GENERAL LIABILITY COVERAGE FORM",
    "COMMERCIAL PROPERTY CONDITIONS",
    "COMMERCIAL PROPERTY COVERAGE FORM",
    "CAUSES OF LOSS",
    "TERRORISM RISK INSURANCE ACT",
    "NUCLEAR HAZARD EXCLUSION",
    "EXCLUSION OF TERRORISM",
    "POLICYHOLDER DISCLOSURE",
    "NOTICE OF TERRORISM",
    "Section 102(1) of the Act",
    "means activities against persons",
    "intimidate or coerce a government",
    "Â© Insurance Services Office",
    "Â© ISO Properties",
    "Includes copyrighted material",
]


def _score_page(page_text: str) -> float:
    """
    Score a PDF page based on how likely it contains important quote data.
    Higher score = more important.
    """
    text_upper = page_text.upper()
    score = 0.0

    # CRITICAL PAGES: These must ALWAYS be included - give maximum boost
    # These are schedule pages that contain essential structured data
    _critical_page_keywords = [
        "ADDITIONAL NAMED INSURED",
        "SCHEDULE OF LOCATIONS",
        "SCHEDULE OF CLASSES",
        "SCHEDULE OF VALUES",
        "SCHEDULE OF HAZARDS",
        "PREMIUM DETAIL",
        "PREMIUM BREAKDOWN",
        "TOTAL COST OF POLICY",
        "SCHEDULE OF UNDERLYING",
        "DESIGNATED PREMISES",
        "PREMISES AND BUILDINGS",
        "LOCATION 1",
        "LOCATION 2",
        "LOCATION 3",
        "BLDG#",
        "BUILDING DESCRIPTION",
        "ADDITIONAL COVERAGES INCLUDED",
        "POLICY COVERAGES",
        "EXCESS LIABILITY POLICY",
        "ATTACHING EXCESS",
        "UNDERLYING LIMITS",
        "UNDERLYING INSURANCE",
        "ATTACHMENT POINT",
        "FOLLOWING FORM",
        "EXCESS OF $",
        "EXCESS LIABILITY QUOTATION",
        "LIMITS OF LIABILITY",
    ]
    is_critical = False
    for kw in _critical_page_keywords:
        if kw in text_upper:
            score += 50.0  # Massive boost - these pages are NEVER dropped
            is_critical = True
            break

    # Positive signals: quote/summary content
    for keyword in QUOTE_PAGE_KEYWORDS:
        if keyword.upper() in text_upper:
            score += 2.0

    # Strong positive: contains dollar amounts with commas (premium figures)
    dollar_amounts = re.findall(r'\$\s*[\d,]+(?:\.\d{2})?', page_text)
    if dollar_amounts:
        score += min(len(dollar_amounts) * 0.5, 5.0)

    # Strong positive: contains percentage rates
    rates = re.findall(r'\d+\.\d{2,4}\s*%?', page_text)
    if rates:
        score += min(len(rates) * 0.3, 3.0)

    # Detect if this is a forms SCHEDULE/LIST page (lists form numbers + descriptions)
    # These pages are HIGH VALUE - they list the forms attached to the policy
    # Do NOT penalize them with boilerplate keywords
    is_forms_schedule = any(kw in text_upper for kw in [
        "FORMS SCHEDULE", "ENDORSEMENT SCHEDULE", "FORMS AND ENDORSEMENTS",
        "SCHEDULE OF FORMS", "FORMS & EXCLUSIONS APPLICABLE",
        "FORMS APPLICABLE", "ENDORSEMENTS APPLICABLE",
        "POLICY FORMS AND ENDORSEMENTS",
    ])
    # Also detect if page has many form numbers (e.g., CP 00 10, CG 00 01, etc.)
    form_numbers = re.findall(r'[A-Z]{2,4}\s*\d{2,4}\s+\d{2,4}', page_text)
    # Also detect NASC/NXLL style form numbers (e.g., NASC 0002 08 09, NXLL 110)
    nasc_forms = re.findall(r'(?:NASC|NXLL|CSXC|CSIP)\s*\d{3,4}', page_text)
    all_form_count = len(form_numbers) + len(nasc_forms)
    if all_form_count >= 3:
        is_forms_schedule = True
        score += 10.0  # High boost for pages with form numbers - these are critical
    elif all_form_count >= 1:
        score += 3.0  # Moderate boost for pages with at least one form number

    # Negative signals: boilerplate forms (but NOT forms schedule pages or critical pages)
    if not is_forms_schedule and not is_critical:
        for keyword in BOILERPLATE_KEYWORDS:
            if keyword.upper() in text_upper:
                score -= 3.0

    # Negative: very long pages with mostly prose (forms text)
    # But NOT forms schedule pages or critical pages which are tabular
    if not is_forms_schedule and not is_critical and len(page_text) > 3000 and score < 2:
        # Check if it's mostly prose (few numbers, lots of text)
        num_count = len(re.findall(r'\d+', page_text))
        word_count = len(page_text.split())
        if word_count > 0 and num_count / word_count < 0.05:
            score -= 2.0

    return score


def extract_text_from_pdf_smart(pdf_path: str, max_chars: int = 100000) -> str:
    """
    Smart PDF text extraction that prioritizes quote summary pages
    over forms/endorsements boilerplate.

    Extracts text page-by-page, scores each page, and returns the
    highest-scoring pages up to max_chars.
    """
    try:
        # Get total page count
        result = subprocess.run(
            ["pdfinfo", pdf_path],
            capture_output=True, text=True, timeout=30
        )
        pages_match = re.search(r"Pages:\s+(\d+)", result.stdout)
        total_pages = int(pages_match.group(1)) if pages_match else 0

        if total_pages == 0:
            logger.warning(f"pdfinfo returned 0 pages, falling back to pdfplumber")
            return _extract_with_pdfplumber(pdf_path, max_chars)
        
        # Log file size for debugging
        file_size = os.path.getsize(pdf_path)
        logger.info(f"PDF file size: {file_size} bytes")

        logger.info(f"PDF has {total_pages} pages, extracting page-by-page for scoring")

        # Extract text page by page
        page_texts = []
        for page_num in range(1, total_pages + 1):
            result = subprocess.run(
                ["pdftotext", "-layout", "-f", str(page_num), "-l", str(page_num), pdf_path, "-"],
                capture_output=True, text=True, timeout=30
            )
            if result.returncode == 0:
                page_text = result.stdout.strip()
                if page_text:
                    score = _score_page(page_text)
                    page_texts.append({
                        "page": page_num,
                        "text": page_text,
                        "score": score,
                        "chars": len(page_text)
                    })

        # --- OCR fallback for image-only pages AND thin-critical pages ---
        # Thin-critical: < 500 chars of pdftotext but mentions a tabular keyword
        # (e.g. Tower Hill PREMISES AND BUILDINGS where headers extract but the Bldg# table is image-based)
        if page_texts and HAS_PDF2IMAGE:
            text_pages_map = {p["page"]: p for p in page_texts}
            missing_pages = [p for p in range(1, total_pages + 1) if p not in text_pages_map]
            _thin_crit_kw = (
                "PREMISES AND BUILDINGS", "SCHEDULE OF LOCATIONS", "SCHEDULE OF VALUES",
                "BLDG#", "BUILDING DESCRIPTION", "BUILDING LIMIT", "TOTAL INSURED VALUE",
                "COVERAGE BY LOCATION", "DESIGNATED PREMISES", "LOCATION SCHEDULE",
                "ADDITIONAL COVERAGES INCLUDED", "POLICY COVERAGES",
            )
            thin_pages = []
            for _p in page_texts:
                if _p["chars"] < 500:
                    _up = _p["text"].upper()
                    if any(kw in _up for kw in _thin_crit_kw):
                        thin_pages.append(_p["page"])
            ocr_targets = sorted(set(missing_pages + thin_pages))
            if ocr_targets and len(ocr_targets) <= 10:
                logger.info(f"OCR: {len(missing_pages)} image-only + {len(thin_pages)} thin-critical pages = {ocr_targets}")
                try:
                    import base64
                    from io import BytesIO
                    for pg in ocr_targets:
                        try:
                            imgs = convert_from_path(pdf_path, dpi=200, first_page=pg, last_page=pg)
                            if not imgs:
                                continue
                            buffered = BytesIO()
                            imgs[0].save(buffered, format="JPEG", quality=85)
                            img_b64 = base64.b64encode(buffered.getvalue()).decode()
                            ocr_resp = _get_openai_client().chat.completions.create(
                                model="gpt-4.1-mini",
                                messages=[
                                    {"role": "system", "content": "Extract ALL text visible in this image, preserving tables, columns, and layout."},
                                    {"role": "user", "content": [
                                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}}
                                    ]}
                                ],
                                temperature=0.0,
                                max_tokens=8000
                            )
                            ocr_text = ocr_resp.choices[0].message.content.strip()
                            if ocr_text and len(ocr_text) > 50:
                                if pg in text_pages_map:
                                    _ent = text_pages_map[pg]
                                    _merged = _ent["text"] + "\n" + ocr_text
                                    _ent["text"] = _merged
                                    _ent["chars"] = len(_merged)
                                    _ent["score"] = _score_page(_merged) + 20
                                    logger.info(f"OCR (thin) page {pg}: merged to {len(_merged)} chars, score={_ent['score']}")
                                else:
                                    score = _score_page(ocr_text)
                                    page_texts.append({
                                        "page": pg,
                                        "text": ocr_text,
                                        "score": score + 20,
                                        "chars": len(ocr_text)
                                    })
                                    logger.info(f"OCR page {pg}: {len(ocr_text)} chars, score={score + 20}")
                        except Exception as e:
                            logger.error(f"OCR failed for page {pg}: {e}")
                except Exception as e:
                    logger.error(f"OCR fallback setup failed: {e}")

        if not page_texts:
            logger.warning("No text extracted from any pages via pdftotext, trying pdfplumber fallback")
            return _extract_with_pdfplumber(pdf_path, max_chars)

        # Sort by score (highest first), then by page number for ties
        page_texts.sort(key=lambda x: (-x["score"], x["page"]))

        # Log the top and bottom scored pages
        logger.info(f"Page scoring results ({len(page_texts)} pages with text):")
        for p in page_texts[:10]:
            logger.info(f"  Page {p['page']}: score={p['score']:.1f}, chars={p['chars']}")
        if len(page_texts) > 10:
            logger.info(f"  ... {len(page_texts) - 10} more pages")
            for p in page_texts[-3:]:
                logger.info(f"  Page {p['page']}: score={p['score']:.1f}, chars={p['chars']} (lowest)")

        # Select pages up to max_chars, prioritizing high-score pages
        selected_pages = []
        total_chars = 0
        for p in page_texts:
            if total_chars + p["chars"] > max_chars:
                # If we haven't selected any pages yet, take at least the first one
                if not selected_pages:
                    selected_pages.append(p)
                break
            selected_pages.append(p)
            total_chars += p["chars"]

        # Re-sort selected pages by page number for coherent reading order
        selected_pages.sort(key=lambda x: x["page"])

        logger.info(
            f"Selected {len(selected_pages)} of {len(page_texts)} pages "
            f"({total_chars} chars) for GPT extraction"
        )

        # Combine selected pages with page markers
        parts = []
        for p in selected_pages:
            parts.append(f"\n--- Page {p['page']} ---\n{p['text']}")

        return "\n".join(parts)

    except Exception as e:
        logger.error(f"Smart PDF extraction failed, falling back to pdfplumber: {e}")
        return _extract_with_pdfplumber(pdf_path, max_chars)


def _extract_with_ocr(pdf_path: str, total_pages: int = 0, max_pages: int = 20) -> str:
    """OCR fallback for scanned/image-based PDFs using GPT Vision.
    
    Converts PDF pages to images and sends them to GPT-4.1-mini with vision
    to extract text content. Used when both pdftotext and pdfplumber
    fail to extract any text.
    
    Optimized for speed: low DPI, limited pages, fast model, single API call.
    """
    if not HAS_PDF2IMAGE:
        logger.warning("pdf2image not available, cannot perform OCR")
        return ""
    
    import base64
    from io import BytesIO
    
    try:
        logger.info(f"OCR fallback: converting PDF pages to images for {pdf_path}")
        
        # Limit pages aggressively - key quote info is in first few pages
        pages_to_convert = min(total_pages, max_pages) if total_pages > 0 else max_pages
        images = convert_from_path(
            pdf_path,
            first_page=1,
            last_page=pages_to_convert,
            dpi=150,  # Lower DPI for speed and memory
            fmt="jpeg"
        )
        
        if not images:
            logger.warning("OCR: No images generated from PDF")
            return ""
        
        logger.info(f"OCR: converted {len(images)} pages to images")
        
        # Build all image content for a single GPT Vision call
        content_parts = [
            {"type": "text", "text": f"Extract ALL text from these {len(images)} insurance document page(s). "
             "Include all numbers, dates, policy details, premiums, limits, carrier names, and coverage information. "
             "Return the text content of each page separated by '--- Page X ---' markers."}
        ]
        
        for i, img in enumerate(images):
            # Convert PIL image to base64 JPEG with moderate quality
            buffer = BytesIO()
            img.save(buffer, format="JPEG", quality=70)
            img_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
            
            content_parts.append({
                "type": "image_url",
                "image_url": {
                    "url": f"data:image/jpeg;base64,{img_base64}",
                    "detail": "low"  # Use low detail for speed
                }
            })
            
            # Free memory immediately
            del img
            buffer.close()
        
        # Free the images list
        del images
        
        try:
            response = _get_openai_client().chat.completions.create(
                model="gpt-4.1-mini",  # Faster model for OCR
                messages=[
                    {"role": "system", "content": "You are an expert OCR assistant. Extract all text from the provided insurance document images accurately and completely."},
                    {"role": "user", "content": content_parts}
                ],
                max_tokens=16000,
                temperature=0.1
            )
            
            result_text = response.choices[0].message.content
            if result_text:
                logger.info(f"OCR: extracted {len(result_text)} chars from {pages_to_convert} pages")
                return result_text
            else:
                logger.warning("OCR: GPT Vision returned empty response")
                return ""
                
        except Exception as e:
            logger.error(f"OCR GPT Vision call failed: {e}")
            return ""
        
    except Exception as e:
        logger.error(f"OCR extraction failed: {e}")
        return ""


def _extract_with_pdfplumber(pdf_path: str, max_chars: int = 120000) -> str:
    """Fallback PDF extraction using pdfplumber (pure Python, no system deps)."""
    if not HAS_PDFPLUMBER:
        logger.warning("pdfplumber not available, falling back to basic pdftotext")
        return extract_text_from_pdf(pdf_path)
    
    try:
        logger.info(f"Using pdfplumber for extraction: {pdf_path}")
        file_size = os.path.getsize(pdf_path)
        logger.info(f"PDF file size: {file_size} bytes")
        
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            logger.info(f"pdfplumber: PDF has {total_pages} pages")
            
            # Extract and score pages
            page_texts = []
            for i, page in enumerate(pdf.pages):
                try:
                    text = page.extract_text() or ""
                    if text.strip():
                        score = _score_page(text)
                        page_texts.append({
                            "page": i + 1,
                            "text": text,
                            "score": score,
                            "chars": len(text)
                        })
                except Exception as e:
                    logger.warning(f"pdfplumber: error on page {i+1}: {e}")
                    continue
            
            if not page_texts:
                logger.warning("pdfplumber: No text extracted from any pages, trying OCR fallback")
                return _extract_with_ocr(pdf_path, total_pages)
            
            logger.info(f"pdfplumber: extracted text from {len(page_texts)} of {total_pages} pages")
            
            # Sort by score (highest first)
            page_texts.sort(key=lambda x: (-x["score"], x["page"]))
            
            # Log top scored pages
            for p in page_texts[:10]:
                logger.info(f"  Page {p['page']}: score={p['score']:.1f}, chars={p['chars']}")
            
            # Select pages up to max_chars
            selected_pages = []
            total_chars = 0
            for p in page_texts:
                if total_chars + p["chars"] > max_chars:
                    if not selected_pages:
                        selected_pages.append(p)
                    break
                selected_pages.append(p)
                total_chars += p["chars"]
            
            # Re-sort by page number
            selected_pages.sort(key=lambda x: x["page"])
            
            logger.info(f"pdfplumber: selected {len(selected_pages)} pages ({total_chars} chars)")
            
            parts = []
            for p in selected_pages:
                parts.append(f"\n--- Page {p['page']} ---\n{p['text']}")
            
            return "\n".join(parts)
    
    except Exception as e:
        logger.error(f"pdfplumber extraction failed: {e}")
        return extract_text_from_pdf(pdf_path)


def extract_text_from_pdf(pdf_path: str) -> str:
    """Extract text from a PDF file using pdftotext (basic full extraction)."""
    try:
        result = subprocess.run(
            ["pdftotext", "-layout", pdf_path, "-"],
            capture_output=True, text=True, timeout=60
        )
        if result.returncode == 0 and result.stdout.strip():
            logger.info(f"Extracted {len(result.stdout)} chars from PDF: {pdf_path}")
            return result.stdout
        else:
            # Fallback: try without layout flag
            result = subprocess.run(
                ["pdftotext", pdf_path, "-"],
                capture_output=True, text=True, timeout=60
            )
            if result.returncode == 0:
                logger.info(f"Extracted {len(result.stdout)} chars from PDF (no layout): {pdf_path}")
                return result.stdout
            logger.error(f"pdftotext failed: {result.stderr}")
            return ""
    except Exception as e:
        logger.error(f"PDF extraction error: {e}")
        return ""


def extract_text_from_excel(excel_path: str) -> str:
    """Extract data from an Excel file (SOV or schedule)."""
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        all_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_text.append(f"\n=== Sheet: {sheet_name} ===\n")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    all_text.append(" | ".join(cells))
        text = "\n".join(all_text)
        logger.info(f"Extracted {len(text)} chars from Excel: {excel_path}")
        return text
    except Exception as e:
        logger.error(f"Excel extraction error: {e}")
        return ""


def extract_document(file_path: str) -> str:
    """Extract text from a document based on its file extension."""
    ext = Path(file_path).suffix.lower()
    if ext == ".pdf":
        return extract_text_from_pdf(file_path)
    elif ext in (".xlsx", ".xls"):
        return extract_text_from_excel(file_path)
    else:
        logger.warning(f"Unsupported file type: {ext}")
        return ""


def extract_document_smart(file_path: str) -> str:
    """Extract text from a document using smart extraction for PDFs."""
    ext = Path(file_path).suffix.lower()
    if ext == ".pdf":
        return extract_text_from_pdf_smart(file_path)
    elif ext in (".xlsx", ".xls"):
        return extract_text_from_excel(file_path)
    else:
        logger.warning(f"Unsupported file type: {ext}")
        return ""


EXTRACTION_SYSTEM_PROMPT = """You are an expert insurance data extraction assistant specializing in hotel and hospitality insurance. You extract structured data from insurance quote documents.

CRITICAL RULES:
1. Extract EVERY form and endorsement number with full description and date
2. NEVER summarize - list everything exactly as shown
3. NEVER write "Additional forms as listed in policy"
4. Extract additional coverages ONLY if they literally appear in the quote. If a peril or coverage is not mentioned, OMIT it from additional_coverages - do NOT invent a row. Only use "Excluded" or "NOT COVERED" when the document literally contains those exact words, a $0 limit, or an explicit exclusion.
5. For Property Flood and Earthquake: include ONLY if the quote explicitly lists them with a limit, "Excluded", "Not Covered", or a $0 value. If the quote is silent on them, OMIT (do not assume excluded and do not emit a row).
6. Include ALL taxes, fees, surcharges in premium calculations
7. Exclude TRIA premiums from totals
8. Extract ALL deductibles
8a. Do NOT include deductibles for perils that are "NOT COVERED" in the sublimits section (e.g., if Named Windstorm sublimit is "NOT COVERED", do NOT extract a Named Windstorm/Named Storm deductible)
9. Extract ALL limits
10. Note carrier name and whether admitted or non-admitted
11. Property additional_coverages is a LITERAL PASS-THROUGH of the "Additional Coverages", "Policy Coverages", "Sublimits", or "Extensions of Coverage" section of the quote. Extract EVERY line item literally present with its exact limit - do NOT add line items that are not in the document, and do NOT drop any that are. Business Income w/ Extra Expense, Ordinance or Law (A/B/C), Equipment Breakdown components (Data Restoration, Hazardous Substances, Spoilage), Debris Removal, Pollutant Cleanup, Valuable Papers, Accounts Receivable, Water Backup, Outdoor Signs, etc. must flow through exactly as written when present, and must be OMITTED when not present.
12. Property forms_endorsements: HARD RULE - emit ONLY form numbers that LITERALLY appear on a dedicated "Form Schedule", "Schedule of Forms", "Schedule of Forms and Endorsements", or "Forms and Endorsements" page within the PROPERTY section of this quote. You are FORBIDDEN from generating forms from training data, generic ISO boilerplate (CP DS 00, CP 00 90, CP 99 03, CP 17 96, CP 17 97, CP 04 21, CP 01 54, IL 00 21, IL 09 99, etc.), or carrier templates. If no such schedule page is found in the PROPERTY section, return "forms_endorsements": [] and add a string to warnings: "Property forms schedule page not found - forms intentionally empty". DO NOT GUESS. DO NOT SYNTHESIZE.
13. GL forms_endorsements: HARD RULE - emit ONLY form numbers that LITERALLY appear on a dedicated forms schedule page within the GL section. If absent, return "forms_endorsements": [] and add a warning. DO NOT synthesize or fill in from training data.
14. GL designated_premises is MANDATORY when CG2144/NXLL110 form exists - extract ALL addresses
15. GL schedule_of_classes MUST include ALL class codes with actual dollar exposure amounts
16. Named insureds MUST be exact legal entity names from the quote - do NOT concatenate hotel brand names into entity names
17. For carrier names: Use the ISSUING carrier name (e.g., "Associated Industries Insurance Company" or "Technology Insurance Company" for AmTrust policies, "Palms Insurance Company" for Palms). Do NOT use the wholesale broker name as the carrier. For RT Specialty quotes: RT Specialty (RSG Specialty) is a WHOLESALE BROKER, NOT the carrier. The issuing carrier is typically "Certain Underwriters at Lloyd's, London" placed through the Owners First Association program. Use "Certain Underwriters at Lloyd's, London" as the carrier name. Similarly, CRC, Amwins, Burns & Wilcox are wholesale brokers — always use the actual issuing carrier.
18. Property coinsurance: emit ONE entry per coverage that LITERALLY has a coinsurance value, monthly limitation, or valuation label in the quote. If the quote shows "Business Income with Extra Expense - 1/6", emit {"coverage": "Business Income", "limitation": "1/6"}. If a coverage has NO coinsurance value shown in the quote, OMIT it - do NOT emit "NOT SHOWN", "N/A", "Blank", or placeholder text.
19. For LAYERED property programs with multiple carriers: Use "property" for the primary layer, "excess_property" for the first excess layer, and "excess_property_2" for the second excess layer. Each layer has its own carrier, limits, deductibles, forms, coinsurance, and subjectivities.
20. COMPETING / ALTERNATIVE QUOTES FOR THE SAME COVERAGE TYPE: When the uploaded documents contain quotes from DIFFERENT carriers for the SAME coverage type (e.g., two separate property quotes from Starr and Markel, or two GL quotes from different carriers), you MUST extract ALL of them. Use the base key for the first quote (e.g., "property") and append "_alt_1", "_alt_2" etc. for additional competing quotes of the same type (e.g., "property_alt_1", "general_liability_alt_1"). Each alternative quote gets its own full coverage entry with carrier, premium, limits, forms, subjectivities, etc. - identical structure to the primary. IMPORTANT: Do NOT discard or merge competing quotes. If two different carriers each provide a property quote, both must appear in the output. Look at the FILE headers to identify separate quote documents from different carriers.
    a) MULTIPLE GL POLICIES FOR DIFFERENT INSUREDS IN ONE PROGRAM: When two GL quotes are uploaded and the Named Insured differs between them (e.g., one carrier issues GL to the management company 'ABC Hospitality Management, Inc.' and another issues GL to a property-owning entity like 'Premier Hotels, LLC' for a specific location), treat this as TWO parallel GL policies that will BOTH bind. Extract the first as "general_liability" and the second as "general_liability_alt_1". Do NOT drop either one just because the insureds don't match the 'primary' named insured. Do NOT merge their limits or premiums into a single entry. Each policy retains its own carrier, premium, limits, forms, and named insured.
21. Property coverage_by_location is MANDATORY for multi-location/multi-building quotes. When the quote shows a "COVERAGE SUMMARY" or "DESCRIPTION OF PREMISES" table with per-premise/per-building values (Building, BPP, Business Income amounts per location), extract EACH row into coverage_by_location. The "limits" array should contain the TOTAL/COMBINED values, while coverage_by_location has the per-location breakdown. Do NOT combine per-location values into single limits.

Return your extraction as a JSON object with the following structure. Only include sections that are present in the documents."""

EXTRACTION_USER_PROMPT = """Extract ALL insurance data from the following quote document(s). Return a JSON object.

The JSON structure should be:
{{
  "client_info": {{
    "named_insured": "Full legal name",
    "dba": "DBA name if any",
    "address": "Full address",
    "entity_type": "LLC/Corp/etc",
    "effective_date": "MM/DD/YYYY",
    "expiration_date": "MM/DD/YYYY",
    "sales_exposure_basis": "Revenue/payroll amount if shown"
  }},
  "coverages": {{
    "property": {{
      "carrier": "Carrier name",
      "carrier_admitted": true or false,
      "am_best_rating": "A+ XV or similar",
      "premium": 0,
      "taxes_fees": 0,
      "total_premium": 0,
      "tria_premium": 0,
      "tiv": "$X Total Insured Value from quote or SOV (e.g., $56,390,000)",
      "limits": [
        {{"description": "Building", "limit": "$X"}},
        {{"description": "Business Personal Property", "limit": "$X"}},
        {{"description": "Business Income", "limit": "ALS or $X"}}
      ],
      "coverage_by_location": [
        {{"premise": 1, "building": 1, "address": "Street Address, City, ST ZIP", "building_value": "$X", "bpp_value": "$X", "business_income": "$X or ALS", "other_values": {{"Outside Signs": "$X"}}, "cause_of_loss": "Special or Basic", "coinsurance": "80%", "valuation": "RC or ACV"}}
      ],
      "deductibles": [
        {{"description": "All Other Perils", "amount": "$X"}},
        {{"description": "Named Storm", "amount": "X% or $X"}},
        {{"description": "Wind/Hail", "amount": "$X"}}
      ],
      "additional_coverages": [
        {{"description": "Line item exactly as written in quote (e.g., Business Income with Extra Expense Coverage - 1/6)", "limit": "exact limit from quote (e.g., $2,100,000)"}}
      ],
      "coinsurance": [
        {{"coverage": "Exact coverage label from quote (e.g., Building, Business Personal Property, Business Income)", "percentage": "exact percentage from quote if shown (e.g., 80%)", "limitation": "exact monthly limitation from quote if shown (e.g., 1/6)"}}
      ],
      "valuation": "Replacement Cost or Actual Cash Value or Agreed Value",
      "forms_endorsements": [
        {{"form_number": "CP 00 10 06/07", "description": "Building and Personal Property Coverage Form"}}
      ],
      "subjectivities": ["List of binding requirements"]
    }},
    "excess_property": {{
      "carrier": "Excess property layer 1 carrier name (e.g., Kinsale)",
      "carrier_admitted": true or false,
      "am_best_rating": "A+ XV or similar",
      "premium": 0,
      "taxes_fees": 0,
      "total_premium": 0,
      "tria_premium": 0,
      "layer_description": "$Xm xs $Xm (e.g., $10,000,000 xs $10,000,000)",
      "tiv": "$X Total Insured Value",
      "limits": [
        {{"description": "Per Occurrence", "limit": "$X"}},
        {{"description": "Excess Of", "limit": "$X"}}
      ],
      "deductibles": [
        {{"description": "All Other Perils", "amount": "As per underlying or $X"}},
        {{"description": "Named Storm", "amount": "As per underlying or $X"}}
      ],
      "coinsurance": [
        {{"coverage": "Building", "percentage": "0% or N/A or per underlying"}}
      ],
      "forms_endorsements": [
        {{"form_number": "XPF1000-1224", "description": "Excess Property Insurance Policy Declarations"}}
      ],
      "subjectivities": ["List of binding requirements"]
    }},
    "excess_property_2": {{
      "carrier": "Excess property layer 2 carrier name (e.g., Gotham via Coaction)",
      "carrier_admitted": true or false,
      "am_best_rating": "A+ XV or similar",
      "premium": 0,
      "taxes_fees": 0,
      "total_premium": 0,
      "tria_premium": 0,
      "layer_description": "$Xm xs $Xm (e.g., $10,050,000 xs $20,000,000)",
      "tiv": "$X Total Insured Value",
      "limits": [
        {{"description": "Per Occurrence", "limit": "$X"}},
        {{"description": "Excess Of", "limit": "$X"}}
      ],
      "deductibles": [
        {{"description": "All Other Perils", "amount": "As per underlying or $X"}},
        {{"description": "Named Storm", "amount": "As per underlying or $X"}}
      ],
      "coinsurance": [
        {{"coverage": "Building", "percentage": "0% or N/A or per underlying"}}
      ],
      "forms_endorsements": [
        {{"form_number": "PN049937", "description": "How to Report a Claim"}}
      ],
      "subjectivities": ["List of binding requirements"]
    }},
    "general_liability": {{
      "carrier": "Carrier name",
      "carrier_admitted": true or false,
      "am_best_rating": "Rating",
      "premium": 0,
      "taxes_fees": 0,
      "total_premium": 0,
      "tria_premium": 0,
      "gl_deductible": "$X Per Occurrence (or $0 if none)",
      "defense_basis": "In Addition to Limits or Within Limits",
      "limits": [
        {{"description": "Each Occurrence", "limit": "$X"}},
        {{"description": "General Aggregate", "limit": "$X"}},
        {{"description": "Products/Completed Operations", "limit": "$X"}},
        {{"description": "Personal & Advertising Injury", "limit": "$X"}},
        {{"description": "Damage to Rented Premises", "limit": "$X"}},
        {{"description": "Medical Payments", "limit": "$X"}},
                    {{"description": "Employee Benefits - Each Claim", "limit": "$X"}},
                    {{"description": "Employee Benefits - Aggregate", "limit": "$X"}},
                    {{"description": "Sexual Abuse - Each Act", "limit": "$X"}},
                    {{"description": "Sexual Abuse - Aggregate", "limit": "$X"}},
                    {{"description": "Hired & Non-Owned Auto", "limit": "$X"}},
                    {{"description": "Assault and Battery - Each Event", "limit": "$X"}},
                    {{"description": "Assault and Battery - Aggregate", "limit": "$X"}}
      ],
      "aggregate_applies": "Per Location or Per Policy",
                "total_sales": "$X (total gross sales from rate basis line, e.g. $1,151,719)",
      "schedule_of_classes": [
        {{"location": "Loc 1", "address": "Street Address", "brand_dba": "Hotel Brand or DBA Name", "classification": "Hotels/Motels", "class_code": "XXXXX", "rate": "Rate per $100 or flat amount", "exposure_basis": "Sales/Revenue/Area/Units", "exposure": "$X or number", "premium": "$X"}}
      ],
      "additional_coverages": [
        {{"description": "Coverage name", "limit": "$X", "deductible": "$X"}}
      ],
      "designated_premises": [
        "Full address 1 from CG2144/NXLL110 designated premises form",
        "Full address 2"
      ],
      "forms_endorsements": [
        {{"form_number": "CG 00 01 04/13", "description": "Commercial General Liability Coverage Form"}}
      ],
      "subjectivities": []
    }},
    "umbrella": {{
      "carrier": "Primary layer carrier name",
      "carrier_admitted": true or false,
      "am_best_rating": "Rating",
      "premium": 0,
      "taxes_fees": 0,
      "total_premium": 0,
      "tria_premium": 0,
      "limits": [
        {{"description": "Each Occurrence", "limit": "$X"}},
        {{"description": "Aggregate", "limit": "$X"}},
        {{"description": "Self-Insured Retention", "limit": "$X"}}
      ],
      "underlying_insurance": [
        {{"carrier": "Carrier", "coverage": "Auto Liability", "limits": "$X CSL"}},
        {{"carrier": "Carrier", "coverage": "General Liability", "limits": "$X Occ / $X Agg"}}
      ],
      "tower_structure": [
        {{"layer": "Primary", "carrier": "Carrier", "limits": "$5M xs Primary", "premium": 0, "total_cost": 0}}
      ],
      "first_dollar_defense": true,
      "tria_included": true,
      "forms_endorsements": [],
      "subjectivities": []
    }},
    "umbrella_layer_2": {{
      "carrier": "Second layer carrier name (if a second excess layer exists)",
      "carrier_admitted": true or false,
      "am_best_rating": "Rating",
      "premium": 0,
      "taxes_fees": 0,
      "total_premium": 0,
      "limits": [
        {{"description": "Each Occurrence", "limit": "$X"}},
        {{"description": "Aggregate", "limit": "$X"}}
      ],
      "tower_structure": [
        {{"layer": "2nd Excess", "carrier": "Carrier", "limits": "$5M xs $5M", "premium": 0, "total_cost": 0}}
      ],
      "forms_endorsements": [],
      "subjectivities": []
    }},
    "umbrella_layer_3": {{
      "carrier": "Third layer carrier name (if a third excess layer exists)",
      "carrier_admitted": true or false,
      "am_best_rating": "Rating",
      "premium": 0,
      "taxes_fees": 0,
      "total_premium": 0,
      "limits": [
        {{"description": "Each Occurrence", "limit": "$X"}},
        {{"description": "Aggregate", "limit": "$X"}}
      ],
      "tower_structure": [
        {{"layer": "3rd Excess", "carrier": "Carrier", "limits": "$5M xs $10M", "premium": 0, "total_cost": 0}}
      ],
      "forms_endorsements": [],
      "subjectivities": []
    }},
    "workers_comp": {{
      "carrier": "Carrier name",
      "carrier_admitted": true or false,
      "am_best_rating": "Rating",
      "premium": 0,
      "taxes_fees": 0,
      "total_premium": 0,
      "limits": [
        {{"description": "Workers Compensation", "limit": "Statutory"}},
        {{"description": "EL - Each Accident", "limit": "$X"}},
        {{"description": "EL - Disease Policy Limit", "limit": "$X"}},
        {{"description": "EL - Disease Each Employee", "limit": "$X"}}
      ],
      "deductible": {{"amount": "$X", "type": "Per Claim or Per Accident"}},
      "rating_basis": [
        {{"state": "XX", "location": "1", "class_code": "XXXX", "classification": "Hotels", "payroll": "$X", "rate": "X.XX"}}
      ],
      "forms_endorsements": [],
      "subjectivities": []
    }},
    "commercial_auto": {{
      "carrier": "Carrier name",
      "carrier_admitted": true or false,
      "am_best_rating": "Rating",
      "premium": 0,
      "taxes_fees": 0,
      "total_premium": 0,
      "limits": [
        {{"description": "Liability CSL", "limit": "$X"}},
        {{"description": "Uninsured Motorist", "limit": "$X"}},
        {{"description": "Medical Payments", "limit": "$X"}}
      ],
      "vehicle_schedule": [
        {{"year": "XXXX", "make": "Make", "model": "Model", "vin": "VIN"}}
      ],
      "forms_endorsements": [],
      "subjectivities": []
    }},
    "terrorism": {{
      "carrier": "Carrier name",
      "carrier_admitted": true or false,
      "am_best_rating": "Rating",
      "premium": 0,
      "taxes_fees": 0,
      "total_premium": 0,
      "limits": [
        {{"description": "Certified Acts of Terrorism", "limit": "$X or Policy Limit"}},
        {{"description": "Non-Certified Acts / Active Assailant", "limit": "$X or N/A"}}
      ],
      "additional_coverages": [
        {{"description": "Coverage name", "limit": "$X"}}
      ],
      "forms_endorsements": [],
      "subjectivities": []
    }},
    "crime": {{
      "carrier": "Carrier name (e.g., Federal Insurance Company for Chubb)",
      "carrier_admitted": true or false,
      "am_best_rating": "Rating",
      "premium": 0,
      "taxes_fees": 0,
      "total_premium": 0,
      "insuring_clauses": [
        {{"clause": "Employee Theft", "limit": "$X", "retention": "$X"}},
        {{"clause": "Forgery or Alteration", "limit": "$X", "retention": "$X"}},
        {{"clause": "Inside the Premises - Theft of Money & Securities", "limit": "$X", "retention": "$X"}},
        {{"clause": "Inside the Premises - Robbery/Safe Burglary of Other Property", "limit": "$X", "retention": "$X"}},
        {{"clause": "Outside the Premises", "limit": "$X", "retention": "$X"}},
        {{"clause": "Computer and Funds Transfer Fraud", "limit": "$X", "retention": "$X"}},
        {{"clause": "Money Orders and Counterfeit Money", "limit": "$X", "retention": "$X"}},
        {{"clause": "Social Engineering Fraud", "limit": "$X", "retention": "$X"}}
      ],
      "limits": [
        {{"description": "Per Loss Limit", "limit": "$X"}},
        {{"description": "Aggregate Limit", "limit": "$X"}}
      ],
      "forms_endorsements": [],
      "subjectivities": []
    }},
    "cyber": {{
      "carrier": "Carrier name",
      "carrier_admitted": true or false,
      "am_best_rating": "Rating",
      "premium": 0,
      "taxes_fees": 0,
      "total_premium": 0,
      "limits": [
        {{"description": "Aggregate Limit", "limit": "$X"}},
        {{"description": "Retention/Deductible", "limit": "$X"}}
      ],
      "additional_coverages": [
        {{"description": "Coverage name", "limit": "$X"}}
      ],
      "forms_endorsements": [],
      "subjectivities": []
    }},
    "epli": {{
      "carrier": "Carrier name",
      "carrier_admitted": true or false,
      "am_best_rating": "Rating",
      "premium": 0,
      "taxes_fees": 0,
      "total_premium": 0,
      "defense_provisions": "Duty to Defend or Non-Duty to Defend",
      "limits": [
        {{"description": "EPL Aggregate Limit of Liability", "limit": "$X"}},
        {{"description": "Third Party Discrimination / Harassment", "limit": "$X or N/A"}},
        {{"description": "Additional Defense Limit", "limit": "$X or N/A"}},
        {{"description": "Retention Per Claim", "limit": "$X"}}
      ],
      "additional_coverages": [
        {{"description": "Sublimit name (e.g. Wage and Hour Defense Costs, Workplace Violence Expenses, Immigration Claim Investigation, WARN Act Defense, Biometric Claims Defense, Employee Privacy Violation Defense)", "limit": "$X"}}
      ],
      "notable_endorsements": [
        {{"description": "Endorsement name", "detail": "Coverage detail (e.g. Yes - absolute language, Excluded, Sublimited)"}}
      ],
      "forms_endorsements": [],
      "subjectivities": []
    }}
  }},
  "named_insureds": [
    {{"name": "Full legal entity name", "dba": "DBA/trade name if shown"}}
  ],
  "additional_named_insureds": [
    {{"name": "Entity name", "dba": "DBA if shown"}}
  ],
  "additional_insureds": [
    {{"name": "Entity or person name", "relationship": "Franchisor/Mortgagee/Manager/etc", "description": "Additional details"}}
  ],
  "additional_interests": [
    {{"type": "Mortgagee/Loss Payee/etc", "name_address": "Full name and address", "description": "Description"}}
  ],
  "locations": [
    {{"number": "1", "name": "Property/Hotel name or DBA", "corporate_entity": "Entity name", "address": "Street", "city": "City", "state": "ST", "zip": "XXXXX", "description": "Hotel/Motel", "tiv": 0}}
  ],
  "expiring_premiums": {{
    "property": 0,
    "excess_property": 0,
    "excess_property_2": 0,
    "general_liability": 0,
    "umbrella": 0,
    "workers_comp": 0,
    "commercial_auto": 0,
    "total": 0
  }},
  "payment_options": [
    {{"carrier": "Carrier", "coverage_type": "Property, General Liability, Umbrella / Excess, Workers Compensation, Crime, Terrorism, Equipment Breakdown, EPLI, Cyber, Flood, Auto", "terms": "Payment terms (exclude commission/broker fee info)", "mep": "Minimum earned premium"}}
  ],
  "warnings": ["String entries describing any value the extractor could not verify (e.g., Property forms schedule page not found - forms intentionally empty)"]
}}

IMPORTANT:
- COVERAGE CLASSIFICATION: A standalone terrorism/TRIA policy is NOT general liability. If a document is from Lloyd's of London, AEGIS, or similar and covers ONLY terrorism/TRIA/certified acts of terrorism/active assailant, classify it as "terrorism" NOT "general_liability". General Liability covers bodily injury, property damage, personal & advertising injury with occurrence/aggregate limits. Terrorism covers certified/non-certified acts of terrorism. If a single policy bundles both, put the terrorism portion in "terrorism" and the GL portion in "general_liability".
- Only include coverage sections that appear in the documents
- Extract EVERY form number and endorsement exactly as written
- Include form dates (e.g., "06/07" in "CP 00 10 06/07")
- For total_premium: This MUST be the all-in out-the-door number — the HIGHEST total shown on the quote. Look for "Total Cost of Policy", "Total Package Cost", "Total Policy Cost", "Total Policy Premium", "Total Due", "Grand Total", "Total Amount Due", "Total Estimated Cost", or any final total line. It includes base premium + broker fees + surplus lines tax + stamping fee + fire marshal tax + inspection fees + FSLSO fees + EMPA surcharge + any other taxes/fees/surcharges. ALWAYS use the HIGHEST total amount shown — this includes broker fees. If no single total line exists, calculate total_premium = premium + taxes_fees. CRITICAL: total_premium must ALWAYS be >= premium. If the quote shows separate line items for taxes and fees, ADD them ALL (including broker fees) to the base premium to get total_premium. For example if GL premium is $8,874, broker fee is $250, surplus lines tax is $430.39, and stamping fee is $3.55, then total_premium = $8,874 + $250 + $430.39 + $3.55 = $9,557.94. NEVER use the base premium as total_premium when taxes/fees exist. PREFERRED: If the quote prints a "Total Cost of Policy" line that includes broker fee, use that EXACT number (with cents) as total_premium. DOUBLE-CHECK: After extracting premium and total_premium, verify that total_premium = premium + ALL individual tax/fee line items (including broker fee). If your calculated total doesn't match the document's stated total, use the document's stated total number — always the HIGHEST one.
- For GL policies that include BOTH General Liability AND Liquor Liability in a single package: The "premium" field should be the combined package premium (GL + Liquor), and "total_premium" should be the Total Package Cost (premium + broker fee + surplus lines tax + stamping fee + all other fees). CRITICAL: Use the EXACT dollar amount printed on the "Total Package Cost" or "Total Cost" line - do NOT manually sum individual line items, as you may miss fees or introduce rounding errors. For example, if the quote shows Total General Liability Premium: $61,899, Total Liquor Liability Premium: $2,200, Total Package Premium: $64,099, Broker Fee: $2,500, Surplus Lines Tax: $3,108.80, Stamping Fee: $25.64, Total Package Cost: $69,733.44 - then premium=64099 and total_premium=69733.44. ALWAYS preserve cents.
- For GL gl_deductible: Extract the per-occurrence deductible if one exists. Look for "Deductible Per Occurrence", "Deductible Liability", "$X,000 Deductible Per Occurrence Including Loss Adjustment Expense", or similar. Include the full description (e.g., "$5,000 Per Occurrence Including Loss Adjustment Expense"). If no GL deductible, set to "$0" or "None".
- For GL defense_basis: Look for "Defense Basis" or whether defense costs are "In Addition to Limits" or "Within Limits of Liability".
- For GL schedule_of_classes: Extract the exposure schedule. This may be location-based OR class-code-based. For class-code-based quotes (like AmTrust), extract each class code entry with: class_code (e.g., "45190"), classification/description, rate (e.g., "9.964" per $100), exposure amount (e.g., "$8,748,612"), and exposure_basis (e.g., "Gross Sales", "Per Acre", "Area", "Liquor Sales", "FLAT"). For location-based quotes, include address, brand_dba, classification, exposure, and premium. Include vacant land, restaurants, liquor, sundry, hired auto, loss control, and all non-hotel entries. Include ALL exposure classes for each location (e.g., Hotel/Motel, Restaurant, Liquor Liability as separate rows). CRITICAL: Always capture the actual dollar amount for exposure (e.g., $8,748,612 not just "Gross Sales"). The exposure_basis describes what the number represents (Gross Sales, Revenue, Area, etc.)
- GL DESIGNATED PREMISES LOCATIONS: If the GL quote includes a form like CG 21 44, CG2144, NXLL110, or any "Limitation of Coverage to Designated Premises" form, extract EVERY location listed in that form. These are ALL the locations covered under the GL policy. Do TWO things: (1) Add each address as a separate schedule_of_classes entry with its full address. (2) Also populate the "designated_premises" array with each full address string exactly as written (e.g., "4285 Highway 51, LaPlace, LA 70068"). The designated_premises array is the AUTHORITATIVE list of GL covered locations. CRITICAL: The CG2144/NXLL110 form typically lists addresses in a numbered format like "1) 4285 Highway 51, LaPlace, LA 70068" followed by "2) 4281 Highway 51..." etc. Extract ALL numbered addresses, not just the first few. There may be 8 or more addresses. Also look for addresses that may appear with labels like "Office:" or "Hotels:" before the numbered list. Extract those too. If the form text is split across multiple pages, combine all addresses from all pages.
- ALWAYS preserve cents in premium amounts (e.g., $60,513.35 not $60,513)
- Mark excluded coverages explicitly
- For Property tiv: Extract the Total Insured Value (TIV) from the property quote or SOV. Look for "Total Insured Value", "TIV", "Total Values", or the sum total on the Schedule of Values. This should be the total of Building + Contents/BPP + Business Income/Rents across all locations. For example if the SOV shows Building Total $42,800,000 + Contents Total $7,700,000 + BI/Rents Total $5,550,000 = TIV $56,390,000. Use the actual SOV/quote total, NOT the per-location coverage limits. TOWER HILL / VANTAGE RISK SPECIFIC: For Tower Hill property quotes, the overall Total Insured Value does NOT appear as a single labeled line. Instead, you MUST sum the "Limit" column across ALL Bldg# rows in the "PREMISES AND BUILDINGS" section. For a single-location Tower Hill quote with one Bldg# that has Building Limit $15,080,000 and Business Personal Property Limit $1,500,000, property.tiv = 16580000. NEVER leave property.tiv at 0 when a Tower Hill Bldg# table is present.
- For locations: Extract ALL property locations. For each location, set "name" to the hotel/property name (e.g., "Hacienda Hotel", "Hampton Inn") and "tiv" to the Total Insured Value for that specific location. For single-location quotes, use the "Account Name", "Applicant", or "Named Insured" as the location name, and the total TIV from the quote (e.g., "Total Insured Values: $4,660,000") as the location tiv. Always extract tiv as a number (no $ or commas).
- TOWER HILL / VANTAGE RISK PROPERTY QUOTES - CRITICAL EXTRACTION PROCEDURE: Tower Hill Insurance Group underwrites on behalf of Vantage Risk Specialty Insurance Company. Their property quotes do NOT use the header "Schedule of Locations" or "Schedule of Values". Instead, locate the section whose heading reads exactly "PREMISES AND BUILDINGS" (may appear as "PREMISES  AND  BUILDINGS" with double spaces). Under that heading you will see (in order): (a) a few "Coverage for All Premises" rows (Employee Dishonesty, Forgery - SKIP these, they are policy-wide limits, not locations), then (b) one or more centered sub-headings of the form "Location 1", "Location 2", etc. Under each "Location N" heading the street address appears on line 1 and "CITY, ST ZIP" appears on line 2. Immediately below is a small table with column headers "Bldg#", "Building Description", and "Limit". Each Bldg# row in that table has the FULL property address in "Building Description" and a dollar "Limit" value (typically one row for Building limit and one for Business Personal Property / BPP). To extract this correctly:
- TOWER HILL PER-PREMISE POLICY COVERAGES AND ADDITIONAL COVERAGES INCLUDED: (1) From the "PREMISES AND BUILDINGS" page, under each "Location N" heading, extract the Bldg# table rows for Building and Business Personal Property into property.limits[]. (2) The "Business Income with Extra Expense Coverage" row (often shown as "- 1/6" meaning a 2-month period) MUST be extracted as an entry in property.limits[] with description "Business Income" and the exact dollar value shown - this is required so TIV sums Building + BPP + Business Income. (3) All OTHER per-premise "Policy Coverages" rows (Ordinance or Law A, Ordinance or Law B/C, Valuable Papers, Accounts Receivable, Debris Removal, Pollutant Removal and Clean Up, Water Backup and Sump Overflow, Outdoor Signs, Fire Department Service Charge, etc.) go into property.additional_coverages[] with the exact description and limit. (4) A separate page titled "Additional Coverages Included" lists the Tower Hill form-included coverages (often 15 to 25 line items) - pass through EVERY line on that page into property.additional_coverages[] exactly as written, with the description and any limit shown. (5) NEVER put Southlake or GL-only optional-coverage packages (Breach Response, EPLI, Enviro Pack, Workplace Violence, Hired and Non-Owned Auto) into property.additional_coverages[] - those are GL coverages and belong only in general_liability.additional_coverages[].
    1. For locations[N].name: If the quote shows a hotel/DBA name (e.g., "Hilton Garden Inn Winter Park"), use that; otherwise use the corporate entity or franchise flag. Do NOT default to the LLC name when a franchise DBA is available.
    2. For locations[N].address / city / state / zip: Parse the two lines under the "Location N" heading - line 1 = street address, line 2 = "CITY, ST ZIP" (split on the comma and the last space). You may also confirm with the "Building Description" column.
    3. For locations[N].tiv: Sum the "Limit" column across ALL Bldg# rows that belong to this Location N (typically Building + BPP). Example: if Bldg# 1 shows Building $15,080,000 and BPP $1,500,000, locations[0].tiv = 16580000.
    4. For coverages.property.tiv at the policy level: Sum the limits across EVERY Bldg# row under EVERY Location N (this equals the total insured value for the property policy). For a single-location Tower Hill quote, this equals locations[0].tiv.
    5. NEVER emit a location whose name/address comes from the Insured LLC when the LLC name contains a street that does NOT match the PREMISES AND BUILDINGS address. The property-schedule address wins over the LLC name.
- LLC/ENTITY NAME IS NOT THE PROPERTY ADDRESS: The Insured / Named Insured is often an LLC whose legal name contains a street address that is DIFFERENT from the actual insured property (e.g., "1345 Lee Rd LLC" owning a hotel at 1275 Lee Rd). NEVER copy the LLC name or any address embedded in the LLC name into locations[].address. Always use the address from the property schedule / PREMISES AND BUILDINGS / Building Description / Schedule of Locations section. If the only address you can find IS the one inside the LLC name, leave the location address empty rather than guess.
- REJECT PHANTOM LOCATION PHRASES: Do NOT create a location entry from generic phrases like "See attached for Schedule of Locations", "Per Schedule", "See SOV", "Various", "TBD", "Pending", or any similar placeholder language. Only emit a location when you have an actual street address or hotel/property name from the quote.
- LOCATIONS FROM SOV - AUTHORITATIVE SOURCE: When a Statement of Values (SOV) Excel file is provided (look for FILE (Excel): markers with "SOV" in the filename), treat it as the AUTHORITATIVE source for the locations[] array. Include EVERY distinct row from the SOV as a location, regardless of which named insured owns that row. The SOV commonly contains rows owned by the primary named insured AND by secondary named insureds (e.g., Premier Hotels LLC, Maruti Hospitality Inc, Milan Properties Inc) - ALL of them are valid locations on this placement. Do NOT filter SOV rows by the primary named insured's LLC name. Deduplicate only when two rows share the same street address (multi-building rows at the same address collapse to ONE location; set tiv = sum of building limits for that address). For each SOV location: name = DBA column (e.g., "Quality Inn & Suites Universal", "Express by Marriott"), address = Address column, city/state/zip = the respective columns. If a location appears in BOTH the SOV and a GL premises schedule, the SOV row's DBA name wins for the location name.
- For Property: ALWAYS include Flood and Earthquake rows even if excluded
- For Property deductibles: Do NOT extract deductibles for perils marked "NOT COVERED" in the sublimits. If Named Windstorm sublimit says "NOT COVERED", omit the Named Storm/Named Windstorm deductible entirely. Only extract deductibles for perils that actually have coverage on this specific policy.
- For Property additional_coverages (sublimits/extensions): This section is MANDATORY. Extract ALL sublimits of liability, also called extensions of coverage or additional coverages. Common property sublimits include: Flood, Earthquake, Equipment Breakdown, Ordinance or Law, Spoilage, Business Income Extended Period, Sign Coverage, Accounts Receivable, Valuable Papers, Fine Arts, Newly Acquired Property, Transit, Debris Removal, Pollutant Cleanup, Utility Services, Green Building, Sewer/Drain Backup, Water Damage, Mold/Fungi, and any other sublimit or extension listed in the quote. Include the limit and deductible for each.
- For Property forms_endorsements: This section is MANDATORY. Extract EVERY policy form and endorsement listed in the property quote. Include the exact form number (e.g., CP 00 10 06/07, PR 001, PR 902, SSPN-018, LMA 5401, NMA1191) and description. These are typically listed under "Endorsements/Additional Endorsements" or "Forms Schedule" - may span MULTIPLE PAGES. Extract ALL items (a through z, aa through zz, etc.). Do NOT skip this section even if the list is long (50+ forms is normal for property).
- For General Liability forms_endorsements: This section is MANDATORY. Extract EVERY form and endorsement listed under "PRIMARY GENERAL LIABILITY FORMS & ENDORSEMENTS" or similar GL-specific forms schedule. These forms have form numbers starting with CG, AD, AI, DE, JA, IL (liability-specific), etc. Do NOT copy property forms (CP, MS PR, HSIC, MS DEC, MS EBC) into the GL section. Each coverage type must have ONLY its own forms.
- For Umbrella/Excess forms_endorsements: Extract the forms listed under the umbrella/excess liability quote. If the umbrella quote shares a forms schedule with GL (common with Admiral), extract the umbrella-specific forms. Do NOT copy property forms into the umbrella section.
- FORMS SEPARATION RULE: Each coverage's forms_endorsements array must contain ONLY forms from that specific coverage's quote document. Property forms (CP, PR, MS PR, HSIC, SSPN, LMA, NMA, 6133x forms) go ONLY in the property section. GL forms (CG, AD, AI, DE, JA forms) go ONLY in the general_liability section. Umbrella/excess forms (SCX, NXLL, CSXC forms) go ONLY in the umbrella section. EPLI forms (BR, EMD, EMO, EGD, PN forms) go ONLY in the epli section. Crime forms should contain ONLY crime/fidelity-specific forms (e.g., CR, bond, fidelity forms) — do NOT copy property, GL, umbrella, or EPLI forms into the crime section. When a single carrier (e.g., Coalition) provides both EPLI and Crime, extract separate forms for each — crime gets only crime-specific endorsements, EPLI gets only EPLI endorsements. Shared policy jacket forms (IL, EMN, EMJ) should go in the primary coverage only, not duplicated across both.
- For General Liability limits: Extract ALL limits of liability listed on the quote, not just the standard 6 CGL limits. Many hotel GL policies include additional limits for Employee Benefits (Each Claim and Aggregate), Sexual Abuse (Each Act and Aggregate), Hired & Non-Owned Auto, and Assault & Battery (Each Event and Aggregate). Include EVERY limit line item shown on the carrier quote in the "limits" array. Also extract the ACTUAL dollar amounts from the quote - do not use defaults like $100,000 for Damage to Rented Premises or $5,000 for Medical Payments if the quote shows different amounts.
- For General Liability total_sales and schedule_of_classes exposure: The "total_sales" field must contain the ACTUAL total gross sales figure from the quote's rate basis line. Look for text like "Per $1,000 Gross Sales ($X)" or "Gross Sales: $X" and extract $X as total_sales. Do NOT fabricate or estimate per-class exposure amounts in schedule_of_classes - if the quote does not show individual per-class exposure breakdowns, leave the exposure field empty for each class entry. The total_sales field is the authoritative source for the Information Summary.
- For EPLI / Employment Practices Liability / Management Liability (ProEx): Extract as coverage_type "epli". ProEx Management Liability proposals from carriers like Coalition, Travelers, or Hartford contain EPL coverage. Look for "Employment Practices", "EPL", "EPLI", "Management Liability", or "ProEx" in the document. Extract the carrier name, AM Best rating, premium, surplus lines tax, total_premium (premium + SLT only, no broker fees), defense provisions (Duty to Defend or Non-Duty to Defend), aggregate limit, third-party discrimination/harassment sublimit, additional defense limit, retention per claim, and all sublimits (wage & hour, workplace violence, immigration, WARN Act, biometric, employee privacy). Also extract notable endorsements like "Bodily Injury & Property Damage Exclusion" or "Physical or Sexual Abuse Exclusion" with their coverage detail (e.g., "Yes - absolute language"). CRITICAL: The ProEx/Management Liability PDF is a SEPARATE coverage from General Liability — do NOT merge EPLI data into the GL section.
- For ALL coverage types subjectivities: This section is CRITICAL. Extract ALL conditions, subjectives, binding requirements, and binding conditions listed in the quote. These are often on a page titled "CONDITIONS & SUBJECTIVES", "BINDING REQUIREMENTS", "BINDING SUBJECTIVITIES", or "BINDING CONDITIONS". Each bullet point or numbered item should be a separate string in the subjectivities array. Include items like: loss control report requirements, certificates of insurance requirements, named insured confirmation, application requirements, ACORD application deadlines, terrorism form requirements, payment of state taxes, inspection/audit contact requirements, and any other conditions the carrier requires before or after binding. Do NOT skip or summarize - extract each condition verbatim as written in the quote.
- For named_insureds: Extract each named insured as an object with "name" and "dba" fields. Do NOT repeat the same entity twice (case-insensitive). If a named insured has a DBA or trade name EXPLICITLY listed in the quote (e.g., "Q Hotels Management LLC DBA Best Western"), split into name="Q Hotels Management LLC" and dba="Best Western". CRITICAL RULES: (1) Only include DBAs that are EXPLICITLY written as "DBA", "d/b/a", or "doing business as" in the documents. (2) Do NOT infer DBAs from hotel brand names, location names, or SOV entries. (3) Do NOT fabricate entity names like "Cajun Lodging LLC" unless that exact name appears in the quote documents. (4) If a named insured appears as "Q HOTEL MANAGEMENT, LLC" in ALL CAPS, extract it exactly as written - the generator will handle proper case formatting. (5) Do NOT create separate named insured entries for each hotel brand - those are locations, not named insureds.
- For additional_named_insureds: CRITICAL - Search ALL pages thoroughly for "Additional Named Insured", "Additional Named Insureds Schedule", "Named Insured Schedule", or similar headings. These are often on a SEPARATE PAGE listing 5-15+ entities (LLCs, management companies with DBAs). You MUST extract EVERY SINGLE entity listed - do NOT stop early or truncate. Count the entities and verify your count matches the document. Each entity is typically an LLC with a DBA hotel brand name (e.g., "PORT PLAZA HOTEL LLC DBA HOME2SUITES BY HILTON"). Extract each one as {{name: "LLC name", dba: "brand name"}}. Do NOT duplicate entities already in named_insureds.
- For additional_insureds: Search for "Additional Insured", "Additional Insured Schedule", or endorsement pages listing additional insureds (franchisors, mortgagees, managers). Extract all of them.
- CRIME COVERAGE: For crime/fidelity bond policies (e.g., Chubb ForeFront Portfolio, Travelers Crime), extract ALL insuring clauses with their individual limits and retentions. Common insuring clauses include: Employee Theft, Forgery or Alteration, Inside the Premises (Theft of Money & Securities), Inside the Premises (Robbery/Safe Burglary), Outside the Premises, Computer and Funds Transfer Fraud, Money Orders and Counterfeit Money, Social Engineering Fraud. Also extract all endorsements from the forms schedule. If the policy is claims-made, note the retroactive date.
- RT SPECIALTY / LLOYD'S OF LONDON PROPERTY QUOTES: RT Specialty quotes have a two-part format: (1) an RT "Insurance Proposal" cover with Cost Summary, and (2) a "Quotation Memorandum" from the actual carrier (Lloyd's via Owners First Association). CRITICAL extraction rules: Use the Quotation Memorandum "PREMIUM AND FEES" section Total as total_premium (this includes brokerage fees and association fees). The RT cover "Total Policy Cost" includes surplus lines tax which should NOT be in total_premium. Extract TIV from "RISK DETAILS" section. Extract deductibles from the lettered A/B/C structure (e.g., A=$10K AOP, B=$25K per building for water damage). Extract ALL lettered sublimits A through FF+ from "SCHEDULE OF SUBLIMITS". If Wind/Hail or Named Windstorm says "EXCLUDED", mark as "NOT COVERED" in additional_coverages. Extract WARRANTIES as subjectivities. Extract SUBJECTIVITIES as binding conditions. The SOV page (Schedule of Values) contains per-location breakdowns with LID, address, year built, construction, stories, units, sprinklered status, gross sq ft, building value, BPP value, BI value, and TIV — use this for coverage_by_location. Set carrier_admitted to false (Lloyd's is non-admitted/surplus lines).
- LAYERED PROPERTY PROGRAMS: When a property quote contains multiple carriers in a layered/shared program (e.g., Lexington primary + Kinsale excess + Gotham/Coaction excess), extract EACH layer separately. Use "property" for the primary layer, "excess_property" for the first excess layer, and "excess_property_2" for the second excess layer. Each layer has its own carrier, premium, limits, deductibles, forms, subjectivities, and coinsurance. The layer_description should show the attachment point (e.g., "$10,000,000 xs $10,000,000"). Look for terms like "Excess", "xs", "excess of", or "Per Schedule" to identify excess layers. Common excess property carriers include Kinsale, Gotham (via Coaction), and others.
- COINSURANCE & VALUATION: For ALL property layers (primary and excess), extract the coinsurance percentage for Building, Business Income, and BPP. Also extract the Monthly Limitation for Business Income (e.g., "1/4 Monthly", "1/3 Monthly"). This is a CRITICAL field that must ALWAYS be included in property quotes. Look for "Coinsurance", "Monthly Limitation", "Coinsurance & Valuation" sections. If coinsurance is waived or 0%, still include it as "0%". Also extract the valuation basis (Replacement Cost, Actual Cash Value, Agreed Value).
- UMBRELLA/EXCESS LAYERS: When multiple umbrella/excess liability quotes are provided (e.g., separate PDFs for different layers), extract EACH layer as a separate coverage entry. Use "umbrella" for the primary excess layer, "umbrella_layer_2" for the second excess layer ($XM xs $XM), and "umbrella_layer_3" for the third excess layer ($XM xs $XM). Each layer has its own carrier, premium, limits, forms, and subjectivities. The tower_structure field should show that layer's position. Look for "Controlling Underlying" or "Schedule of Underlying" to determine the layer position. If a quote says it sits excess of another carrier's layer, it is a higher layer.
- EXCESS LAYER COUNTING (MANDATORY): Before finalizing your output, COUNT the distinct excess/umbrella PDFs in the document set. Each separate PDF with its own carrier for "Excess Liability", "Commercial Excess Liability", "Excess Umbrella", or "XS Liability" is a SEPARATE layer. If N such PDFs exist, your output MUST contain exactly N layer entries: 1 PDF = umbrella only; 2 PDFs = umbrella + umbrella_layer_2; 3 PDFs = umbrella + umbrella_layer_2 + umbrella_layer_3. NEVER merge two separate excess PDFs into one layer. NEVER drop a layer because you cannot determine the attachment point — if unsure, place it at the next open layer slot and mark tower_structure as "position uncertain". The attachment order from lowest to highest is determined by the "xs $X" or "excess of $X" amount shown on each quote (e.g., "$5M x P" = primary = umbrella; "$5M x $5M" = second layer; "$5M x $10M" = third layer).
- CRITICAL DISTINCTION - EXCESS LIABILITY vs EXCESS PROPERTY: "Excess Liability" is NOT the same as "Excess Property". If a quote says "Excess Liability", "Excess Liability Quotation", or "XS Liability" and its Schedule of Underlying Insurance references an Umbrella or General Liability policy, it is an UMBRELLA/EXCESS LIABILITY layer - use "umbrella" or "umbrella_layer_2" or "umbrella_layer_3". Do NOT classify it as "property" or "excess_property". Excess Property layers sit excess of a primary PROPERTY policy and cover physical damage to buildings/contents. Excess Liability layers sit excess of an Umbrella or GL policy and cover bodily injury/property damage liability claims. If the underlying schedule shows an umbrella or GL carrier, it is ALWAYS an excess liability layer, never excess property.
- SAME CARRIER FOR GL AND EXCESS: When the SAME carrier (e.g., Admiral Insurance Company) provides BOTH a General Liability quote AND an Excess Liability/Umbrella quote in separate PDF files, these MUST be extracted as SEPARATE coverage entries. Extract the GL quote under "general_liability" and the Excess Liability quote under "umbrella". Do NOT merge or combine them into one entry just because they share a carrier name. Look at the coverage type stated on each document ("Coverage: Excess Liability" vs "Coverage: General Liability") and the document title ("Commercial Excess Liability Quote" vs "Commercial General Liability Quote") to distinguish them.
- EXCESS TOWER FILENAME HINTS: When excess/umbrella PDFs have sparse body text (scanned pages) but the FILENAME contains shorthand like "5x5", "5 x 10", "5xP", "5 x P", "10 x 5", "5M x 10M", etc., USE the filename to determine the attachment point and emit the appropriate excess layer. Shorthand decoder: first number is the LIMIT, second value after "x" is the ATTACH point. "xP" or "x P" or "x Primary" = primary excess (coverage key = umbrella). "x5" or "5x5" or "x $5M" = sits above $5M (coverage key = umbrella_layer_2, attach = $5M). "x10" or "5x10" or "x $10M" = sits above $10M (coverage key = umbrella_layer_3, attach = $10M). Example: filename "Ascot 5x10 Excess Liability Quote" with scanned body emits umbrella_layer_3 with carrier = Ascot Specialty Insurance Company, each_occurrence = $5,000,000, aggregate = $5,000,000, attach_point = $10,000,000, even if limits fields on page 1 are blank. Extract the premium/fees from readable pages. Do this ONLY when body text does not contradict the filename inference.
- MULTI-OPTION EXCESS QUOTES: Some excess liability quotes present multiple limit options in columns (e.g., $1M/$2M/$3M Each Loss Event with different premiums for each). Extract the HIGHEST limit option as the primary "umbrella" entry. If the user needs a different option, they can adjust in the editor.
- COMPETING QUOTES (MULTIPLE CARRIERS FOR SAME COVERAGE): When documents contain quotes from DIFFERENT carriers for the SAME line of coverage (e.g., Starr Property quote AND Markel Property quote in separate PDFs), extract EACH as a separate coverage entry. Use the base key for the first (e.g., "property") and "_alt_1", "_alt_2" suffixes for additional competing quotes (e.g., "property_alt_1", "general_liability_alt_1"). Do NOT discard any carrier's quote. Do NOT confuse competing quotes with layered programs - layered programs have excess/xs relationships, while competing quotes are independent quotes at the same attachment point from different carriers.

DOCUMENT TEXT:
{document_text}"""


async def extract_and_structure_data(file_paths: list[str]) -> dict:
    """
    Extract text from all uploaded documents and use GPT to structure
    the data into a standardized format for proposal generation.

    Uses smart page-level extraction for PDFs to prioritize quote
    summary pages over forms/endorsements boilerplate.
    """
    # Step 1: Extract text from all documents using smart extraction
    all_text = []
    for fp in file_paths:
        fname = Path(fp).name
        text = extract_document_smart(fp)
        if text:
            all_text.append(f"\n{'='*60}\nFILE: {fname}\n{'='*60}\n{text}")
            logger.info(f"Smart extraction from {fname}: {len(text)} chars")
        else:
            logger.warning(f"No text extracted from: {fname}")

    if not all_text:
        return {"error": "Could not extract text from any uploaded documents."}

    combined_text = "\n".join(all_text)

    # Final safety truncation (should rarely be needed with smart extraction)
    max_chars = 300000
    if len(combined_text) > max_chars:
        logger.warning(f"Combined text truncated from {len(combined_text)} to {max_chars} chars")
        combined_text = combined_text[:max_chars]

    logger.info(f"Sending {len(combined_text)} chars to GPT for extraction")

    # Step 2: Send to GPT for structured extraction
    try:
        response = _get_openai_client().chat.completions.create(
            model=GPT_MODEL,
            messages=[
                {"role": "system", "content": EXTRACTION_SYSTEM_PROMPT},
                {"role": "user", "content": EXTRACTION_USER_PROMPT.format(document_text=combined_text)}
            ],
            response_format={"type": "json_object"},
            temperature=0.1,
            max_tokens=32000
        )

        result_text = response.choices[0].message.content
        finish_reason = response.choices[0].finish_reason
        logger.info(f"GPT response: {len(result_text)} chars, finish_reason={finish_reason}")

        if finish_reason == "length":
            logger.warning("GPT response was truncated (hit max_tokens). Attempting to parse partial JSON.")

        data = json.loads(result_text)
        
        # Normalize coverages: GPT sometimes returns a list instead of dict
        covs = data.get("coverages", {})
        if isinstance(covs, list):
            normalized = {}
            for item in covs:
                if isinstance(item, dict):
                    cov_type = item.get("coverage_type", item.get("type", "unknown"))
                    if cov_type not in normalized:
                        normalized[cov_type] = item
                    else:
                        # Competing quote - find next available alt slot
                        for alt_n in range(1, 5):
                            alt_key = f"{cov_type}_alt_{alt_n}"
                            if alt_key not in normalized:
                                normalized[alt_key] = item
                                logger.info(f"Competing quote: {cov_type} already exists, stored as {alt_key}")
                                break
            data["coverages"] = normalized
            logger.info(f"Normalized coverages from list ({len(covs)} items) to dict ({list(normalized.keys())})")
        elif not isinstance(covs, dict):
            data["coverages"] = {}
        
        # Also fix individual coverage values that are lists instead of dicts
        covs = data.get("coverages", {})
        if isinstance(covs, dict):
            for key, val in list(covs.items()):
                if isinstance(val, list):
                    if len(val) >= 1 and isinstance(val[0], dict):
                        covs[key] = val[0]
                        logger.info(f"Unwrapped list for coverage '{key}'")
                    else:
                        covs[key] = {}
                elif not isinstance(val, dict):
                    covs[key] = {}
        
        logger.info(f"GPT extraction successful. Coverages found: {list(data.get('coverages', {}).keys())}")

        # POST-PROCESSING: Validate and fix common extraction issues
        
        # Fix 1: Clean up named insureds - remove entries with multiple hotel brand names
        _brand_names = {"marriott", "hilton", "ihg", "wyndham", "best western", "choice",
                       "hampton inn", "hampton", "holiday inn", "holiday inn express",
                       "candlewood", "towneplace", "staybridge", "springhill",
                       "comfort inn", "comfort suites", "quality inn", "sleep inn"}
        raw_named = data.get("named_insureds", [])
        cleaned_named = []
        for ni in raw_named:
            ni_name = ni.get("name", "") if isinstance(ni, dict) else str(ni)
            ni_lower = ni_name.lower()
            brand_count = sum(1 for b in _brand_names if b in ni_lower)
            if brand_count >= 3:
                # This is likely a hallucinated concatenation - try to extract just the entity
                import re as _re_fix
                m = _re_fix.match(r'^(.+?\b(?:LLC|LP|LLP|Inc|Corp)\b)', ni_name, _re_fix.IGNORECASE)
                if m:
                    if isinstance(ni, dict):
                        ni["name"] = m.group(1).strip()
                        ni["dba"] = ""  # Clear the hallucinated DBA
                    else:
                        ni = {"name": m.group(1).strip(), "dba": ""}
                    logger.warning(f"Fixed hallucinated named insured: '{ni_name}' -> '{ni.get('name', ni) if isinstance(ni, dict) else ni}'")
            cleaned_named.append(ni)
        data["named_insureds"] = cleaned_named
        
        # Fix 2: Log GL carrier name (keep original from quote)
        gl_cov = data.get("coverages", {}).get("general_liability", {})
        if gl_cov:
            carrier = gl_cov.get("carrier", "")
            logger.info(f"GL carrier from extraction: '{carrier}'")
        
        # Fix 3: Validate that forms_endorsements is not empty for property and GL
        for cov_key in ["property", "general_liability"]:
            cov = data.get("coverages", {}).get(cov_key, {})
            if cov and not cov.get("forms_endorsements"):
                logger.warning(f"{cov_key} has no forms_endorsements extracted - may need manual review")
        
        # Fix 4: Validate additional_coverages for property
        prop_cov = data.get("coverages", {}).get("property", {})
        if prop_cov and not prop_cov.get("additional_coverages"):
            logger.warning("Property has no additional_coverages (sublimits) extracted - may need manual review")

        # Validate and fix total_premium for each coverage
        def _to_num(val):
            if isinstance(val, (int, float)):
                return float(val)
            if isinstance(val, str):
                try:
                    return float(val.replace(",", "").replace("$", ""))
                except (ValueError, TypeError):
                    return 0.0
            return 0.0

        for key, cov in data.get("coverages", {}).items():
            premium = _to_num(cov.get("premium", 0))
            taxes_fees = _to_num(cov.get("taxes_fees", 0))
            total_premium = _to_num(cov.get("total_premium", 0))
            surplus_lines_tax = _to_num(cov.get("surplus_lines_tax", 0))
            stamping_fee = _to_num(cov.get("stamping_fee", 0))
            tria_premium = _to_num(cov.get("tria_premium", 0))

            # Store cleaned numeric values back
            cov["premium"] = premium
            cov["taxes_fees"] = taxes_fees
            cov["total_premium"] = total_premium

            # Cross-check: for GL, sum schedule_of_classes premiums vs reported premium
            if key == "general_liability" and cov.get("schedule_of_classes"):
                soc_total = 0
                for cls in cov["schedule_of_classes"]:
                    if isinstance(cls, dict):
                        cls_prem = _to_num(cls.get("premium", 0))
                        soc_total += cls_prem
                if soc_total > 0 and premium > 0:
                    diff = abs(soc_total - premium)
                    if diff > 1:  # Allow $1 rounding
                        logger.warning(f"  GL schedule_of_classes premium sum ({soc_total}) != "
                                     f"reported premium ({premium}), diff={diff}")
                        # If the schedule total is closer to total_premium, GPT may have
                        # confused premium with total_premium
                        if abs(soc_total - total_premium) < abs(soc_total - premium) and soc_total < total_premium:
                            logger.info(f"  GL: schedule sum {soc_total} is closer to total_premium {total_premium}, "
                                       f"likely correct extraction")

            # Fallback: if total_premium is less than premium, recalculate
            if total_premium < premium and taxes_fees > 0:
                corrected = premium + taxes_fees
                logger.warning(f"  {key}: total_premium ({total_premium}) < premium ({premium}). "
                             f"Correcting to premium + taxes_fees = {corrected}")
                cov["total_premium"] = corrected
                total_premium = corrected
            elif total_premium == 0 and premium > 0:
                cov["total_premium"] = premium + taxes_fees
                total_premium = cov["total_premium"]
                logger.info(f"  {key}: total_premium was 0, set to premium + taxes_fees = {total_premium}")
            
            # Additional check: if taxes_fees is 0 but total_premium > premium, derive taxes_fees
            if taxes_fees == 0 and total_premium > premium and premium > 0:
                cov["taxes_fees"] = total_premium - premium
                taxes_fees = cov["taxes_fees"]
                logger.info(f"  {key}: derived taxes_fees = {taxes_fees} from total_premium - premium")

            logger.info(f"  {key}: carrier={cov.get('carrier', 'N/A')}, premium={premium}, "
                       f"taxes_fees={taxes_fees}, total_premium={total_premium}")

        # CRITICAL FIX: Override total_premium with exact values from raw PDF text
        # GPT frequently miscalculates taxes/fees. Scan the raw text for "Total Cost of Policy",
        # "Total Package Cost", etc. and use the exact number from the document.
        _total_cost_patterns = [
            (r'Total\s+Cost\s+of\s+Policy[:\s]*[\$]?\s*([\d,]+(?:\.\d{2})?)', 'Total Cost of Policy'),
            (r'Total\s+Package\s+Cost[:\s]*[\$]?\s*([\d,]+(?:\.\d{2})?)', 'Total Package Cost'),
            (r'Total\s+Policy\s+(?:Cost|Premium)[:\s]*[\$]?\s*([\d,]+(?:\.\d{2})?)', 'Total Policy Cost/Premium'),
            (r'Total\s+Estimated\s+(?:Cost|Premium)[:\s]*[\$]?\s*([\d,]+(?:\.\d{2})?)', 'Total Estimated Cost'),
            (r'Grand\s+Total[:\s]*[\$]?\s*([\d,]+(?:\.\d{2})?)', 'Grand Total'),
        ]
        
        # Extract all total cost values from the raw text
        raw_totals = []
        for pattern, label in _total_cost_patterns:
            matches = re.findall(pattern, combined_text, re.IGNORECASE)
            for m in matches:
                try:
                    val = float(m.replace(',', ''))
                    if val > 1000:  # Ignore trivially small values
                        raw_totals.append((val, label))
                        logger.info(f"  Raw text premium found: {label} = ${val:,.2f}")
                except (ValueError, TypeError):
                    pass
        
        if raw_totals:
            # Match raw totals to coverages by finding the closest match
            for key, cov in data.get("coverages", {}).items():
                gpt_total = _to_num(cov.get("total_premium", 0))
                gpt_premium = _to_num(cov.get("premium", 0))
                if gpt_premium <= 0:
                    continue
                
                # Find the raw total that's closest to GPT's total_premium but >= premium
                best_match = None
                best_diff = float('inf')
                for raw_val, raw_label in raw_totals:
                    # The raw total should be >= the base premium (it includes taxes/fees)
                    if raw_val >= gpt_premium * 0.95:  # Allow 5% tolerance
                        diff = abs(raw_val - gpt_total) if gpt_total > 0 else abs(raw_val - gpt_premium)
                        # Only match if reasonably close (within 20% of premium)
                        if diff < gpt_premium * 0.20 and diff < best_diff:
                            best_match = (raw_val, raw_label)
                            best_diff = diff
                
                if best_match and abs(best_match[0] - gpt_total) > 1:  # Only override if different
                    old_total = gpt_total
                    cov["total_premium"] = best_match[0]
                    # Recalculate taxes_fees from the corrected total
                    if gpt_premium > 0:
                        cov["taxes_fees"] = best_match[0] - gpt_premium
                    logger.warning(f"  PREMIUM OVERRIDE for {key}: GPT total_premium ${old_total:,.2f} "
                                  f"-> raw text '{best_match[1]}' ${best_match[0]:,.2f} "
                                  f"(diff: ${abs(best_match[0] - old_total):,.2f})")

        return data

    except json.JSONDecodeError as e:
        logger.error(f"GPT returned invalid JSON: {e}")
        logger.error(f"Raw response (first 500 chars): {result_text[:500] if 'result_text' in dir() else 'N/A'}")
        return {"error": f"Failed to parse extraction results: {e}"}
    except Exception as e:
        error_str = str(e)
        # Retry on rate limit errors (429)
        if "429" in error_str or "rate_limit" in error_str.lower():
            logger.warning(f"Rate limit hit, waiting 60 seconds and retrying...")
            import asyncio
            await asyncio.sleep(60)
            try:
                response = _get_openai_client().chat.completions.create(
                    model=GPT_MODEL,
                    messages=[
                        {"role": "system", "content": EXTRACTION_SYSTEM_PROMPT},
                        {"role": "user", "content": EXTRACTION_USER_PROMPT.format(document_text=combined_text)}
                    ],
                    response_format={"type": "json_object"},
                    temperature=0.1,
                    max_tokens=32000
                )
                result_text = response.choices[0].message.content
                data = json.loads(result_text)
                logger.info(f"Retry successful. Coverages found: {list(data.get('coverages', {}).keys())}")
                return data
            except Exception as retry_e:
                logger.error(f"Retry also failed: {retry_e}")
                return {"error": f"AI extraction failed after retry: {retry_e}"}
        logger.error(f"GPT extraction failed: {e}")
        return {"error": f"AI extraction failed: {e}"}


def format_verification_message(data: dict) -> str:
    """
    Format the extracted data into a verification message for the user
    to review before generating the proposal.
    """
    if "error" in data:
        return f"Extraction Error: {data['error']}"

    lines = []
    lines.append("----------------------------------------")
    lines.append("ð EXTRACTED DATA - PLEASE VERIFY")
    lines.append("-----------------------------------------")

    # Client Info
    ci = data.get("client_info", {})
    if ci:
        lines.append("")
        lines.append("-¸ CLIENT INFORMATION")
        lines.append(f"  Named Insured: {ci.get('named_insured', 'N/A')}")
        if ci.get("dba"):
            lines.append(f"  DBA: {ci['dba']}")
        if ci.get("address"):
            lines.append(f"  Address: {ci['address']}")
        if ci.get("effective_date"):
            lines.append(f"  Effective: {ci['effective_date']}")
        if ci.get("expiration_date"):
            lines.append(f"  Expiration: {ci['expiration_date']}")

    # Premium Summary
    coverages = data.get("coverages", {})
    if coverages:
        lines.append("")
        lines.append("-¸ PREMIUM SUMMARY")
        lines.append(f"  {'Coverage':<25} {'Carrier':<20} {'Total Premium':>15}")
        lines.append(f"  {'-'*25} {'-'*20} {'-'*15}")

        grand_total = 0
        coverage_names = {
            "property": "Property",
            "general_liability": "General Liability",
            "umbrella": "Umbrella",
            "workers_comp": "Workers Comp",
            "commercial_auto": "Commercial Auto",
            "cyber": "Cyber",
            "epli": "EPLI",
            "flood": "Flood",
            "terrorism": "Terrorism / TRIA",
            "crime": "Crime",
            "inland_marine": "Inland Marine"
        }

        for key, display_name in coverage_names.items():
            cov = coverages.get(key)
            if cov:
                carrier = cov.get("carrier", "N/A")
                total = cov.get("total_premium", 0)
                if not isinstance(total, (int, float)):
                    try:
                        total = float(str(total).replace(",", "").replace("$", ""))
                    except (ValueError, TypeError):
                        total = 0
                admitted = "" if cov.get("carrier_admitted", True) else " (Non-Admitted)"
                lines.append(f"  {display_name:<25} {carrier[:20]:<20} ${total:>12,.0f}")
                if admitted:
                    lines.append(f"  {'':25} {admitted}")
                grand_total += total

        lines.append(f"  {'-'*25} {'-'*20} {'-'*15}")
        lines.append(f"  {'TOTAL':<25} {'':20} ${grand_total:>12,.0f}")

    # Coverage Details
    for key, display_name in [("property", "PROPERTY"), ("general_liability", "GENERAL LIABILITY"),
                               ("umbrella", "UMBRELLA"), ("workers_comp", "WORKERS COMP"),
                               ("commercial_auto", "COMMERCIAL AUTO"), ("cyber", "CYBER"),
                               ("epli", "EPLI"), ("flood", "FLOOD"),
                               ("terrorism", "TERRORISM / TRIA"), ("crime", "CRIME"),
                               ("inland_marine", "INLAND MARINE")]:
        cov = coverages.get(key)
        if not cov:
            continue

        lines.append("")
        lines.append(f"-¸ {display_name}")
        lines.append(f"  Carrier: {cov.get('carrier', 'N/A')}")
        lines.append(f"  AM Best: {cov.get('am_best_rating', 'N/A')}")

        # Limits
        limits = cov.get("limits", [])
        if limits and isinstance(limits, list):
            lines.append("  Limits:")
            for lim in limits:
                if isinstance(lim, dict):
                    lines.append(f"    -¢ {lim.get('description', '')}: {lim.get('limit', '')}")
                elif isinstance(lim, str):
                    lines.append(f"    -¢ {lim}")

        # Deductibles
        deductibles = cov.get("deductibles", [])
        if deductibles and isinstance(deductibles, list):
            lines.append("  Deductibles:")
            for ded in deductibles:
                if isinstance(ded, dict):
                    lines.append(f"    -¢ {ded.get('description', '')}: {ded.get('amount', '')}")
                elif isinstance(ded, str):
                    lines.append(f"    -¢ {ded}")

        # Additional Coverages
        addl = cov.get("additional_coverages", [])
        if addl and isinstance(addl, list):
            lines.append("  Additional Coverages:")
            for ac in addl:
                if isinstance(ac, dict):
                    ded_str = f" (Ded: {ac['deductible']})" if ac.get("deductible") else ""
                    lines.append(f"    -¢ {ac.get('description', '')}: {ac.get('limit', '')}{ded_str}")
                elif isinstance(ac, str):
                    lines.append(f"    -¢ {ac}")

        # Forms count
        forms = cov.get("forms_endorsements", [])
        if forms and isinstance(forms, list):
            lines.append(f"  Forms & Endorsements: {len(forms)} extracted")
            for f in forms[:5]:
                if isinstance(f, dict):
                    lines.append(f"    -¢ {f.get('form_number', '')} - {f.get('description', '')}")
                elif isinstance(f, str):
                    lines.append(f"    -¢ {f}")
            if len(forms) > 5:
                lines.append(f"    ... and {len(forms) - 5} more")

        # Subjectivities / Conditions
        subjs = cov.get("subjectivities", [])
        if subjs and isinstance(subjs, list):
            lines.append(f"  Conditions & Subjectivities: {len(subjs)} items")
            for s in subjs[:5]:
                s_text = s if isinstance(s, str) else str(s)
                # Truncate long items for display
                if len(s_text) > 100:
                    s_text = s_text[:97] + "..."
                lines.append(f"    - {s_text}")
            if len(subjs) > 5:
                lines.append(f"    ... and {len(subjs) - 5} more")

    # Locations
    locations = data.get("locations", [])
    if locations:
        lines.append("")
        lines.append(f"-¸ LOCATIONS: {len(locations)} found")
        for loc in locations[:5]:
            lines.append(f"  {loc.get('number', '?')}. {loc.get('address', '')} {loc.get('city', '')}, {loc.get('state', '')} {loc.get('zip', '')}")
        if len(locations) > 5:
            lines.append(f"  ... and {len(locations) - 5} more")

    # Named Insureds
    named = data.get("named_insureds", [])
    if named:
        lines.append("")
        lines.append(f"-¸ NAMED INSUREDS: {len(named)}")
        for ni in named[:5]:
            lines.append(f"  -¢ {ni}")
        if len(named) > 5:
            lines.append(f"  ... and {len(named) - 5} more")

    lines.append("")
    lines.append("-----------------------------------------")
    lines.append("- ï¸ PLEASE VERIFY ALL DATA ABOVE")
    lines.append("")
    lines.append("Reply with:")
    lines.append("  - /proposal confirm - to generate the proposal")
    lines.append("  -ï¸ Send corrections as a message")
    lines.append("  - /proposal cancel - to cancel")
    lines.append("-----------------------------------------")

    return "\n".join(lines)


async def apply_corrections(data: dict, corrections_text: str) -> dict:
    """Use GPT to apply user corrections to the extracted data."""
    try:
        response = _get_openai_client().chat.completions.create(
            model=GPT_MODEL,
            messages=[
                {"role": "system", "content": "You are an insurance data correction assistant. Apply the user's corrections to the extracted data JSON. Return the corrected JSON object. Only modify the fields mentioned in the corrections, keep everything else the same."},
                {"role": "user", "content": f"Current extracted data:\n{json.dumps(data, indent=2)}\n\nUser corrections:\n{corrections_text}\n\nReturn the corrected JSON:"}
            ],
            response_format={"type": "json_object"},
            temperature=0.1,
            max_tokens=32000
        )
        corrected = json.loads(response.choices[0].message.content)
        logger.info("Corrections applied successfully")
        return corrected
    except Exception as e:
        logger.error(f"Failed to apply corrections: {e}")
        return data


class ProposalExtractor:
    """
    Class wrapper around the extraction functions for use by proposal_handler.
    Provides an OOP interface to the module-level extraction functions.
    """

    def extract_pdf_text(self, pdf_path: str) -> str:
        """Extract text from a PDF file using smart extraction."""
        return extract_text_from_pdf_smart(pdf_path)

    def extract_excel_data(self, excel_path: str) -> str:
        """Extract text from an Excel file."""
        return extract_text_from_excel(excel_path)

    def structure_insurance_data(
        self,
        pdf_texts: list[dict],
        excel_data: list[dict],
        client_name: str,
    ) -> dict:
        """
        Use GPT to structure extracted text into a standardized insurance data format.

        Args:
            pdf_texts: List of dicts with 'filename' and 'text' keys
            excel_data: List of dicts with 'filename' and 'data' keys
            client_name: Name of the client/insured
        """
        # Combine all text sources with FAIR BUDGET ALLOCATION
        # Ensures every file gets represented even when total exceeds max_chars
        max_chars = 150000
        all_items = []
        for item in pdf_texts:
            header = f"\n{'='*60}\nFILE: {item['filename']}\n{'='*60}\n"
            all_items.append({"header": header, "text": item["text"], "filename": item["filename"]})
        for item in excel_data:
            header = f"\n{'='*60}\nFILE (Excel): {item['filename']}\n{'='*60}\n"
            all_items.append({"header": header, "text": item["data"], "filename": item["filename"]})

        if not all_items:
            return {"error": "No text extracted from any documents."}

        # Calculate total and check if truncation is needed
        total_chars = sum(len(it["header"]) + len(it["text"]) for it in all_items)
        if total_chars > max_chars:
            # Fair allocation: give each file a proportional budget, but ensure
            # every file gets at least 20K chars (or its full text if shorter)
            n_files = len(all_items)
            min_per_file = min(20000, max_chars // n_files)
            # First pass: allocate minimum to each, then distribute remainder
            remaining = max_chars
            budgets = []
            for it in all_items:
                full_len = len(it["header"]) + len(it["text"])
                budgets.append(min(full_len, min_per_file))
                remaining -= budgets[-1]
            # Second pass: distribute remaining budget proportionally to files that need more
            needs_more = [(i, len(it["header"]) + len(it["text"]) - budgets[i])
                          for i, it in enumerate(all_items) if len(it["header"]) + len(it["text"]) > budgets[i]]
            if needs_more and remaining > 0:
                total_need = sum(need for _, need in needs_more)
                for i, need in needs_more:
                    extra = int(remaining * need / total_need) if total_need > 0 else 0
                    full_len = len(all_items[i]["header"]) + len(all_items[i]["text"])
                    budgets[i] = min(full_len, budgets[i] + extra)
            
            # Build combined text with per-file budgets
            parts = []
            for i, it in enumerate(all_items):
                text_budget = budgets[i] - len(it["header"])
                truncated_text = it["text"][:max(0, text_budget)]
                parts.append(it["header"] + truncated_text)
                if len(it["text"]) > text_budget:
                    logger.warning(f"File '{it['filename']}' truncated from {len(it['text'])} to {text_budget} chars")
                else:
                    logger.info(f"File '{it['filename']}': {len(it['text'])} chars (full)")
            combined_text = "\n".join(parts)
            logger.info(f"Fair budget allocation: {total_chars} total chars -> {len(combined_text)} chars across {n_files} files")
        else:
            combined_text = "\n".join(it["header"] + it["text"] for it in all_items)
            logger.info(f"All files fit within budget: {total_chars} chars across {len(all_items)} files")

        logger.info(f"Sending {len(combined_text)} chars to GPT for extraction")

        try:
            response = _get_openai_client().chat.completions.create(
                model=GPT_MODEL,
                messages=[
                    {"role": "system", "content": EXTRACTION_SYSTEM_PROMPT},
                    {
                        "role": "user",
                        "content": EXTRACTION_USER_PROMPT.format(
                            document_text=combined_text
                        ),
                    },
                ],
                response_format={"type": "json_object"},
                temperature=0.1,
                max_tokens=32000,
            )

            result_text = response.choices[0].message.content
            finish_reason = response.choices[0].finish_reason
            logger.info(f"GPT response: {len(result_text)} chars, finish_reason={finish_reason}")

            data = json.loads(result_text)
            data["client_name"] = client_name
            
            # Normalize coverages: GPT sometimes returns a list instead of dict
            covs = data.get("coverages", {})
            if isinstance(covs, list):
                # Convert list of coverage dicts to keyed dict
                # Handle duplicate coverage types by promoting to _alt_N keys
                normalized = {}
                for item in covs:
                    if isinstance(item, dict):
                        cov_type = item.get("coverage_type", item.get("type", "unknown"))
                        if cov_type not in normalized:
                            normalized[cov_type] = item
                        else:
                            # Competing quote - find next available alt slot
                            for alt_n in range(1, 5):
                                alt_key = f"{cov_type}_alt_{alt_n}"
                                if alt_key not in normalized:
                                    normalized[alt_key] = item
                                    logger.info(f"Competing quote: {cov_type} already exists, stored as {alt_key} (carrier: {item.get('carrier', 'unknown')})")
                                    break
                data["coverages"] = normalized
                covs = normalized
            elif not isinstance(covs, dict):
                data["coverages"] = {}
                covs = data["coverages"]
            
            # Also fix individual coverage values that are lists instead of dicts
            if isinstance(covs, dict):
                for key, val in list(covs.items()):
                    if isinstance(val, list):
                        if len(val) >= 1 and isinstance(val[0], dict):
                            covs[key] = val[0]
                            logger.info(f"Unwrapped list for coverage '{key}'")
                        else:
                            covs[key] = {}
                    elif not isinstance(val, dict):
                        covs[key] = {}
            
            logger.info(
                f"GPT extraction successful. Coverages found: "
                f"{list(covs.keys()) if isinstance(covs, dict) else covs}"
            )

            # Log coverage details
            if isinstance(covs, dict):
                for key, cov in covs.items():
                    if isinstance(cov, dict):
                        logger.info(f"  {key}: carrier={cov.get('carrier', 'N/A')}, premium={cov.get('premium', 0)}, total={cov.get('total_premium', 0)}")

            # ===== UMBRELLA/EXCESS VALIDATION =====
            # Check if combined_text contains excess liability content but no umbrella was extracted
            _has_umbrella = any(k.startswith("umbrella") for k in covs.keys()) if isinstance(covs, dict) else False
            if not _has_umbrella:
                _text_lower = combined_text.lower()
                _excess_indicators = [
                    "excess liability", "commercial excess liability",
                    "xs liability", "excess liability quote",
                    "excess liability quotation", "excess of underlying",
                    "schedule of underlying insurance",
                    "commercial excess liability quote"
                ]
                _has_excess_content = any(ind in _text_lower for ind in _excess_indicators)
                if _has_excess_content:
                    logger.warning("UMBRELLA MISSING: combined_text contains excess liability content but no umbrella coverage was extracted. Attempting targeted re-extraction...")
                    try:
                        # Extract just the excess-related sections
                        _xs_keywords = ["excess liability", "commercial excess", "underlying insurance",
                                        "excess premium", "total excess", "each loss event",
                                        "policy aggregate", "xs quote", "excess quotation"]
                        _xs_text = self._extract_relevant_sections(combined_text, _xs_keywords, context_chars=15000)
                        if len(_xs_text) > 500:
                            _xs_prompt = f"""The main extraction MISSED an Excess Liability / Umbrella coverage that is clearly present in the documents. 
    Extract ONLY the umbrella/excess liability coverage from the text below. Return a JSON object with a single key "umbrella" containing the full coverage data (carrier, premium, total_premium, limits, underlying_insurance, forms_endorsements, etc).
    If the quote shows multiple limit options in columns, extract the HIGHEST limit option.
    The coverage type on this document says "Excess Liability" or "Commercial Excess Liability" - this maps to the "umbrella" key, NOT "general_liability" or "excess_property".

    TEXT:
    {_xs_text}"""
                            _xs_response = _get_openai_client().chat.completions.create(
                                model=GPT_MODEL,
                                messages=[
                                    {"role": "system", "content": "You are an insurance data extraction specialist. Extract the umbrella/excess liability coverage data and return valid JSON."},
                                    {"role": "user", "content": _xs_prompt},
                                ],
                                response_format={"type": "json_object"},
                                temperature=0.1,
                                max_tokens=8000,
                            )
                            _xs_data = json.loads(_xs_response.choices[0].message.content)
                            if "umbrella" in _xs_data and isinstance(_xs_data["umbrella"], dict):
                                _umb = _xs_data["umbrella"]
                                if _umb.get("carrier") or _umb.get("premium"):
                                    data["coverages"]["umbrella"] = _umb
                                    logger.info(f"UMBRELLA RECOVERED: carrier={_umb.get('carrier', 'N/A')}, premium={_umb.get('premium', 0)}, total={_umb.get('total_premium', 0)}")
                                else:
                                    logger.warning("Umbrella re-extraction returned empty data")
                            else:
                                logger.warning(f"Umbrella re-extraction did not return umbrella key. Keys: {list(_xs_data.keys())}")
                    except Exception as e:
                        logger.error(f"Umbrella re-extraction failed: {e}")
            # ===== TOWER VALIDATION (MULTI-LAYER EXCESS) =====
            # If the source contains multiple distinct excess PDFs but fewer layers were
            # extracted, re-run a targeted GPT call to recover missing layers.
            try:
                _covs_tv = data.get("coverages", {}) if isinstance(data.get("coverages"), dict) else {}
                _layer_keys = [k for k in _covs_tv.keys() if k == "umbrella" or k.startswith("umbrella_layer_")]
                _layer_count = len(_layer_keys)
                _text_lower_tv = combined_text.lower()
                # Count distinct excess PDFs via FILE: markers
                _file_markers = re.findall(r"file:\s*([^\n]+\.pdf)", _text_lower_tv)
                _excess_file_count = sum(1 for f in _file_markers if any(w in f for w in ("excess", " xs ", "umbrella", "x p", "x $", "x5m", "x10m")))
                # Count distinct attachment points
                _attach_matches = re.findall(r"(?:xs|x\s|excess\s+of)\s*\$?\s*(\d{1,3})\s*m", _text_lower_tv)
                _distinct_attach = set(_attach_matches)
                _expected_layers = max(_excess_file_count, len(_distinct_attach))
                if _expected_layers > _layer_count and _expected_layers >= 2 and _expected_layers <= 5:
                    logger.warning(f"TOWER MISMATCH: expected {_expected_layers} excess layers but only {_layer_count} extracted ({_layer_keys}). Re-running tower extraction.")
                    _tower_kw = ["excess liability", "excess of", "underlying insurance",
                                 "attachment point", "attaching excess", "xs of",
                                 "following form", "each loss event", "policy aggregate",
                                 "ascot", "markel", "endurance", "arch", "scottsdale",
                                 "partnerre", "summit", "alta", "dual", "apollo"]
                    _tower_text = self._extract_relevant_sections(combined_text, _tower_kw, context_chars=25000)
                    if len(_tower_text) > 500:
                        # Parse excess filenames for attach-point shorthand (e.g., "5x10", "5xP") to help GPT recover missing layers
                        _fn_hint_lines = []
                        for _fn in _file_markers:
                            _fn_lower = _fn.lower()
                            if not any(w in _fn_lower for w in ("excess", " xs ", "umbrella", "following form")):
                                continue
                            _m = re.search(r"(\d+)\s*[xX]\s*(P|p|\d+)", _fn.strip())
                            if _m:
                                _lim = _m.group(1); _att = _m.group(2)
                                if _att.lower() == "p":
                                    _fn_hint_lines.append(f"  FILE {_fn.strip()}: limit ${_lim}M attaches PRIMARY (emit as coverage key 'umbrella')")
                                else:
                                    _an = int(_att)
                                    if _an == 5:
                                        _fn_hint_lines.append(f"  FILE {_fn.strip()}: limit ${_lim}M attaches at $5M (emit as coverage key 'umbrella_layer_2')")
                                    elif _an == 10:
                                        _fn_hint_lines.append(f"  FILE {_fn.strip()}: limit ${_lim}M attaches at $10M (emit as coverage key 'umbrella_layer_3')")
                                    else:
                                        _fn_hint_lines.append(f"  FILE {_fn.strip()}: limit ${_lim}M attaches at ${_an}M")
                            else:
                                _fn_hint_lines.append(f"  FILE {_fn.strip()}: (no shorthand detected)")
                        _filename_hints = ("FILENAME HINTS (AUTHORITATIVE for attach points and coverage keys when body pages are scanned/blank):\n" + "\n".join(_fn_hint_lines) + "\n\n") if _fn_hint_lines else ""
                        
                        _tower_prompt = (
                            f"The main extraction found only {_layer_count} excess/umbrella layer(s) but the document set appears to contain {_expected_layers} distinct excess quotes.\n\n"
                            "Extract ALL excess/umbrella layers from the text below. Return a JSON object with keys \"umbrella\", \"umbrella_layer_2\", \"umbrella_layer_3\" as applicable.\n"
                            "- \"umbrella\" = primary excess (sits above GL, typical attachment $1M)\n"
                            "- \"umbrella_layer_2\" = second excess layer (e.g., $5M xs $5M)\n"
                            "- \"umbrella_layer_3\" = third excess layer (e.g., $5M xs $10M)\n"
                            "For each layer include: carrier, premium, total_premium, limits (Each Occurrence, Aggregate), attachment_point, tower_structure, forms_endorsements, subjectivities.\n"
                            "Do NOT merge separate quotes into one layer. Each quote with its own carrier = one layer.\n\n"
                            f"{_filename_hints}TEXT:\n{_tower_text}"
                        )
                        _tower_response = _get_openai_client().chat.completions.create(
                            model=GPT_MODEL,
                            messages=[
                                {"role": "system", "content": "You are an insurance data extraction specialist. Extract all excess/umbrella tower layers as separate entries. Never merge distinct quotes."},
                                {"role": "user", "content": _tower_prompt},
                            ],
                            response_format={"type": "json_object"},
                            temperature=0.1,
                            max_tokens=8000,
                        )
                        _tower_data = json.loads(_tower_response.choices[0].message.content)
                        for _lk in ("umbrella", "umbrella_layer_2", "umbrella_layer_3"):
                            if _lk in _tower_data and isinstance(_tower_data[_lk], dict):
                                _lay = _tower_data[_lk]
                                if _lay.get("carrier") or _lay.get("premium"):
                                    _existing = data.get("coverages", {}).get(_lk, {})
                                    if not isinstance(_existing, dict) or not _existing.get("carrier"):
                                        data.setdefault("coverages", {})[_lk] = _lay
                                        logger.info(f"TOWER LAYER RECOVERED [{_lk}]: carrier={_lay.get('carrier','N/A')}, premium={_lay.get('premium',0)}")
            except Exception as _e_tv:
                logger.error(f"Tower validation failed: {_e_tv}")


            # ===== MULTI-PASS EXTRACTION =====
            # Pass 2: Focused forms & endorsements extraction for coverages missing them
            data = self._pass2_forms_extraction(data, combined_text)
            
            # Pass 3: Focused address extraction for GL missing designated_premises

            # Pass 2b: GL limits re-extraction if only standard 6 limits found
            _gl_cov = data.get("coverages", {}).get("general_liability", {})
            _gl_limits = _gl_cov.get("limits", []) if isinstance(_gl_cov, dict) else []
            if isinstance(_gl_cov, dict) and _gl_cov.get("carrier") and len(_gl_limits) < 10:
                logger.info(f"Pass 2b: GL has only {len(_gl_limits)} limits, re-extracting additional limits")
                _addl_kw = [
                    "employee benefits", "sexual abuse", "assault and battery",
                    "assault & battery", "hired and non-owned", "hired & non-owned",
                    "hnoa", "hired auto", "abuse or molestation", "each act",
                    "each event", "each claim", "sublimit", "sublimits of liability",
                    "additional limits", "endorsement schedule", "limits of insurance",
                    _gl_cov.get("carrier", "").split()[0] if _gl_cov.get("carrier") else ""
                ]
                _addl_text = self._extract_relevant_sections(combined_text, _addl_kw, context_chars=20000)
                if len(_addl_text) > 500:
                    _existing_descs = {(l.get("description", "") or "").lower() for l in _gl_limits if isinstance(l, dict)}
                    _addl_prompt = f"""The initial extraction found only {len(_gl_limits)} GL limits.
Hotel GL policies typically include additional limits beyond the standard 6 CGL limits.

Look for these ADDITIONAL limits in the text:
- Employee Benefits Liability (Each Claim and Aggregate)
- Sexual Abuse / Abuse & Molestation (Each Act/Occurrence and Aggregate)
- Assault and Battery (Each Event and Aggregate)
- Hired & Non-Owned Auto (HNOA) Liability

Already extracted: {', '.join(sorted(_existing_descs))}

Return JSON: {{"additional_limits": [{{"description": "Name", "limit": "$X"}}]}}
Only include limits NOT already extracted. If none found, return {{"additional_limits": []}}.

TEXT:
{_addl_text}"""
                    try:
                        _addl_resp = _get_openai_client().chat.completions.create(
                            model="gpt-4.1-mini",
                            messages=[
                                {"role": "system", "content": "Extract additional GL limits beyond the standard 6 CGL limits."},
                                {"role": "user", "content": _addl_prompt}
                            ],
                            response_format={"type": "json_object"},
                            temperature=0.0,
                            max_tokens=4000
                        )
                        _addl_data = json.loads(_addl_resp.choices[0].message.content)
                        _addl_lims = _addl_data.get("additional_limits", [])
                        if _addl_lims and isinstance(_addl_lims, list):
                            for al in _addl_lims:
                                if isinstance(al, dict) and al.get("description"):
                                    if al["description"].lower() not in _existing_descs:
                                        _gl_limits.append(al)
                                        _existing_descs.add(al["description"].lower())
                            _gl_cov["limits"] = _gl_limits
                            logger.info(f"Pass 2b: GL now has {len(_gl_limits)} limits")
                        else:
                            logger.info("Pass 2b: No additional GL limits found")
                    except Exception as e:
                        logger.error(f"Pass 2b GL limits re-extraction failed: {e}")
            data = self._pass3_address_extraction(data, combined_text, all_items)
            
            # Pass 4: Focused sublimits extraction for property missing additional_coverages
            data = self._pass4_sublimits_extraction(data, combined_text)

            return data

        except json.JSONDecodeError as e:
            logger.error(f"GPT returned invalid JSON: {e}")
            return {"error": f"Failed to parse extraction results: {e}"}
        except Exception as e:
            logger.error(f"GPT extraction failed: {e}\n{traceback.format_exc()}")
            return {"error": f"AI extraction failed: {e}\n\nTraceback:\n{traceback.format_exc()}"}

    @staticmethod
    def _extract_relevant_sections(combined_text: str, keywords: list, context_chars: int = 8000) -> str:
        """Extract sections of text surrounding keyword matches to reduce token usage.

        Instead of sending the entire document to GPT for focused passes,
        extract windows of text around relevant keywords. Falls back to
        truncated full text if no keywords match.
        """
        text_lower = combined_text.lower()
        # Find all keyword positions
        positions = set()
        for kw in keywords:
            start = 0
            kw_lower = kw.lower()
            while True:
                idx = text_lower.find(kw_lower, start)
                if idx == -1:
                    break
                positions.add(idx)
                start = idx + 1

        if not positions:
            # No keyword matches - fall back to first + last portions of text
            max_len = context_chars * 2
            if len(combined_text) <= max_len:
                return combined_text
            return combined_text[:context_chars] + "\n\n[...]\n\n" + combined_text[-context_chars:]

        # Build windows around each match position, merge overlapping windows
        half = context_chars // 2
        windows = []
        for pos in sorted(positions):
            win_start = max(0, pos - half)
            win_end = min(len(combined_text), pos + half)
            windows.append((win_start, win_end))

        # Merge overlapping windows
        merged = [windows[0]]
        for ws, we in windows[1:]:
            prev_s, prev_e = merged[-1]
            if ws <= prev_e:
                merged[-1] = (prev_s, max(prev_e, we))
            else:
                merged.append((ws, we))

        # Extract and join sections
        sections = []
        for ws, we in merged:
            sections.append(combined_text[ws:we])

        result = "\n\n[...]\n\n".join(sections)
        # Cap at reasonable size
        if len(result) > context_chars * 4:
            result = result[:context_chars * 4]
        return result

    def _pass2_forms_extraction(self, data: dict, combined_text: str) -> dict:
        """Pass 2: Focused extraction of forms & endorsements for coverages that have too few.
        Uses gpt-4.1-mini for speed since this is a focused extraction task."""
        covs = data.get("coverages", {})
        if not isinstance(covs, dict):
            return data
        
        PASS_MODEL = "gpt-4.1-mini"  # Faster model for focused extraction passes
        
        # Check which coverages need forms extraction
        # GL/Property quotes typically have 30-60+ forms - use higher thresholds
        # to ensure Pass 2 re-extracts when initial pass captured only a partial list
        _forms_thresholds = {
            "general_liability": 25,  # GL forms schedules are typically 30-60+ forms
            "property": 25,           # Property typically 20-60+ forms (Starr can have 50+)
            "umbrella": 10,           # Umbrella typically 10-25 forms
            "umbrella_layer_2": 10,
            "umbrella_layer_3": 10,
            "crime": 5,
        }
        needs_forms = []
        for key in ["property", "general_liability", "umbrella", "umbrella_layer_2", "umbrella_layer_3", "crime"]:
            cov = covs.get(key, {})
            if not cov or not cov.get("carrier"):
                continue
            existing_forms = cov.get("forms_endorsements", [])
            form_count = len(existing_forms) if isinstance(existing_forms, list) else 0
            threshold = _forms_thresholds.get(key, 5)
            if form_count < threshold:
                needs_forms.append(key)
                logger.info(f"Pass 2: {key} has only {form_count} forms (threshold={threshold}), will re-extract")
        
        if not needs_forms:
            logger.info("Pass 2 (forms): All coverages have sufficient forms, skipping")
            return data
        
        logger.info(f"Pass 2 (forms): Extracting forms for {needs_forms}")
        
        coverage_display = {
            "property": "Property", "general_liability": "General Liability",
            "umbrella": "Umbrella/Excess Layer 1", "umbrella_layer_2": "Excess Layer 2",
            "umbrella_layer_3": "Excess Layer 3", "crime": "Crime/Fidelity"
        }
        
        for cov_key in needs_forms:
            cov = covs[cov_key]
            carrier = cov.get("carrier", "unknown")
            display = coverage_display.get(cov_key, cov_key)
            
            # Use smart text selection to reduce token usage
            forms_keywords = [
                "forms schedule", "endorsement schedule", "forms list", "schedule of forms",
                "CP 00", "CG 00", "CG 20", "CG 21", "CG 22", "CG 24", "CG 34", "CG 40",
                "IL 00", "IL 01", "IL 09", "IL DS", "NASC", "NXLL", "CSXC",
                "FUT ", "GLF", "EPL", "CYB", "WPA ", "EP100",
                "EXL ", "HS XS", "HS IL", "CX 00", "CX 21",
                "PR 0", "PR 9", "SSPN", "NMA", "LMA", "Starr", "6133",
                "form number", "endorsement", "form name", "edition", "coverage line",
                carrier.split()[0] if carrier else "",
                display.lower()
            ]
            relevant_text = self._extract_relevant_sections(combined_text, forms_keywords, context_chars=25000)

            prompt = f"""Extract EVERY form number and endorsement from this insurance document for the {display} coverage issued by {carrier}.

Rules:
- Extract EVERY form/endorsement number with its full description and edition date
- Format: {{"form_number": "XX 00 00 MM/YY", "description": "Full description"}}
- Include ALL forms from the forms schedule, endorsement schedule, or forms list
- Do NOT skip any forms even if the list is very long
- Include standard forms (e.g., CP 00 10, CG 00 01) AND manuscript/carrier-specific forms
- For NASC/NXLL/CSXC forms, include the full number and description

Return a JSON object with exactly one key:
{{"forms_endorsements": [{{"form_number": "...", "description": "..."}}]}}

DOCUMENT TEXT:
{relevant_text}"""
            
            try:
                response = _get_openai_client().chat.completions.create(
                    model=PASS_MODEL,
                    messages=[
                        {"role": "system", "content": "You are an expert insurance forms extraction assistant. Extract every form number and endorsement exactly as written."},
                        {"role": "user", "content": prompt}
                    ],
                    response_format={"type": "json_object"},
                    temperature=0.0,
                    max_tokens=16000
                )
                result = json.loads(response.choices[0].message.content)
                forms = result.get("forms_endorsements", [])
                # ---- FORMS PREFIX VALIDATION ----
                # Property uses CP/TC/VR/EC/EB/CPF/CFP prefixes (ISO + Vantage Risk/Tower Hill)
                # GL uses CG/GLF/AD/AI/DE/JA/NXLL/NASC; Liquor uses LL; Auto uses CA
                # Umbrella uses CSXC/EXL/HS XS/FUT/XS/CX
                _prop_pfx = ("CP ", "CPF", "CFP", "TC ", "VR ", "EC ", "EB-", "EB0", "MS PR", "MS DEC", "MS EBC", "HSIC SP", "HSIC SOS", "MS GEN")
                _gl_pfx = ("CG ", "AD ", "AI ", "DE ", "JA ", "NXLL", "NASC", "GLF", "GL ")
                _umb_pfx = ("CSXC", "EXL ", "HS XS", "FUT ", "XS ", "NXLL", "CX ")
                _liq_pfx = ("LL ", "LL-", "LL FLIQL")
                _auto_pfx = ("CA ", "CA-")
                _gl_pkg_pfx = ("EPL", "CYB", "WPA", "EP1", "FLSL", "SSIC", "FUT-SS", "FUT SS", "GL STATE")
                if forms and cov_key in ("general_liability", "umbrella", "umbrella_layer_2", "umbrella_layer_3"):
                    prop_ct = sum(1 for f in forms if isinstance(f, dict) and
                                 str(f.get("form_number", "")).upper().startswith(_prop_pfx))
                    rel_ct = 0
                    if cov_key == "general_liability":
                        rel_ct = sum(1 for f in forms if isinstance(f, dict) and
                                    str(f.get("form_number", "")).upper().startswith(_gl_pfx))
                    else:
                        rel_ct = sum(1 for f in forms if isinstance(f, dict) and
                                    str(f.get("form_number", "")).upper().startswith(_umb_pfx + _gl_pfx))
                    if prop_ct > 5 and rel_ct < 3:
                        logger.warning(f"Pass 2: REJECTED {len(forms)} forms for {cov_key} - "
                                      f"{prop_ct} property prefixes vs {rel_ct} relevant. "
                                      f"These are property forms incorrectly extracted for {cov_key}.")
                        forms = []
                    elif prop_ct > 0 and rel_ct > 0:
                        filtered = [f for f in forms if isinstance(f, dict) and
                                   not str(f.get("form_number", "")).upper().startswith(_prop_pfx)]
                        logger.info(f"Pass 2: Filtered {len(forms) - len(filtered)} property forms from {cov_key}, keeping {len(filtered)}")
                        forms = filtered
                # Inverse: reject GL/liquor/auto/umbrella forms leaking into property coverage
                if forms and cov_key in ("property", "excess_property", "excess_property_2"):
                    _nonprop = _gl_pfx + _umb_pfx + _liq_pfx + _auto_pfx + _gl_pkg_pfx
                    nonprop_ct = sum(1 for f in forms if isinstance(f, dict) and
                                    str(f.get("form_number", "")).upper().startswith(_nonprop))
                    prop_ct2 = sum(1 for f in forms if isinstance(f, dict) and
                                  str(f.get("form_number", "")).upper().startswith(_prop_pfx))
                    if nonprop_ct > 5 and prop_ct2 < 3:
                        logger.warning(f"Pass 2: REJECTED {len(forms)} forms for {cov_key} - "
                                      f"{nonprop_ct} non-property prefixes vs {prop_ct2} property. "
                                      f"These are GL/liquor/auto forms incorrectly extracted for property.")
                        forms = []
                    elif nonprop_ct > 0:
                        filtered = [f for f in forms if isinstance(f, dict) and
                                   not str(f.get("form_number", "")).upper().startswith(_nonprop)]
                        logger.info(f"Pass 2: Filtered {len(forms) - len(filtered)} non-property forms from {cov_key}, keeping {len(filtered)}")
                        forms = filtered
                existing_count = len(cov.get("forms_endorsements", []) or [])
                if forms and isinstance(forms, list) and len(forms) > existing_count:
                    cov["forms_endorsements"] = forms
                    logger.info(f"Pass 2: Extracted {len(forms)} forms for {cov_key} (was {existing_count})")
                elif forms and isinstance(forms, list) and len(forms) > 0:
                    logger.info(f"Pass 2: Found {len(forms)} forms for {cov_key} but not better than existing {existing_count}")
                else:
                    logger.warning(f"Pass 2: No forms found for {cov_key} in focused extraction")
            except Exception as e:
                logger.error(f"Pass 2 forms extraction failed for {cov_key}: {e}")
        
        return data

    def _pass3_address_extraction(self, data: dict, combined_text: str, all_items: list = None) -> dict:
        """Pass 3: Focused extraction of covered addresses for GL.
        Uses gpt-4.1-mini for speed since this is a focused extraction task."""
        PASS_MODEL = "gpt-4.1"      
        covs = data.get("coverages", {})
        gl = covs.get("general_liability", {})
        
        if not gl or not gl.get("carrier"):
            logger.info("Pass 3 (addresses): No GL coverage found, skipping")
            return data
        
        # Check if GL data is complete - trigger re-extraction when:
        # 1. designated_premises < 3 (GPT often captures only 1-2 in initial pass)
        # 2. schedule_of_classes count is much less than SOV location count (truncated table)
        dp = gl.get("designated_premises", [])
        dp_count = len(dp) if isinstance(dp, list) else 0
        soc = gl.get("schedule_of_classes", [])
        soc_count = len(soc) if isinstance(soc, list) else 0

        # Get SOV location count for comparison
        sov_data = data.get("sov_data")
        sov_count = len(sov_data.get("locations", [])) if sov_data and isinstance(sov_data, dict) else 0

        # Determine if re-extraction is needed
        dp_seems_complete = dp_count >= max(sov_count, 3)
        soc_seems_complete = soc_count >= max(sov_count, 5)

        if dp_seems_complete and soc_seems_complete:
            logger.info(f"Pass 3 (addresses): GL has {dp_count} premises and {soc_count} classes "
                       f"(SOV has {sov_count} locations), both seem complete - skipping")
            return data

        logger.info(f"Pass 3 (addresses): GL has {dp_count} premises and {soc_count} classes "
                   f"(SOV has {sov_count} locations) - running focused re-extraction")

        # Use smart text selection for address-related content
        # Use larger context window to capture full Schedule of Classes tables
        address_keywords = [
            "CG 21 44", "NXLL 110", "designated premises", "schedule of hazards",
            "schedule of locations", "covered premises", "insured locations",
            "location address", "schedule of classes", "Hotels/Motels",
            "class code", "exposure basis", "Gross Sales", "FUT 1004", "FUT 1005",
            "location#", "Primary", "45191", "16910", "58173", "Gross Sales", "named insured", "FUT 1007"
        ]
        # Use the FULL text of the GL file (not the truncated combined_text)
        # The GL file in combined_text may be heavily truncated by budget allocation
        gl_source_text = combined_text  # fallback
        if all_items:
            gl_keywords = ["gl", "general liability", "commercial general"]
            for item in all_items:
                fn_lower = item.get("filename", "").lower()
                if any(kw in fn_lower for kw in gl_keywords):
                    full_gl_text = item.get("header", "") + item.get("text", "")
                    if len(full_gl_text) > len(gl_source_text) * 0.5:
                        gl_source_text = full_gl_text
                        gl_fn = item.get("filename", "unknown")
                        logger.info(f"Pass 3: Using full GL file text ({len(full_gl_text)} chars) "
                                   f"from {gl_fn!r} instead of truncated combined_text")
                        break
        relevant_text = self._extract_relevant_sections(gl_source_text, address_keywords, context_chars=40000)

        prompt = f"""From this General Liability insurance document, extract TWO things:

1. ALL physical street addresses that represent covered locations (designated_premises)
2. The COMPLETE Schedule of Classes table (schedule_of_classes) - EVERY row, EVERY location

Look for addresses in:
- CG 21 44 or NXLL 110 (Limitation of Coverage to Designated Premises) form
- Schedule of Hazards / Schedule of Locations / FUT 1005
- Any numbered list of addresses (e.g., "1) 4285 Highway 51, LaPlace, LA 70068")
- Any section listing covered premises, designated locations, or insured locations
- The declarations page showing location addresses

For Schedule of Classes (FUT 1004 or similar form):
- Extract EVERY row - there may be 10-15+ rows spanning multiple locations
- Each row has: Location (e.g., "Primary", "location#3", "location#8"), Class Code (e.g., 45191), Description, Exposure Amount, Rate, Premium
- Include ALL locations: Primary, location#3, location#4, ..., location#12, etc.
- Include restaurant/liquor entries (class codes 16910, 58173) as separate rows
- CRITICAL: Do NOT stop after 3-4 rows. Extract the ENTIRE table.
- The exposure amount is the dollar figure (e.g., $3,200,000) - this represents Gross Sales for that location

Rules:
- Extract the COMPLETE street address including street number, street name, city, state, and zip
- Include ALL addresses, even if they span multiple pages
- Do NOT include PO Boxes or mailing addresses - only physical location addresses
- Each address should be a separate entry

Return a JSON object:
{{"designated_premises": ["Full address 1", "Full address 2", ...],
  "schedule_of_classes": [{{"location": "Primary", "address": "Street, City, ST Zip", "brand_dba": "Hotel name", "classification": "Hotels and Motels - with pools", "class_code": "45191", "exposure_basis": "Gross Sales", "exposure": "$3,200,000", "rate": "5.8653", "premium": "$18,769"}}]}}

DOCUMENT TEXT:
{relevant_text}"""
        
        try:
            response = _get_openai_client().chat.completions.create(
                model=PASS_MODEL,
                messages=[
                    {"role": "system", "content": "You are an expert at extracting location data from insurance documents. Extract every covered location address AND every row from the Schedule of Classes table. Do NOT truncate - include ALL rows."},
                    {"role": "user", "content": prompt}
                ],
                response_format={"type": "json_object"},
                temperature=0.0,
                max_tokens=16000
            )
            result = json.loads(response.choices[0].message.content)
            
            addresses = result.get("designated_premises", [])
            if addresses and isinstance(addresses, list) and len(addresses) > dp_count:
                gl["designated_premises"] = addresses
                logger.info(f"Pass 3: Extracted {len(addresses)} designated premises for GL (was {dp_count})")
            elif addresses and isinstance(addresses, list) and len(addresses) > 0:
                logger.info(f"Pass 3: Found {len(addresses)} addresses but not better than existing {dp_count}")
            
            # Also update schedule_of_classes if it was empty or incomplete
            soc = result.get("schedule_of_classes", [])
            existing_soc = gl.get("schedule_of_classes", [])
            if soc and isinstance(soc, list) and len(soc) >= len(existing_soc):
                # Use Pass 3 result - it used focused prompt with larger context window
                gl["schedule_of_classes"] = soc
                logger.info(f"Pass 3: Updated schedule_of_classes to {len(soc)} entries (was {len(existing_soc)})")
            return data
        except Exception as e:
            logger.warning(f"Pass 3 address extraction failed: {e}")
            return data

    def _pass4_sublimits_extraction(self, data: dict, combined_text: str) -> dict:
        """Pass 4: Focused extraction of property sublimits/extensions.
        Uses gpt-4.1-mini for speed since this is a focused extraction task."""
        PASS_MODEL = "gpt-4.1-mini"
        covs = data.get("coverages", {})
        prop = covs.get("property", {})
        
        if not prop or not prop.get("carrier"):
            logger.info("Pass 4 (sublimits): No property coverage found, skipping")
            return data
        
        # Check if additional_coverages is already populated
        ac = prop.get("additional_coverages", [])
        if ac and isinstance(ac, list) and len(ac) >= 5:
            logger.info(f"Pass 4 (sublimits): Property already has {len(ac)} sublimits, skipping")
            return data
        
        logger.info(f"Pass 4 (sublimits): Property has only {len(ac) if ac else 0} sublimits, running focused extraction")
        
        # Use smart text selection for sublimit-related content
        sublimit_keywords = [
            "sublimit", "extension of coverage", "additional coverage",
            "coverage extension", "supplemental", "flood", "earthquake",
            "equipment breakdown", "ordinance or law", "spoilage",
            "business income", "debris removal", "pollutant cleanup",
            "utility services", "sewer", "drain backup", "mold", "fungi",
            "ingress", "egress", "contingent business",
            "additional coverages included", "policy coverages", "valuable papers", "accounts receivable", "water backup", "outdoor signs", "fire department service charge",
        ]
        relevant_text = self._extract_relevant_sections(combined_text, sublimit_keywords, context_chars=10000)

        prompt = f"""From this Property insurance document, extract ALL sublimits of liability, extensions of coverage, and additional coverages.

Look for sections titled:
- Sublimits of Liability
- Extensions of Coverage
- Additional Coverages
- Coverage Extensions
- Supplemental Coverages
- Additional Coverages Included (Tower Hill / Vantage Risk quotes use this exact section name listing form-included coverages)
- Policy Coverages (Tower Hill per-premise sub-table listing Ordinance or Law A, Ordinance or Law B/C, Valuable Papers, Business Income with Extra Expense Coverage, Accounts Receivable, Debris Removal, Pollutant Removal and Clean Up, Water Backup)
- Any table or list showing coverage descriptions with dollar limits

Common property sublimits to look for:
- Flood (per occurrence and aggregate)
- Earthquake
- Equipment Breakdown
- Ordinance or Law (Coverage A, B, C)
- Spoilage
- Business Income Extended Period (days)
- Sign Coverage
- Accounts Receivable
- Valuable Papers
- Fine Arts
- Newly Acquired Property
- Transit
- Debris Removal
- Pollutant Cleanup
- Utility Services (Direct Damage and Time Element)
- Green Building
- Sewer/Drain Backup
- Water Damage
- Mold/Fungi
- Electronic Data
- Brands and Labels
- Civil/Military Authority
- Ingress/Egress
- Service Interruption
- Contingent Business Income

Rules:
- Extract EVERY sublimit with its dollar amount or status (Included, Excluded, NOT COVERED)
- Include deductibles for each sublimit if shown
- If a sublimit has different per-occurrence and aggregate limits, include both
- Mark excluded coverages as "Excluded" or "NOT COVERED"
- Include the ACTUAL dollar amounts, not just "Included"

Return a JSON object:
{{"additional_coverages": [{{"description": "Coverage name", "limit": "$X or Excluded", "deductible": "$X or N/A"}}]}}

DOCUMENT TEXT:
{relevant_text}"""
        
        try:
            response = _get_openai_client().chat.completions.create(
                model=PASS_MODEL,
                messages=[
                    {"role": "system", "content": "You are an expert at extracting property insurance sublimits and extensions of coverage. Extract every sublimit with its exact dollar amount."},
                    {"role": "user", "content": prompt}
                ],
                response_format={"type": "json_object"},
                temperature=0.0,
                max_tokens=8000
            )
            result = json.loads(response.choices[0].message.content)
            
            sublimits = result.get("additional_coverages", [])
            if sublimits and isinstance(sublimits, list) and len(sublimits) > len(ac or []):
                prop["additional_coverages"] = sublimits
                logger.info(f"Pass 4: Extracted {len(sublimits)} sublimits for property (was {len(ac or [])}")
            else:
                logger.info(f"Pass 4: Focused extraction found {len(sublimits) if sublimits else 0} sublimits (not better than existing {len(ac or [])})")
            
        except Exception as e:
            logger.error(f"Pass 4 sublimits extraction failed: {e}")
        
        return data

    def apply_adjustments(self, data: dict, instructions: str) -> dict:
        """
        Apply user corrections/adjustments to the extracted data using GPT.
        This is a synchronous wrapper; the caller should use asyncio.to_thread.
        """
        import asyncio

        loop = asyncio.new_event_loop()
        try:
            return loop.run_until_complete(apply_corrections(data, instructions))
        finally:
            loop.close()
