"""
SOV (Statement of Values) Parser for Hotel Insurance Proposals.

Parses .xlsx SOV spreadsheets (Amwins, Starr, and similar formats) to extract
per-location property details for use in proposal generation.
"""

import logging
import re
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)

# Common header keywords used to detect the header row in SOV spreadsheets.
# We look for rows containing several of these terms.
HEADER_KEYWORDS = {
    "address", "city", "state", "zip", "building value", "contents value",
    "tiv", "rooms", "location", "construction", "year built", "hotel flag",
    "dba", "occupancy", "square footage", "sprinkler", "flood zone",
    "roof", "county", "stories"
}

# Mapping from common SOV column header variants to our normalized field names.
COLUMN_MAP = {
    # Location identifiers
    "client name": "client_name",
    "corporate name": "corporate_name",
    "corporate name (llc)": "corporate_name",
    "dba": "dba",
    "hotel flag": "hotel_flag",
    "brand": "hotel_flag",
    "location name": "dba",
    "location #": "location_num",
    "location number": "location_num",
    "loc #": "location_num",
    "building #": "building_num",
    "building number": "building_num",
    "bldg #": "building_num",
    "* bldg no.": "building_num",
    "*bldg no.": "building_num",

    # Address fields
    "address": "address",
    "street address": "address",
    "street": "address",
    "city": "city",
    "state": "state",
    "zip": "zip_code",
    "zip code": "zip_code",
    "zipcode": "zip_code",
    "county": "county",

    # Dates
    "effective date": "effective_date",
    "expiration date": "expiration_date",

    # Property values
    "building value": "building_value",
    "* building value": "building_value",
    "*building value": "building_value",
    "bldg value": "building_value",
    "real property value": "building_value",
    "*real property value": "building_value",
    "* real property value": "building_value",
    "*real property value ($)": "building_value",
    "contents value": "contents_value",
    "*contents value": "contents_value",
    "* contents value": "contents_value",
    "contents": "contents_value",
    "personal property value": "contents_value",
    "personal property value ($)": "contents_value",
    "business income/rents": "bi_value",
    "buisness income/rents": "bi_value",  # common typo in templates
    "*buisness income/rents": "bi_value",
    "* buisness income/rents": "bi_value",
    "*business income/rents": "bi_value",
    "* business income/rents": "bi_value",
    "bi/rental income": "bi_value",
    "bi/rental income ($)": "bi_value",
    "bi value": "bi_value",
    "bi": "bi_value",
    "pool value": "pool_value",
    "pools": "pool_value",
    "sign value": "sign_value",
    "signs": "sign_value",
    "other values": "other_value",
    "other": "other_value",
    "tiv": "tiv",
    "total insured value": "tiv",
    "*total tiv": "tiv",

    # Construction details
    "occupancy description": "occupancy",
    "occupancy": "occupancy",
    "construction code (iso)*": "construction_code",
    "construction code": "construction_code",
    "*iso const": "construction_code",
    "iso const": "construction_code",
    "iso construction": "construction_code",
    "construction type": "construction_type",
    "construction": "construction_type",
    "no. of stories": "stories",
    "stories": "stories",
    "number of stories": "stories",
    "year built": "year_built",
    "square footage": "square_footage",
    "sq ft": "square_footage",
    "sqft": "square_footage",

    # Systems/updates
    "year electrical updated": "year_electrical",
    "year plumbing updated": "year_plumbing",
    "year hvac updated": "year_hvac",
    "roof full replacement year": "roof_year",
    "roof year": "roof_year",
    "roof type": "roof_type",

    # Occupancy and protection
    "occupancy %": "occupancy_pct",
    "sprinklered %": "sprinkler_pct",
    "sprinklered": "sprinkler_pct",
    "sprinkler": "sprinkler_pct",

    # Valuation
    "valuation": "valuation_per_sqft",
    "# of rooms": "num_rooms",
    "rooms": "num_rooms",
    "number of rooms": "num_rooms",
    "*# of units": "num_rooms",
    "# of units": "num_rooms",
    "units": "num_rooms",

    # Other
    "eifs?": "eifs",
    "eifs": "eifs",
    "flood coverage required?": "flood_required",
    "flood limit requested": "flood_limit",
    "flood zone": "flood_zone",
    "flood zone (select zone)": "flood_zone",
    "aop deductible": "aop_deductible",
    "aop deductible (choose one)": "aop_deductible",
    "earthquake required?": "earthquake_required",
    "earthquake limit requested": "earthquake_limit",
    "any stoves in rooms?": "stoves",
    "fireplace in rooms?": "fireplace",
    "wiring type": "wiring_type",
    "total losses last 5 years ($)": "total_losses_5yr",
    "loss details (if any)": "loss_details",
}


def _normalize_header(text: str) -> str:
    """Normalize a header string for matching."""
    if not text:
        return ""
    return re.sub(r'\s+', ' ', str(text).strip().lower())


def _find_header_row(ws) -> Optional[int]:
    """
    Scan the worksheet to find the header row by looking for rows
    that contain multiple known SOV column keywords.
    """
    for row_idx in range(1, min(ws.max_row + 1, 30)):  # Check first 30 rows
        row_values = []
        for cell in ws[row_idx]:
            if cell.value:
                row_values.append(_normalize_header(str(cell.value)))

        # Count how many known keywords appear in this row
        matches = 0
        for val in row_values:
            for keyword in HEADER_KEYWORDS:
                if keyword in val:
                    matches += 1
                    break

        # If we find 5+ keyword matches, this is likely the header row
        if matches >= 5:
            logger.info(f"SOV header row found at row {row_idx} with {matches} keyword matches")
            return row_idx

    return None


def _extract_summary(ws, header_row: int) -> dict:
    """
    Extract summary information from the rows above the header row.
    Looks for: First Named Insured, Mailing Address, Effective Date, Total TIV.
    """
    summary = {}
    for row_idx in range(1, header_row):
        for cell in ws[row_idx]:
            val = str(cell.value).strip().lower() if cell.value else ""
            if "first named insured" in val:
                # The value is typically in the next column
                next_cell = ws.cell(row=row_idx, column=cell.column + 1)
                if next_cell.value:
                    summary["named_insured"] = str(next_cell.value).strip()
            elif "mailing address" in val:
                next_cell = ws.cell(row=row_idx, column=cell.column + 1)
                if next_cell.value:
                    summary["mailing_address"] = str(next_cell.value).strip()
                # Check next row for city/state/zip continuation
                below_cell = ws.cell(row=row_idx + 1, column=cell.column + 1)
                if below_cell.value:
                    summary["mailing_address"] += ", " + str(below_cell.value).strip()
            elif "desired effective date" in val or "effective date" in val:
                next_cell = ws.cell(row=row_idx, column=cell.column + 1)
                if next_cell.value:
                    summary["effective_date"] = str(next_cell.value).strip()
            elif "total tiv" in val:
                next_cell = ws.cell(row=row_idx, column=cell.column + 1)
                if next_cell.value:
                    try:
                        summary["total_tiv"] = float(str(next_cell.value).replace(",", "").replace("$", ""))
                    except (ValueError, TypeError):
                        summary["total_tiv"] = str(next_cell.value)

    return summary


def _map_columns(ws, header_row: int) -> dict:
    """
    Map column indices to normalized field names using the header row.
    Returns {column_index: field_name}.
    """
    col_map = {}
    # Track which fields have already been mapped to prevent duplicates
    # (first column wins for each field name)
    mapped_fields = set()
    
    for cell in ws[header_row]:
        if cell.value:
            raw = str(cell.value)
            normalized = _normalize_header(raw)
            # Strip leading asterisks and clean for matching
            cleaned = re.sub(r'^\*+\s*', '', normalized).strip()
            # Also strip trailing ($) or similar
            cleaned_no_dollar = re.sub(r'\s*\(\$\)\s*$', '', cleaned).strip()
            
            matched_field = None
            
            # Try exact match on normalized, cleaned, and cleaned_no_dollar
            for variant in [normalized, cleaned, cleaned_no_dollar]:
                if variant in COLUMN_MAP:
                    matched_field = COLUMN_MAP[variant]
                    break
            
            # Try partial match only if no exact match found
            # Use word-boundary-aware matching to avoid false positives
            # (e.g., "addressnum" should NOT match "address")
            if not matched_field:
                for header_variant, field_name in COLUMN_MAP.items():
                    # Skip very short variants for partial matching
                    if len(header_variant) < 4:
                        continue
                    # Check if variant appears as a word boundary match in the header
                    pattern = r'(?:^|\b|\s)' + re.escape(header_variant) + r'(?:$|\b|\s)'
                    if re.search(pattern, normalized) or re.search(pattern, cleaned):
                        matched_field = field_name
                        break
            
            # Only map if field hasn't been mapped yet (first column wins)
            if matched_field and matched_field not in mapped_fields:
                col_map[cell.column] = matched_field
                mapped_fields.add(matched_field)

    logger.info(f"Mapped {len(col_map)} columns: {list(col_map.values())}")
    return col_map


def _safe_float(val) -> float:
    """Safely convert a value to float, returning 0.0 on failure."""
    if val is None:
        return 0.0
    try:
        cleaned = str(val).replace(",", "").replace("$", "").strip()
        if not cleaned or cleaned.lower() in ("n/a", "na", "none", "-", "0"):
            return 0.0
        return float(cleaned)
    except (ValueError, TypeError):
        return 0.0


def _safe_int(val) -> int:
    """Safely convert a value to int, returning 0 on failure."""
    return int(_safe_float(val))


def _safe_str(val) -> str:
    """Safely convert a value to string."""
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ("none", "0", "0.0"):
        return ""
    return s


def parse_sov(file_path: str) -> dict:
    """
    Parse an SOV spreadsheet and return structured location data.

    Returns:
        {
            "source": "sov",
            "summary": {
                "named_insured": str,
                "mailing_address": str,
                "effective_date": str,
                "total_tiv": float,
                "num_locations": int
            },
            "locations": [
                {
                    "location_num": int,
                    "client_name": str,
                    "corporate_name": str,
                    "dba": str,
                    "hotel_flag": str,
                    "address": str,
                    "city": str,
                    "state": str,
                    "zip_code": str,
                    "county": str,
                    "num_rooms": int,
                    "building_value": float,
                    "contents_value": float,
                    "bi_value": float,
                    "pool_value": float,
                    "sign_value": float,
                    "other_value": float,
                    "tiv": float,
                    "construction_type": str,
                    "construction_code": str,
                    "stories": int,
                    "year_built": int,
                    "square_footage": int,
                    "sprinkler_pct": str,
                    "roof_type": str,
                    "roof_year": int,
                    "flood_zone": str,
                    "aop_deductible": float,
                    "occupancy": str,
                    ...
                },
                ...
            ],
            "totals": {
                "building_value": float,
                "contents_value": float,
                "bi_value": float,
                "tiv": float,
                "num_rooms": int,
                "num_locations": int
            }
        }
    """
    logger.info(f"Parsing SOV file: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Find the header row
    header_row = _find_header_row(ws)
    if header_row is None:
        logger.warning("Could not find SOV header row")
        return {"error": "Could not detect SOV format. No header row found with expected column names."}

    # Extract summary from rows above header
    summary = _extract_summary(ws, header_row)

    # Map columns
    col_map = _map_columns(ws, header_row)
    if not col_map:
        return {"error": "Could not map any columns from the SOV header row."}

    # Parse data rows
    locations = []
    value_fields = {"building_value", "contents_value", "bi_value", "pool_value",
                    "sign_value", "other_value", "tiv", "aop_deductible",
                    "valuation_per_sqft", "flood_limit", "earthquake_limit",
                    "total_losses_5yr"}
    int_fields = {"location_num", "building_num", "stories", "year_built",
                  "square_footage", "num_rooms", "year_electrical", "year_plumbing",
                  "year_hvac", "roof_year"}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        # Check if row has data (at least an address or client name)
        row_has_data = False
        for col_idx, field_name in col_map.items():
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None and str(cell_val).strip():
                if field_name in ("address", "client_name", "dba", "city"):
                    row_has_data = True
                    break

        if not row_has_data:
            continue

        location = {}
        for col_idx, field_name in col_map.items():
            cell_val = ws.cell(row=row_idx, column=col_idx).value

            if field_name in value_fields:
                location[field_name] = _safe_float(cell_val)
            elif field_name in int_fields:
                location[field_name] = _safe_int(cell_val)
            elif field_name == "effective_date":
                if cell_val:
                    if hasattr(cell_val, 'strftime'):
                        location[field_name] = cell_val.strftime("%m/%d/%Y")
                    else:
                        location[field_name] = _safe_str(cell_val)
            elif field_name == "sprinkler_pct":
                val = _safe_float(cell_val)
                if val >= 1:
                    location[field_name] = "Yes (100%)"
                elif val > 0:
                    location[field_name] = f"Partial ({val*100:.0f}%)"
                else:
                    location[field_name] = "No"
            else:
                location[field_name] = _safe_str(cell_val)

        # Calculate TIV if not provided
        if location.get("tiv", 0) == 0:
            location["tiv"] = (
                location.get("building_value", 0) +
                location.get("contents_value", 0) +
                location.get("bi_value", 0) +
                location.get("pool_value", 0) +
                location.get("sign_value", 0) +
                location.get("other_value", 0)
            )

        # Set location number if not provided
        if location.get("location_num", 0) == 0:
            location["location_num"] = len(locations) + 1

        locations.append(location)

    # Calculate totals
    totals = {
        "building_value": sum(loc.get("building_value", 0) for loc in locations),
        "contents_value": sum(loc.get("contents_value", 0) for loc in locations),
        "bi_value": sum(loc.get("bi_value", 0) for loc in locations),
        "tiv": sum(loc.get("tiv", 0) for loc in locations),
        "num_rooms": sum(loc.get("num_rooms", 0) for loc in locations),
        "num_locations": len(locations),
    }

    # Update summary
    summary["num_locations"] = len(locations)
    if "total_tiv" not in summary:
        summary["total_tiv"] = totals["tiv"]

    result = {
        "source": "sov",
        "summary": summary,
        "locations": locations,
        "totals": totals,
    }

    logger.info(f"SOV parsed: {len(locations)} locations, Total TIV: ${totals['tiv']:,.0f}")
    return result


def aggregate_locations(sov_data: dict) -> dict:
    """
    Aggregate building-level SOV rows into location-level summaries.
    Many SOVs list multiple buildings per location (e.g., Days Inn has 4 buildings).
    This groups them by building_num (location identifier) and sums values.
    Returns a new sov_data dict with aggregated locations and updated totals.
    """
    if "error" in sov_data:
        return sov_data
    
    raw_locations = sov_data.get("locations", [])
    if not raw_locations:
        return sov_data
    
    from collections import OrderedDict
    agg = OrderedDict()
    
    for loc in raw_locations:
        # Use building_num as the location grouping key (this is the location number in AmRisc SOVs)
        key = loc.get("building_num", loc.get("location_num", 0))
        if key not in agg:
            agg[key] = {
                "location_num": key,
                "dba": loc.get("dba", ""),
                "hotel_flag": loc.get("hotel_flag", ""),
                "corporate_name": loc.get("corporate_name", ""),
                "client_name": loc.get("client_name", ""),
                "address": loc.get("address", ""),
                "city": loc.get("city", ""),
                "state": loc.get("state", ""),
                "zip_code": loc.get("zip_code", ""),
                "county": loc.get("county", ""),
                "construction_type": loc.get("construction_type", ""),
                "construction_code": loc.get("construction_code", ""),
                "year_built": loc.get("year_built", 0),
                "stories": loc.get("stories", 0),
                "num_rooms": 0,
                "building_value": 0.0,
                "contents_value": 0.0,
                "bi_value": 0.0,
                "other_value": 0.0,
                "pool_value": 0.0,
                "sign_value": 0.0,
                "tiv": 0.0,
                "square_footage": 0,
                "occupancy": loc.get("occupancy", ""),
                "sprinkler_pct": loc.get("sprinkler_pct", ""),
                "flood_zone": loc.get("flood_zone", ""),
                "_building_count": 0,
            }
        entry = agg[key]
        entry["num_rooms"] += loc.get("num_rooms", 0)
        entry["building_value"] += loc.get("building_value", 0)
        entry["contents_value"] += loc.get("contents_value", 0)
        entry["bi_value"] += loc.get("bi_value", 0)
        entry["other_value"] += loc.get("other_value", 0)
        entry["pool_value"] += loc.get("pool_value", 0)
        entry["sign_value"] += loc.get("sign_value", 0)
        entry["tiv"] += loc.get("tiv", 0)
        entry["square_footage"] += loc.get("square_footage", 0)
        entry["_building_count"] += 1
        # Use highest stories value
        if loc.get("stories", 0) > entry["stories"]:
            entry["stories"] = loc["stories"]
    
    aggregated = list(agg.values())
    
    # Recalculate totals
    totals = {
        "building_value": sum(loc.get("building_value", 0) for loc in aggregated),
        "contents_value": sum(loc.get("contents_value", 0) for loc in aggregated),
        "bi_value": sum(loc.get("bi_value", 0) for loc in aggregated),
        "tiv": sum(loc.get("tiv", 0) for loc in aggregated),
        "num_rooms": sum(loc.get("num_rooms", 0) for loc in aggregated),
        "num_locations": len(aggregated),
    }
    
    result = dict(sov_data)
    result["locations"] = aggregated
    result["locations_raw"] = raw_locations  # Keep raw building-level data
    result["totals"] = totals
    result["summary"] = dict(sov_data.get("summary", {}))
    result["summary"]["num_locations"] = len(aggregated)
    
    logger.info(f"Aggregated {len(raw_locations)} buildings into {len(aggregated)} locations")
    return result


def is_sov_file(file_path: str) -> bool:
    """
    Quick check to determine if an .xlsx file looks like an SOV spreadsheet.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active
        header_row = _find_header_row(ws)
        wb.close()
        return header_row is not None
    except Exception as e:
        logger.warning(f"Error checking SOV file: {e}")
        return False


def format_sov_summary(sov_data: dict) -> str:
    """
    Format SOV data as a human-readable summary for Telegram display.
    """
    if "error" in sov_data:
        return f"SOV Error: {sov_data['error']}"

    summary = sov_data.get("summary", {})
    totals = sov_data.get("totals", {})
    locations = sov_data.get("locations", [])

    lines = []
    lines.append("📋 Statement of Values Parsed")
    lines.append("")

    if summary.get("named_insured"):
        lines.append(f"Named Insured: {summary['named_insured']}")

    lines.append(f"Locations: {len(locations)}")
    lines.append(f"Total Rooms: {totals.get('num_rooms', 0):,}")
    lines.append(f"Total TIV: ${totals.get('tiv', 0):,.0f}")
    lines.append(f"  Building: ${totals.get('building_value', 0):,.0f}")
    lines.append(f"  Contents: ${totals.get('contents_value', 0):,.0f}")
    lines.append(f"  BI/Rents: ${totals.get('bi_value', 0):,.0f}")
    lines.append("")

    for loc in locations:
        loc_num = loc.get("location_num", "?")
        dba = loc.get("dba", "") or loc.get("hotel_flag", "")
        addr = loc.get("address", "")
        city = loc.get("city", "")
        state = loc.get("state", "")
        rooms = loc.get("num_rooms", 0)
        tiv = loc.get("tiv", 0)

        loc_line = f"  #{loc_num}"
        if dba:
            loc_line += f" {dba}"
        loc_line += f" — {addr}, {city}, {state}"
        if rooms:
            loc_line += f" ({rooms} rooms)"
        loc_line += f" — TIV: ${tiv:,.0f}"
        lines.append(loc_line)

    return "\n".join(lines)
